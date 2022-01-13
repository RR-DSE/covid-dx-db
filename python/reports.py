# reports.py (utf-8)
# 
# Edited by: RR-DSE
# Timestamp: 22-01-13 16:46:17

# ------------
# Dependencies
# ------------

import fpdf
import subprocess
import mysql.connector
import re
import datetime
import sys
import io
import os
import xlrd
import xlwt
import email
import email.encoders
import email.mime.base
import email.mime.multipart
import email.mime.text
import smtplib
import ssl
import xml.etree.ElementTree as ElementTree
from copy import copy
import unicodedata
from openpyxl import Workbook, load_workbook
import listools

# ---------
# Constants
# ---------

dCounts = None

sLogFileTimestampFormat = "%Y_%m_%d_%H_%M_%S"
sLogFolder = "logs"
sConfigFolder = "config"
sConfigXML = "config.xml"

sCode = "utf-8"
sDateFormat = "%Y-%m-%d"
sDateTimeFormat = "%Y-%m-%d %H:%M:%S"
sTimestampFormat = "%Y-%m-%d %H:%M:%S"

dEmailStockReportInputs = None

dConfigReport = {
  "page_unit":  "mm",
  "page_format": "A4",
  "page_orientation": "P",
  "page_margin_left": 10,
  "page_margin_right": 10,
  "page_margin_top": 10,
  "page_margin_bottom": 10,
  "page_body_width": 188,
  "page_body_height": 277,
  "page_char_advance": 1.30,
  "page_header_sep": 2,
  "page_section_advance": 5,
  "page_section_sep": 2,
  "page_subsection_sep": 1.5,
  "font_family_header_1": "helvetica",
  "font_style_header_1": "",
  "font_size_header_1": 8,
  "font_height_header_1": 4,
  "font_sep_header_1": 1,
  "font_color_r_header_1": 100,
  "font_color_g_header_1": 100,
  "font_color_b_header_1": 100,
  "font_family_title_1": "helvetica",
  "font_style_title_1": "B",
  "font_size_title_1": 12,
  "font_height_title_1": 5,
  "font_sep_title_1": 2,
  "font_color_r_title_1": 0,
  "font_color_g_title_1": 0,
  "font_color_b_title_1": 0,
  "font_family_section_1": "helvetica",
  "font_style_section_1": "B",
  "font_size_section_1": 8,
  "font_height_section_1": 5,
  "font_sep_section_1": 1,
  "font_color_r_section_1": 0,
  "font_color_g_section_1": 0,
  "font_color_b_section_1": 0,
  "font_family_section_3": "helvetica",
  "font_style_section_3": "B",
  "font_size_section_3": 7,
  "font_height_section_3": 4,
  "font_sep_section_3": 0,
  "font_color_r_section_3": 0,
  "font_color_g_section_3": 0,
  "font_color_b_section_3": 0,
  "font_family_normal": "helvetica",
  "font_style_normal": "",
  "font_size_normal": 8,
  "font_height_normal": 4,
  "font_sep_normal": 0,
  "font_color_r_normal": 0,
  "font_color_g_normal": 0,
  "font_color_b_normal": 0,
  "font_family_normal_bold": "helvetica",
  "font_style_normal_bold": "B",
  "font_size_normal_bold": 8,
  "font_height_normal_bold": 4,
  "font_sep_normal_bold": 0,
  "font_color_r_normal_bold": 0,
  "font_color_g_normal_bold": 0,
  "font_color_b_normal_bold": 0,
  "font_family_large_1": "helvetica",
  "font_style_large_1": "",
  "font_size_large_1": 9,
  "font_height_large_1": 4,
  "font_sep_large_1": 0,
  "font_color_r_large_1": 0,
  "font_color_g_large_1": 0,
  "font_color_b_large_1": 0,
  "font_family_large_1_bold": "helvetica",
  "font_style_large_1_bold": "B",
  "font_size_large_1_bold": 9,
  "font_height_large_1_bold": 4,
  "font_sep_large_1_bold": 0,
  "font_color_r_large_1_bold": 0,
  "font_color_g_large_1_bold": 0,
  "font_color_b_large_1_bold": 0,
  "font_family_small_1": "helvetica",
  "font_style_small_1": "",
  "font_size_small_1": 7,
  "font_height_small_1": 3.5,
  "font_sep_small_1": 0,
  "font_color_r_small_1": 0,
  "font_color_g_small_1": 0,
  "font_color_b_small_1": 0,
  "font_family_small_1_bold": "helvetica",
  "font_style_small_1_bold": "B",
  "font_size_small_1_bold": 7,
  "font_height_small_1_bold": 3.5,
  "font_sep_small_1_bold": 0,
  "font_color_r_small_1_bold": 0,
  "font_color_g_small_1_bold": 0,
  "font_color_b_small_1_bold": 0,
  "font_family_small_1_bold_red": "helvetica",
  "font_style_small_1_bold_red": "B",
  "font_size_small_1_bold_red": 7,
  "font_height_small_1_bold_red": 3.5,
  "font_sep_small_1_bold_red": 0,
  "font_color_r_small_1_bold_red": 180,
  "font_color_g_small_1_bold_red": 0,
  "font_color_b_small_1_bold_red": 0,
  "font_family_small_1_bold_green": "helvetica",
  "font_style_small_1_bold_green": "B",
  "font_size_small_1_bold_green": 7,
  "font_height_small_1_bold_green": 3.5,
  "font_sep_small_1_bold_green": 0,
  "font_color_r_small_1_bold_green": 0,
  "font_color_g_small_1_bold_green": 180,
  "font_color_b_small_1_bold_green": 0,
  "font_family_small_1_bold_yellow": "helvetica",
  "font_style_small_1_bold_yellow": "B",
  "font_size_small_1_bold_yellow": 7,
  "font_height_small_1_bold_yellow": 3.5,
  "font_sep_small_1_bold_yellow": 0,
  "font_color_r_small_1_bold_yellow": 180,
  "font_color_g_small_1_bold_yellow": 180,
  "font_color_b_small_1_bold_yellow": 0,
  "font_family_small_2": "helvetica",
  "font_style_small_2": "",
  "font_size_small_2": 5,
  "font_height_small_2": 3,
  "font_sep_small_2": 0,
  "font_color_r_small_2": 0,
  "font_color_g_small_2": 0,
  "font_color_b_small_2": 0,
  "font_family_small_2_bold": "helvetica",
  "font_style_small_2_bold": "B",
  "font_size_small_2_bold": 5,
  "font_height_small_2_bold": 3,
  "font_sep_small_2_bold": 0,
  "font_color_r_small_2_bold": 0,
  "font_color_g_small_2_bold": 0,
  "font_color_b_small_2_bold": 0,
  "font_family_small_2_bold_red": "helvetica",
  "font_style_small_2_bold_red": "B",
  "font_size_small_2_bold_red": 5,
  "font_height_small_2_bold_red": 3,
  "font_sep_small_2_bold_red": 0,
  "font_color_r_small_2_bold_red": 180,
  "font_color_g_small_2_bold_red": 0,
  "font_color_b_small_2_bold_red": 0,
  "font_family_small_3": "helvetica",
  "font_style_small_3": "",
  "font_size_small_3": 6,
  "font_height_small_3": 3.2,
  "font_sep_small_3": 0,
  "font_color_r_small_3": 0,
  "font_color_g_small_3": 0,
  "font_color_b_small_3": 0,
  "font_family_small_3_bold": "helvetica",
  "font_style_small_3_bold": "B",
  "font_size_small_3_bold": 6,
  "font_height_small_3_bold": 3.2,
  "font_sep_small_3_bold": 0,
  "font_color_r_small_3_bold": 0,
  "font_color_g_small_3_bold": 0,
  "font_color_b_small_3_bold": 0,
  "row_height": 4,
  "width_date": 12,
  "width_date_2": 18,
  "width_name": 55,
  "width_name_2": 72,
  "width_age": 5,
  "width_record": 12,
  "width_state_id_1": 12,
  "width_department_tolerance": 130,
  "width_department": 40,
  "width_result": 20,
  "width_result_2": 15,
  "width_result_3": 23,
  "width_history_sep": 3,
  "width_datetime": 20,
  "width_status": 20,
  "width_status_2": 18,
  "width_caption_1": 15,
  "tolerance_address": 150,
  "tolerance_comments": 150,
  "tolerance_comments_2": 130,
  "tolerance_info": 150,
  "row_disp": 0,
  "rows_per_page": 20,
  "row_header_fill_color": (230, 230, 230),
  "cell_normal_fill_color": (230, 230, 230),
  "cell_warning_fill_color": (255, 190, 190),
  "cell_attention_fill_color": (255, 255, 0),
  "max_y_pos": 260,
  "reserved_height_1": 12,
  "reserved_height_2": 50,
  "xls_width_0": 15500,
  "xls_width_1": 5000,
  "xls_width_2": 3500,
  "xls_width_3": 3500,
  "xls_width_4": 15500,
  "xls_width_5": 5000,
  "xls_width_6": 5000,
  "xls_width_7": 8000,
  "width_count_result" : 30,
  "width_count_method" : 20,
  "page_header_sep_summary_1": 10,
  "row_height_summary_1": 8,
  "width_summary_1_label": 70,
  "width_summary_1_content": 100
}

sDestinationFolder = "reports"

# --------------
# Global objects
# --------------

dConfig = None

sHost = None
sUser = None
sPassword = None
sDatabase = None
sSMTPServer = None
uSMTPPort = None
sSMTPSender = None
sSMTPPassword = None

dMemLog = io.StringIO(newline=None)

dMySqlDB = None
dMySqlCursor = None
  
oSMTPServer = None
    
lsSocialInstitutionClasses = list()

# -----------------
# Auxiliary methods
# -----------------

# https://stackoverflow.com/a/517974
def RemoveAccents(sInput):
  sNFKD = unicodedata.normalize('NFKD', sInput)
  sRes = u"".join([c for c in sNFKD if not unicodedata.combining(c)])
  sRes = listools.CleanStr(sRes)
  return sRes

def StrLatin1(sInput):
  sRes = sInput.encode("latin-1", errors="replace").decode("latin-1")
  return sRes

def CleanStrName(sSrc):
  sRes = re.sub(r"^\s+", "", str(sSrc))
  sRes = re.sub(r"\s+$", "", sRes)
  sRes = re.sub("\'|\"|\^|\\.", "", sRes)
  sRes = re.sub(r"\s+", " ", sRes)
  return sRes

def CleanStrComments(sSrc):
  sRes = re.sub(r"\s+", " ", str(sSrc))
  sRes = re.sub("\'|\"|\^|\\.", "", sRes)
  sRes = re.sub(r"^\s+|\s+$", "", sRes)
  return sRes

def CleanStr(sSrc):
  sRes = re.sub(r"^\s+", "", str(sSrc))
  sRes = re.sub(r"\s+$", "", sRes)
  sRes = re.sub("\'|\"|\^|\\.", " ", sRes)
  sRes = re.sub(r"\s+", " ", sRes)
  return sRes

def GetGreeting():
  uHour = datetime.datetime.now().hour
  if uHour >= 20 or uHour < 6:
    return "Boa noite"
  elif uHour >= 12 and uHour < 20:
    return "Boa tarde"
  else:
    return "Bom dia"
        
# -----------
# XML methods
# -----------

# https://stackoverflow.com/a/30923963
def XMLToDict(oXML, bRoot = True):
  if bRoot:
    return {oXML.tag : XMLToDict(oXML, False)}
  dRes = copy(oXML.attrib)
  if oXML.text:
    dRes["_text"] = oXML.text
  for oElement in oXML.findall("./*"):
    if oElement.tag not in dRes:
      dRes[oElement.tag] = []
    dRes[oElement.tag].append(XMLToDict(oElement, False))
  return dRes

# --------------
# Config methods
# --------------
 
def LoadConfigXML():
  global \
    sConfigFolder,\
    sConfigXML,\
    dConfig,\
    sHost,\
    sUser,\
    sPassword,\
    sDatabase,\
    sSMTPServer,\
    uSMTPPort,\
    sSMTPSender,\
    sSMTPPassword,\
    lsSocialInstitutionClasses,\
    dEmailStockReportInputs
  oFile = open("{}\\{}".format(sConfigFolder, sConfigXML), "r", encoding = "utf-8")
  sFile = oFile.read()
  oFile.close()
  oXML = ElementTree.fromstring(sFile)
  dConfig = XMLToDict(oXML)['config']
  sHost = dConfig['host'][0]['_text']
  sUser = dConfig['user'][0]['_text']
  sPassword = dConfig['password'][0]['_text']
  sDatabase = dConfig['database'][0]['_text']
  sSMTPServer = dConfig['smtp_server'][0]['_text']
  uSMTPPort = int(dConfig['smtp_port'][0]['_text'])
  sSMTPSender = dConfig['smtp_sender'][0]['_text']
  sSMTPPassword = dConfig['smtp_password'][0]['_text']
  for dClass in dConfig['social_institution_classes'][0]['class']:
    lsSocialInstitutionClasses.append(dClass['_text'])
  dConfig['method_captions'] = dict()
  for dMethod in dConfig['methods'][0]['method']:
    dConfig['method_captions'][dMethod['id']] = dMethod['_text']
  dConfig['status_captions'] = dict()
  for dStatus in dConfig['status_list'][0]['status']:
    dConfig['status_captions'][dStatus['id']] = dStatus['_text'] if '_text' in dStatus else ""
  dConfig['result_captions'] = dict()
  for dResult in dConfig['results'][0]['result']:
    dConfig['result_captions'][dResult['id']] = dResult['_text']
  dEmailStockReportInputs = dict()
  for dStock in dConfig['stock_references'][0]['stock']:
    if 'acquired' in dStock and dStock['acquired'].lower() == "true":
      sKey = "{}_stock".format(dStock['id'])
      fFactor = float(dStock['factor']) if 'factor' in dStock else 1.0
      sMethod = dStock['method'] if 'method' in dStock else None
      dEmailStockReportInputs[sKey] = ("{} em stock".format(dStock['caption']), sMethod, fFactor)
      sKey = "{}_requested".format(dStock['id'])
      dEmailStockReportInputs[sKey] = ("{} pedidos".format(dStock['caption']), None, None)
    else:
      dEmailStockReportInputs[dStock['id']] = (dStock['caption'], None, None)
  return dConfig

# ----------------
# Database methods
# ----------------

def MySqlCursorStart():
  global dMySqlDB, dMySqlCursor
  global sHost, sUser, sPassword, sDatabase
  dMySqlDB = mysql.connector.connect(host = sHost, user = sUser, passwd = sPassword, database = sDatabase)
  dMySqlCursor = dMySqlDB.cursor(dictionary = True)
  dMySqlCursor.execute('SET NAMES utf8;')
  dMySqlCursor.execute('SET CHARACTER SET utf8;')
  dMySqlCursor.execute('SET character_set_connection=utf8;') 
  return

def MySqlCursorClose():
  global dMySqlDB, dMySqlCursor
  dMySqlCursor.close()
  dMySqlDB.close()
  return

def RunMySqlCursor(sSQLQuery, bFetch = False):
  global dMySqlCursor, sDateFormat, sDateTimeFormat
  dRes = list()
  dMySqlCursor.execute(sSQLQuery)
  if dMySqlCursor.rowcount == 0 or not bFetch:
    return dRes
  else:
    dRes = dMySqlCursor.fetchall()
    for dRow in dRes:
      for sKey in dRow:
        if isinstance(dRow[sKey], int):
          dRow[sKey] = str(dRow[sKey])
        elif isinstance(dRow[sKey], float):
          dRow[sKey] = str(int(dRow[sKey]))
        elif isinstance(dRow[sKey], datetime.datetime):
          dRow[sKey] = dRow[sKey].strftime(sDateTimeFormat)
        elif isinstance(dRow[sKey], datetime.date):
          dRow[sKey] = dRow[sKey].strftime(sDateFormat)
        elif dRow[sKey] == None:
          dRow[sKey] = "NULL"
        else:
          dRow[sKey] = str(dRow[sKey])
  return dRes

def RunMySqlBatch(sSQLQuery, bFetch = False):
  global sHost, sUser, sPassword, sDatabase, sCode
  sCmd = "mysql -h {0} -u {1} -p{2} {3} -B -e \"{4}\"".format(sHost, sUser, sPassword, sDatabase, sSQLQuery)
  lsLines = subprocess.check_output(sCmd, stderr=subprocess.DEVNULL).decode(sCode).splitlines()
  if len(lsLines) == 0:
    return list()
  lsFields = lsLines[0].split("\t")
  ldRes = list()
  for sLine in lsLines[1:]:
    lsLine = sLine.split("\t")
    ldNew = dict()
    uIndex = 0
    for sField in lsLine:
      ldNew[lsFields[uIndex]] = sField
      uIndex = uIndex + 1
    ldRes.append(ldNew)
  return ldRes

def RunMySql(sSQLQuery, bFetch = False):
  #return RunMySqlBatch(sSQLQuery, bFetch)
  return RunMySqlCursor(sSQLQuery, bFetch)

def CommitMySql():
  global dMySqlDB
  if dMySqlDB:
    dMySqlDB.commit()

def GetResult(sResult):
  global dConfig
  return dConfig['result_captions'][sResult]

def GetMethod(sMethod):
  global dConfig
  return dConfig['method_captions'][sMethod]

def GetStatus(sStatus):
  global dConfig
  return dConfig['status_captions'][sStatus] if sStatus in dConfig['status_captions'] else ""

def GetText(sValue):
  if sValue and sValue.upper() != "NULL":
    return sValue
  else:
    return ""

def GetDBString(sValue):
  if sValue == "NULL" or sValue == "null" or not sValue:
    return "NULL"
  else:
    return "'{}'".format(sValue)

# -----------
# Log methods
# -----------

def Log(sText):
  global dMemLog
  dMemLog.write(sText + os.linesep)
  print(sText)

def LogSave():
  global dMemLog, sLogFileTimestampFormat, sLogFolder
  dMemLog.seek(0)
  sTimeStamp = datetime.datetime.now().strftime(sLogFileTimestampFormat)
  oFile = open("{}\\{}_coviddxdb_reports.log".format(sLogFolder, sTimeStamp), "w", encoding = "utf-8")
  oFile.write(dMemLog.read())
  oFile.close()

# -----------
# PDF methods
# -----------

def StartReport():
  global oReport, dConfigReport, sConfigFolder
  oReport = fpdf.FPDF(
    orientation = dConfigReport['page_orientation'],
    unit = dConfigReport['page_unit'],
    format = dConfigReport['page_format'])
  oReport.set_margins(
    dConfigReport['page_margin_left'],
    dConfigReport['page_margin_top'],
    dConfigReport['page_margin_right'])
  #oReport.add_font("helvetica", "", "{}\\fonts\\FreeSans.ttf".format(sConfigFolder), uni = True)
  #oReport.add_font("helvetica", "B", "{}\\fonts\\FreeSansBold.ttf".format(sConfigFolder), uni = True)
  return

def AddPage():
  global oReport, dConfigReport
  oReport.add_page()
  return

def SetFont(sStyle):
  global oReport, dConfigReport
  oReport.set_font(
    dConfigReport["font_family_" + sStyle.lower()] if dConfigReport["font_family_" + sStyle.lower()] != None else "",
    dConfigReport["font_style_" + sStyle.lower()] if dConfigReport["font_style_" + sStyle.lower()] != None else "",
    dConfigReport["font_size_" + sStyle.lower()] if dConfigReport["font_size_" + sStyle.lower()] != None else 1)
  oReport.set_text_color(
    dConfigReport["font_color_r_" + sStyle.lower()] if dConfigReport["font_color_r_" + sStyle.lower()] != None else 0,
    dConfigReport["font_color_g_" + sStyle.lower()] if dConfigReport["font_color_g_" + sStyle.lower()] != None else 0,
    dConfigReport["font_color_b_" + sStyle.lower()] if dConfigReport["font_color_b_" + sStyle.lower()] != None else 0)
  return

def SetPos(fX, fY):
  global oReport, dConfigReport
  oReport.set_xy(dConfigReport['page_margin_left'] + fX, dConfigReport['page_margin_top'] + fY)
  return

def GetPos():
  global oReport, dConfigReport
  fCurrX = oReport.get_x() - dConfigReport['page_margin_left']
  fCurrY = oReport.get_y() - dConfigReport['page_margin_top']
  return (fCurrX, fCurrY)

def MovePos(fX, fY):
  global oReport, dConfigReport
  fCurrX = oReport.get_x()
  fCurrY = oReport.get_y()
  oReport.set_xy(fCurrX + fX, fCurrY + fY)
  return GetPos()

def GetStringWidth(sText, sStyle = None):
  global oReport, dConfigReport
  if sStyle:
    SetFont(sStyle)
  return oReport.get_string_width(sText)

def Write(sText, sStyle):
  global oReport, dConfigReport
  SetFont(sStyle)
  oReport.write(dConfigReport["font_height_" + sStyle.lower()], sText)
  return

def WriteLine(sText, sStyle, uAdvance = 0, fPosX = 0.0):
  global oReport, dConfigReport
  MovePos(dConfigReport["page_section_advance"] * float(uAdvance) + fPosX, 0.0)
  Write(sText + "\n", sStyle)
  MovePos(0.0, dConfigReport["font_sep_" + sStyle.lower()])
  return

def DrawCell(sText, sStyle, fWidth, fHeight, sAlign, sBorder, bFill = False, uAdvance = 0, fPosX = 0.0, tuFillColor = None):
  global oReport, dConfigReport
  if not fHeight:
    fActualHeight = dConfigReport["font_height_" + sStyle.lower()]
  else:
    fActualHeight = fHeight
  MovePos(dConfigReport["page_section_advance"] * float(uAdvance) + fPosX, 0.0)
  SetFont(sStyle)
  if bFill:
    if tuFillColor:
      oReport.set_fill_color(tuFillColor[0], tuFillColor[1], tuFillColor[2])
  oReport.cell(fWidth, h = fActualHeight, txt = sText, border = sBorder.upper(), ln = 1, align = sAlign, fill = bFill)
  return

def DrawRule(fWidth, sStyle = None):
  global oReport, dConfigReport
  fActualX1 = oReport.get_x() + dConfigReport['page_char_advance']
  fActualY1 = oReport.get_y()
  fActualX2 = fActualX1 + fWidth
  fActualY2 = fActualY1
  oReport.line(fActualX1, fActualY1, fActualX2, fActualY2)
  if sStyle:
    MovePos(0.0, dConfigReport["font_sep_" + sStyle.lower()])
  return

def DrawLine(fX1, fY1, fX2, fY2):
  global oReport, dConfigReport
  if not fX1:
    fActualX1 = oReport.get_x() + dConfigReport['page_char_advance']
  else:
    fActualX1 = dConfigReport['page_margin_left'] + fX1 + dConfigReport['page_char_advance']
  if not fY1:
    fActualY1 = oReport.get_y()
  else:
    fActualY1 = dConfigReport['page_margin_top'] + fY1
  if not fX2:
    fActualX2 = oReport.get_x() + dConfigReport['page_char_advance']
  else:
    fActualX2 = dConfigReport['page_margin_left'] + fX2 + dConfigReport['page_char_advance']
  if not fY2:
    fActualY2 = oReport.get_y()
  else:
    fActualY2 = dConfigReport['page_margin_top'] + fY2
  oReport.line(fActualX1, fActualY1, fActualX2, fActualY2)
  return

def DrawCircle(fX1, fY1, fRadius):
  global oReport, dConfigReport
  if not fX1:
    fActualX1 = oReport.get_x() + dConfigReport['page_char_advance']
  else:
    fActualX1 = dConfigReport['page_margin_left'] + fX1 + dConfigReport['page_char_advance']
  if not fY1:
    fActualY1 = oReport.get_y()
  else:
    fActualY1 = dConfigReport['page_margin_top'] + fY1
  oReport.ellipse(fActualX1 - fRadius, fActualY1 - fRadius, fRadius * 2.0, fRadius * 2.0)
  return

# -------------
# Email methods
# -------------

def EmailLogin(sSMTPServer, uSMTPPort, sSMTPUser, sSMTPPassword):
  global oSMTPServer
  oSSLContext = ssl.create_default_context()
  oSMTPServer = smtplib.SMTP(sSMTPServer, uSMTPPort)
  sSMTPHello1 = oSMTPServer.ehlo()
  oSMTPServer.starttls(context = oSSLContext)
  sSMTPHello2 = oSMTPServer.ehlo()
  oSMTPServer.login(sSMTPUser, sSMTPPassword)

def EmailLogout():
  global oSMTPServer
  oSMTPServer.quit()

def EmailSendWithFile(sSender, lsTo, lsCc, sSubject, sBody, sFilename, sFileTitle):
  global oSMTPServer
  oMessage = email.mime.multipart.MIMEMultipart()
  oMessage["From"] = sSender
  oMessage["To"] = ",".join(lsTo)
  oMessage["Cc"] = ",".join(lsCc)
  oMessage["Subject"] = sSubject
  oMessage.attach(email.mime.text.MIMEText(sBody, "plain"))
  with open(sFilename, "rb") as oAttachment:
    oPart = email.mime.base.MIMEBase("application", "octet-stream")
    oPart.set_payload(oAttachment.read())
  email.encoders.encode_base64(oPart)
  oPart.add_header("Content-Disposition", f"attachment; filename= {sFileTitle}")
  oMessage.attach(oPart)
  sText = oMessage.as_string()
  oSMTPServer.sendmail(sSender, lsTo + lsCc, sText)

def EmailSendWithMultipleFiles(sSender, lsTo, lsCc, sSubject, sBody, ltsFiles):
  global oSMTPServer
  oMessage = email.mime.multipart.MIMEMultipart()
  oMessage["From"] = sSender
  oMessage["To"] = ",".join(lsTo)
  oMessage["Cc"] = ",".join(lsCc)
  oMessage["Subject"] = sSubject
  oMessage.attach(email.mime.text.MIMEText(sBody, "plain"))
  for sFilename, sFileTitle in ltsFiles:
    with open(sFilename, "rb") as oAttachment:
      oPart = email.mime.base.MIMEBase("application", "octet-stream")
      oPart.set_payload(oAttachment.read())
    email.encoders.encode_base64(oPart)
    oPart.add_header("Content-Disposition", f"attachment; filename= {sFileTitle}")
    oMessage.attach(oPart)
  sText = oMessage.as_string()
  oSMTPServer.sendmail(sSender, lsTo + lsCc, sText)

# ------------------
# PDF report methods
# ------------------

def CreatePDFReport_Results(sFolder, sDatetimeStart = None, sDatetimeEnd = None, sGroup = "geral", lsFilterDepartment = None):
  global dConfig, dConfigReport, oReport, sDestinationFolder, dCounts
  if sDatetimeStart:
    oDatetimeStart = datetime.datetime.strptime(sDatetimeStart, "%Y-%m-%d %H:%M:%S")
  else:
    oDatetimeStart = (datetime.datetime.now() - datetime.timedelta(days = 1)).replace(hour = 9, minute = 0, second = 0, microsecond = 0)
  if sDatetimeEnd:
    oDatetimeEnd = datetime.datetime.strptime(sDatetimeEnd, "%Y-%m-%d %H:%M:%S")
  else:
    oDatetimeEnd = datetime.datetime.now().replace(hour = 9, minute = 0, second = 0, microsecond = 0)
  sDatetimeDBStart = oDatetimeStart.strftime("%Y-%m-%d %H:%M:%S")
  sDatetimeFilenameStart = oDatetimeStart.strftime(dConfig['report_1'][0]['datetime_filename'][0]['_text'])
  sDatetimeReportStart = oDatetimeStart.strftime(dConfig['report_1'][0]['datetime_report'][0]['_text'])
  sDatetimeDBEnd = oDatetimeEnd.strftime("%Y-%m-%d %H:%M:%S")
  sDatetimeFilenameEnd = oDatetimeEnd.strftime(dConfig['report_1'][0]['datetime_filename'][0]['_text'])
  sDatetimeReportEnd = oDatetimeEnd.strftime(dConfig['report_1'][0]['datetime_report'][0]['_text'])
  sDatetimeFilenameEmailCurr = datetime.datetime.now().strftime(dConfig['report_1'][0]['datetime_filename_email_curr'][0]['_text'])
  sDatetimeFilenameCurr = datetime.datetime.now().strftime(dConfig['report_1'][0]['datetime_filename_curr'][0]['_text'])
  sDatetimeReportCurr = datetime.datetime.now().strftime(dConfig['report_1'][0]['datetime_report_curr'][0]['_text'])
  sFilename = ("{0}\\{1}\\" + dConfig['report_1'][0]['filename_format'][0]['_text']).format(sDestinationFolder, sFolder, sDatetimeFilenameCurr, sGroup, sDatetimeFilenameStart, sDatetimeFilenameEnd)
  sFilenameTitle = dConfig['report_1'][0]['filename_email_format'][0]['_text'].format(sDatetimeFilenameEmailCurr)
  sQuery = """
    SELECT
      B.name AS Name,
      DATE_FORMAT(B.birthday, '%d/%m/%Y') AS Birthday,
      TIMESTAMPDIFF(YEAR, B.birthday, CURDATE()) AS Age,
      A.department AS TestDepartment,
      B.record AS Record,
      B.state_id_1 AS StateID1,
      B.patient_status AS Status,
      C.title AS DepartmentTitle,
      A.address_1 AS Address1,
      A.address_2 AS Address2,
      DATE_FORMAT(A.sample_date, '%d/%m/%Y') AS SampleDate,
      DATE_FORMAT(A.result_datetime, '%d/%m/%Y %H:%i:%s') AS ResultDateTime,
      A.result_code AS Result,
      A.patient_id AS PatientID,
      A.method_code AS Method,
      A.result_comments AS Comments
    FROM tests AS A
    LEFT JOIN patients AS B
      ON B.id = A.patient_id
    LEFT JOIN departments AS C
      ON C.id = B.department
    WHERE
      A.status <> 2
      AND A.result_code != 'notest'
      AND
      (
        A.result_code = 'waiting'
        OR
        (
          A.result_code != 'waiting'
          AND
          A.result_datetime >= '{}' AND A.result_datetime <= '{}'
        )
      )
    ORDER BY FIELD(A.result_code, 'positive', 'inconclusive', 'error', 'negative', 'waiting', 'notest'), A.name
    """.format(sDatetimeDBStart, sDatetimeDBEnd);
  dDBQueryRes = RunMySql(sQuery, True)
  sQuery2= """
    SELECT
      B.id AS PatientID,
      DATE_FORMAT(C.result_datetime, '%d/%m/%Y') AS ResultDate,
      C.result_code AS Result
    FROM tests AS A
    LEFT JOIN patients AS B
      ON B.id = A.patient_id
    LEFT JOIN tests AS C
      ON C.patient_id = B.id
      AND C.status <> 2
      AND C.result_datetime < A.result_datetime
    WHERE
      A.status <> 2
      AND A.result_code != 'waiting' AND A.result_code != 'notest'
      AND A.result_datetime >= '{}' AND A.result_datetime <= '{}'
    ORDER BY B.id, C.result_datetime DESC
    """.format(sDatetimeDBStart, sDatetimeDBEnd);
  dDBQueryRes2 = RunMySql(sQuery2, True)
  dHistory = dict()
  for dRow in dDBQueryRes2:
    if dRow['ResultDate'] and dRow['ResultDate'].upper() != "NULL":
      if dRow['PatientID'] not in dHistory:
        dHistory[dRow['PatientID']] = dict()
        dHistory[dRow['PatientID']]['results'] = list()
      sResultCode = dRow['Result'].lower()
      sResult = GetResult(sResultCode)
      dHistory[dRow['PatientID']]['results'].append({
        'date': dRow['ResultDate'],
        'result': sResult
        })
  for sPatientID, dEntry in dHistory.items():
    sHistory = ""
    for dResult in dEntry['results'][0:5]:
      sHistory = sHistory + "{}{} ({})".format(
        ", " if sHistory != "" else "",
        dResult['result'],
        dResult['date']
        )
    dHistory[sPatientID]['history'] = sHistory
  dCounts = dict()
  dCounts['_waiting'] = 0
  dCounts['_completed'] = 0
  dCounts['_new_positive'] = 0
  uNewPositiveCount = 0
  for dRow in dDBQueryRes:
    bFilter = False
    if not lsFilterDepartment or lsFilterDepartment[0] == "*" or lsFilterDepartment[0] == "":
      bFilter = True
    else:
      for sFilter in lsFilterDepartment:
        if re.match(sFilter, dRow['TestDepartment'], flags = re.IGNORECASE):
          bFilter = True
          break
    if not bFilter:
      continue
    sResult = dRow['Result'].lower()
    sMethod = dRow['Method'].lower()
    if sResult == "waiting":
      dCounts['_waiting'] = dCounts['_waiting'] + 1
    else:
      dCounts['_completed'] = dCounts['_completed'] + 1
    if sResult not in dCounts:
      dCounts[sResult] = dict()
      dCounts[sResult]['_total'] = 1
    else:
      dCounts[sResult]['_total'] = dCounts[sResult]['_total'] + 1
    if sMethod not in dCounts[sResult]:
      dCounts[sResult][sMethod] = 1
    else:
      dCounts[sResult][sMethod] = dCounts[sResult][sMethod] + 1
    if sMethod not in dCounts:
      dCounts[sMethod] = dict()
      dCounts[sMethod]['_total'] = 1
      if sResult == "waiting":
        dCounts[sMethod]['_waiting'] = 1
        dCounts[sMethod]['_completed'] = 0
      else:
        dCounts[sMethod]['_completed'] = 1
        dCounts[sMethod]['_waiting'] = 0
    else:
      dCounts[sMethod]['_total'] = dCounts[sMethod]['_total'] + 1
      if sResult == "waiting":
        dCounts[sMethod]['_waiting'] = dCounts[sMethod]['_waiting'] + 1
      else:
        dCounts[sMethod]['_completed'] = dCounts[sMethod]['_completed'] + 1
    if sResult not in dCounts[sMethod]:
      dCounts[sMethod][sResult] = 1
    else:
      dCounts[sMethod][sResult] = dCounts[sMethod][sResult] + 1
    if sResult == "positive":
      bNewPositive = True
      if dRow['PatientID'] in dHistory:
        for dResult in dHistory[dRow['PatientID']]['results']:
          oHistoryDatetime = datetime.datetime.strptime(dResult['date'], "%d/%m/%Y")
          oCurrDatetime = datetime.datetime.strptime(dRow['ResultDateTime'], "%d/%m/%Y %H:%M:%S")
          if dResult['result'].lower() == "não detetado":
            bNewPositive = True
            break;
          if dResult['result'].lower() == "detetado" and (oCurrDatetime - oHistoryDatetime).days < 90:
            bNewPositive = False
            break;
      if bNewPositive:
        uNewPositiveCount = uNewPositiveCount + 1
      else:
        dHistory[dRow['PatientID']]['old_positive'] = True
  dCounts['_new_positive'] = uNewPositiveCount
  StartReport()
  sQuantity = str(dCounts['_completed'])
  AddPage()
  SetPos(0,0)
  WriteLine("#INSTITUTION# | #DEPARTMENT#", "header_1")
  WriteLine(dConfig['report_1'][0]['title'][0]['_text'], "title_1")
  WriteLine("Período: {} a {}".format(sDatetimeReportStart, sDatetimeReportEnd), "small_1")
  WriteLine("Data e hora do ficheiro: {}".format(sDatetimeReportCurr), "small_2")
  MovePos(0.0, dConfigReport['page_header_sep'])
  WriteLine("Análises concluídas", "section_1")
  fXStartPos = GetPos()[0]
  fYStartPos = GetPos()[1]
  fXDisp = dConfigReport['page_char_advance'] + dConfigReport['width_count_result']
  fYDisp = 0 
  SetPos(fXStartPos + fXDisp , fYStartPos + fYDisp)
  uMethodCount = 0
  for sMethodCode, sMethodCaption in dConfig['method_captions'].items():
    uMethodCount = uMethodCount + 1
    DrawCell(sMethodCaption, "small_1_bold", dConfigReport['width_count_method'], dConfigReport['row_height'], "C", "LTBR", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_count_method']
    SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
  DrawCell("Total", "small_1_bold", dConfigReport['width_count_method'], dConfigReport['row_height'], "C", "LTBR", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
  fXDisp = dConfigReport['page_char_advance']
  fYDisp = dConfigReport['row_height']
  SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
  for sResultCode, sResultCaption in dConfig['result_captions'].items():
    sFont = "small_1_bold"
    if sResultCode == "positive":
      sFont = "small_1_bold_red"
    if sResultCode == "negative":
      sFont = "small_1_bold_green"
    if sResultCode == "inconclusive":
      sFont = "small_1_bold_yellow"
    if sResultCode == "error":
      sFont = "small_1_bold_yellow"
    DrawCell(sResultCaption, sFont, dConfigReport['width_count_result'], dConfigReport['row_height'], "L", "LTBR", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_count_result']
    SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
    for sMethodCode, sMethodCaption in dConfig['method_captions'].items():
      if sResultCode in dCounts and sMethodCode in dCounts[sResultCode]:
        sResult = str(dCounts[sResultCode][sMethodCode])
      else:
        sResult = "0"
      DrawCell(sResult, "small_1", dConfigReport['width_count_method'], dConfigReport['row_height'], "C", "LTRB")
      fXDisp = fXDisp + dConfigReport['width_count_method']
      SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
    if sResultCode in dCounts:
      sResult = str(dCounts[sResultCode]['_total'])
    else:
      sResult = "0"
    DrawCell(sResult, "small_1_bold", dConfigReport['width_count_method'], dConfigReport['row_height'], "C", "LTRB")
    fXDisp = dConfigReport['page_char_advance']
    fYDisp = fYDisp + dConfigReport['row_height']
    SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
  DrawCell("Total", "small_1_bold", dConfigReport['width_count_result'], dConfigReport['row_height'], "L", "LTBR", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
  fXDisp = fXDisp + dConfigReport['width_count_result']
  SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
  for sMethodCode, sMethodCaption in dConfig['method_captions'].items():
    if sMethodCode in dCounts:
      sResult = str(dCounts[sMethodCode]['_completed'])
    else:
      sResult = "0"
    DrawCell(sResult, "small_1_bold", dConfigReport['width_count_method'], dConfigReport['row_height'], "C", "LTRB")
    fXDisp = fXDisp + dConfigReport['width_count_method']
    SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
  sResult = str(dCounts['_completed'])
  DrawCell(sResult, "small_1_bold", dConfigReport['width_count_method'], dConfigReport['row_height'], "C", "LTRB")
  fYDisp = fYDisp + dConfigReport['row_height'] + dConfigReport['page_section_sep']
  fXDisp = dConfigReport['page_char_advance']
  SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
  DrawCell("Detetados de novo", "small_1_bold_red", dConfigReport['width_count_result'], dConfigReport['row_height'], "L", "LTBR", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
  fXDisp = fXDisp + dConfigReport['width_count_result']
  SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
  sResult = str(uNewPositiveCount)
  DrawCell(sResult, "small_1_bold", dConfigReport['width_count_method'] * (uMethodCount + 1), dConfigReport['row_height'], "C", "LTRB")
  fYDisp = fYDisp + dConfigReport['row_height']
  SetPos(fXStartPos, fYStartPos + fYDisp)
  MovePos(0.0, dConfigReport['page_section_sep'])
  if dDBQueryRes and dCounts['_completed'] > 0:
    fXStartPos = GetPos()[0]
    fYStartPos = GetPos()[1]
    fXDisp = dConfigReport['page_char_advance']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Nome", "small_1_bold", dConfigReport['width_name'], dConfigReport['row_height'], "C", "LTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_name']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("DN", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_date']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Id.", "small_2_bold", dConfigReport['width_age'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_age']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("#INTID1#", "small_2_bold", dConfigReport['width_record'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_record']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("#STATEID1#", "small_2_bold", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_state_id_1']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Departmento", "small_2_bold", dConfigReport['width_department'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_department']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Colheita", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_date']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Resultado", "small_1_bold", dConfigReport['width_result'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_result']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("DH do resultado", "small_2_bold", dConfigReport['width_datetime'], dConfigReport['row_height'], "C", "RTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_datetime']
    SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
    for dRow in dDBQueryRes:
      if dRow['Result'] == "waiting":
        continue
      if dRow['Comments'] and dRow['Comments'].strip() != "" and dRow['Comments'].strip().lower() != "null":
        sComments = dRow['Comments'].strip()
      else:
        sComments = None
      if sComments and 'excluded_comments_search' in dConfig:
        for dXMLElement in dConfig['excluded_comments_search'][0]['search']:
          if re.match(sComments, dXMLElement['_text'], flags = re.IGNORECASE):
            sComments = None
            break
      bFilter = False
      if not lsFilterDepartment or lsFilterDepartment[0] == "*" or lsFilterDepartment[0] == "":
        bFilter = True
      else:
        for sFilter in lsFilterDepartment:
          if re.match(sFilter, dRow['TestDepartment'], flags = re.IGNORECASE):
            bFilter = True
            break
      if not bFilter:
        continue
      fXStartPos = GetPos()[0]
      fYStartPos = GetPos()[1]
      if fYStartPos + dConfigReport['reserved_height_1'] > dConfigReport['max_y_pos']:
        AddPage()
        SetPos(0,0)
        WriteLine("#INSTITUTION# | #DEPARTMENT#", "header_1")
        WriteLine(dConfig['report_1'][0]['title'][0]['_text'], "title_1")
        WriteLine("Período: {} a {}".format(sDatetimeReportStart, sDatetimeReportEnd), "small_1")
        WriteLine("Data e hora do ficheiro: {}".format(sDatetimeReportCurr), "small_2")
        MovePos(0.0, dConfigReport['page_header_sep'])
        WriteLine("Análises concluídas", "section_1")
        sQuantity = str(dCounts['_completed'])
        WriteLine("Quantidade: {} (continuação da página anterior)".format(sQuantity), "small_1")
        MovePos(0.0, dConfigReport['page_section_sep'])
        fXStartPos = GetPos()[0]
        fYStartPos = GetPos()[1]
        fXDisp = dConfigReport['page_char_advance']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Nome", "small_1_bold", dConfigReport['width_name'], dConfigReport['row_height'], "C", "LTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_name']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("DN", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_date']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Id.", "small_2_bold", dConfigReport['width_age'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_age']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("#INTID1#", "small_2_bold", dConfigReport['width_record'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_record']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("#STATEID1#", "small_2_bold", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_state_id_1']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Departmento", "small_2_bold", dConfigReport['width_department'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_department']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Colheita", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_date']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Resultado", "small_1_bold", dConfigReport['width_result'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_result']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("DH do resultado", "small_2_bold", dConfigReport['width_datetime'], dConfigReport['row_height'], "C", "RTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_datetime']
        SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
        fXStartPos = GetPos()[0]
        fYStartPos = GetPos()[1]
      fXDisp = dConfigReport['page_char_advance']
      fXDispStart = fXDisp
      sName = str(dRow['Name']) if str(dRow['Name']).upper() != "NULL" else ""
      sBirthday = str(dRow['Birthday']) if str(dRow['Birthday']).upper() != "NULL" else ""
      sAge = str(dRow['Age']) if str(dRow['Age']).upper() != "NULL" else ""
      sRecord = str(dRow['Record']) if str(dRow['Record']).upper() != "NULL" else ""
      sStateID1 = str(dRow['StateID1']) if str(dRow['StateID1']).upper() != "NULL" else ""
      sDepartmentTitle = str(dRow['DepartmentTitle']) if str(dRow['DepartmentTitle']).upper() != "NULL" else ""
      sDepartmentTitle = str(dRow['TestDepartment']) if str(dRow['TestDepartment']).upper() != "NULL" else ""
      sSampleDate = str(dRow['SampleDate']) if str(dRow['SampleDate']).upper() != "NULL" else ""
      sResult = GetResult(str(dRow['Result']))
      sResultDateTime = str(dRow['ResultDateTime']) if str(dRow['ResultDateTime']).upper() != "NULL" else ""
      sAddress1 = str(dRow['Address1']) if str(dRow['Address1']).upper() != "NULL" else ""
      sAddress2 = str(dRow['Address2']) if str(dRow['Address2']).upper() != "NULL" else ""
      SetPos(fXStartPos + fXDisp, fYStartPos)
      uNameWidth = GetStringWidth(sName, "small_1_bold")
      DrawCell(sName if uNameWidth < dConfigReport['width_name'] else sName.split()[0] + " " + sName.split()[1] + " " + sName.split()[-1], "small_1_bold", dConfigReport['width_name'], dConfigReport['row_height'], "L", "LT")
      fXDisp = fXDisp + dConfigReport['width_name']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sBirthday, "small_2", dConfigReport['width_date'], dConfigReport['row_height'], "C", "T")
      fXDisp = fXDisp + dConfigReport['width_date']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sAge, "small_1", dConfigReport['width_age'], dConfigReport['row_height'], "C", "T")
      fXDisp = fXDisp + dConfigReport['width_age']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sRecord, "small_2", dConfigReport['width_record'], dConfigReport['row_height'], "C", "T")
      fXDisp = fXDisp + dConfigReport['width_record']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sStateID1, "small_2", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "T")
      fXDisp = fXDisp + dConfigReport['width_state_id_1']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      uDepartmentWidth = GetStringWidth(dRow['DepartmentTitle'], "small_2")
      DrawCell(sDepartmentTitle if uDepartmentWidth < dConfigReport['width_department_tolerance'] else sDepartmentTitle[0:20], "small_2", dConfigReport['width_department'], dConfigReport['row_height'], "C", "T")
      fXDisp = fXDisp + dConfigReport['width_department']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sSampleDate, "small_2", dConfigReport['width_date'], dConfigReport['row_height'], "C", "T")
      fXDisp = fXDisp + dConfigReport['width_date']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      if dRow['Result'].lower() == "positive":
        if dRow['PatientID'] in dHistory and 'old_positive' in dHistory[dRow['PatientID']] and dHistory[dRow['PatientID']]['old_positive']:
          bSelFill = True
          tuSelFillColor = dConfigReport['cell_attention_fill_color']
          sSelStyle = "small_1_bold"
        else:
          bSelFill = True
          tuSelFillColor = dConfigReport['cell_warning_fill_color']
          sSelStyle = "small_1_bold"
      elif dRow['Result'].lower() == "inconclusive":
        bSelFill = True
        tuSelFillColor = dConfigReport['cell_attention_fill_color']
        sSelStyle = "small_1"
      else:
        bSelFill = False
        tuSelFillColor = None
        sSelStyle = "small_1"
      if sComments:
        sResult = "{}*".format(sResult)
      DrawCell(sResult, sSelStyle, dConfigReport['width_result'], dConfigReport['row_height'], "C", "T", bFill = bSelFill, tuFillColor = tuSelFillColor)
      fXDisp = fXDisp + dConfigReport['width_result']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sResultDateTime, "small_2", dConfigReport['width_datetime'], dConfigReport['row_height'], "C", "RT")
      fXDisp = fXDisp + dConfigReport['width_datetime']
      fXDispEnd = fXDisp
      fYStartPos = fYStartPos + dConfigReport['row_height'] * 1
      fXDisp = fXDispStart
      SetPos(fXStartPos + fXDispStart, fYStartPos)
      if sComments:
        DrawCell("Comentário", "small_2_bold_red", dConfigReport['width_caption_1'], dConfigReport['row_height'], "C", "L")
      else:
        DrawCell("Morada", "small_2", dConfigReport['width_caption_1'], dConfigReport['row_height'], "C", "L")
      fXDisp = fXDisp + dConfigReport['width_caption_1']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      uCellWidth = fXDispEnd - fXDisp
      sAddress = "{} | {}".format(sAddress2, sAddress1)
      uAddressWidth = GetStringWidth(sAddress, "small_1")
      if sComments:
        uCommentsWidth = GetStringWidth(sComments, "small_1")
        DrawCell(sComments if uCommentsWidth < uCellWidth - 5 else sComments[0:dConfigReport['tolerance_comments']], "small_1", uCellWidth, dConfigReport['row_height'], "L", "R")
      else:
        DrawCell(sAddress if uAddressWidth < uCellWidth - 5 else sAddress[0:dConfigReport['tolerance_address']], "small_1", uCellWidth, dConfigReport['row_height'], "L", "R")
      fYStartPos = fYStartPos + dConfigReport['row_height'] * 1
      fXDisp = fXDispStart
      SetPos(fXStartPos + fXDispStart, fYStartPos)
      DrawCell("Histórico", "small_2", dConfigReport['width_caption_1'], dConfigReport['row_height'], "C", "LB")
      fXDisp = fXDisp + dConfigReport['width_caption_1']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      if dRow['PatientID'] in dHistory:
        for dEntry in dHistory[dRow['PatientID']]['results'][0:5]:
          sEntry = "{} ({})".format(dEntry['result'], dEntry['date'])
          if dEntry['result'].lower() == "detetado":
            bSelFill = True
            tuSelFillColor = dConfigReport['cell_warning_fill_color']
            sSelStyle = "small_2"
          elif dEntry['result'].lower() == "inconclusivo":
            bSelFill = True
            tuSelFillColor = dConfigReport['cell_attention_fill_color']
            sSelStyle = "small_2"
          else:
            bSelFill = True
            tuSelFillColor = dConfigReport['cell_normal_fill_color']
            sSelStyle = "small_2"
          DrawCell(sEntry, sSelStyle, dConfigReport['width_result_3'], dConfigReport['row_height'], "C", "B", bFill = bSelFill, tuFillColor = tuSelFillColor)
          fXDisp = fXDisp + dConfigReport['width_result_3']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          DrawCell("", "small_1", dConfigReport['width_history_sep'], dConfigReport['row_height'], "L", "B")
          fXDisp = fXDisp + dConfigReport['width_history_sep']
          SetPos(fXStartPos + fXDisp, fYStartPos)
      uCellWidth = fXDispEnd - fXDisp
      DrawCell("", "small_1", uCellWidth, dConfigReport['row_height'], "L", "BR")
      SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
  MovePos(0.0, dConfigReport['page_header_sep'])
  fXStartPos = GetPos()[0]
  fYStartPos = GetPos()[1]
  if fYStartPos + dConfigReport['reserved_height_2'] > dConfigReport['max_y_pos']:
    AddPage()
    SetPos(0,0)
    WriteLine("#INSTITUTION# | #DEPARTMENT#", "header_1")
    WriteLine(dConfig['report_1'][0]['title'][0]['_text'], "title_1")
    WriteLine("Data e hora do ficheiro: {}".format(sDatetimeReportCurr), "small_2")
  MovePos(0.0, dConfigReport['page_header_sep'])
  WriteLine("Resultados pendentes", "section_1")
  sQuantity = str(dCounts['_waiting'])
  WriteLine("Quantidade: {}".format(sQuantity), "small_1")
  if dDBQueryRes and dCounts['_waiting'] > 0:
    MovePos(0.0, dConfigReport['page_section_sep'])
    fXStartPos = GetPos()[0]
    fYStartPos = GetPos()[1]
    fXDisp = dConfigReport['page_char_advance']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Nome", "small_1_bold", dConfigReport['width_name'], dConfigReport['row_height'], "C", "LTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_name']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("DN", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_date']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Id.", "small_2_bold", dConfigReport['width_age'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_age']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("#INTID1#", "small_2_bold", dConfigReport['width_record'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_record']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("#STATEID1#", "small_2_bold", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_state_id_1']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Departmento", "small_2_bold", dConfigReport['width_department'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_department']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Colheita", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TBR", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_date']
    SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
    for dRow in dDBQueryRes:
      if dRow['Result'] != "waiting":
        continue
      bFilter = False
      if not lsFilterDepartment or lsFilterDepartment[0] == "*" or lsFilterDepartment[0] == "":
        bFilter = True
      else:
        for sFilter in lsFilterDepartment:
          if re.match(sFilter, dRow['TestDepartment'], flags = re.IGNORECASE):
            bFilter = True
            break
      if not bFilter:
        continue
      fXStartPos = GetPos()[0]
      fYStartPos = GetPos()[1]
      if fYStartPos + dConfigReport['reserved_height_1'] > dConfigReport['max_y_pos']:
        AddPage()
        SetPos(0,0)
        WriteLine("#INSTITUTION# | #DEPARTMENT#", "header_1")
        WriteLine(dConfig['report_1'][0]['title'][0]['_text'], "title_1")
        WriteLine("Data e hora do ficheiro: {}".format(sDatetimeReportCurr), "small_2")
        MovePos(0.0, dConfigReport['page_header_sep'])
        WriteLine("Resultados pendentes", "section_1")
        sQuantity = str(dCounts['_waiting'])
        WriteLine("Quantidade: {} (continuação da página anterior)".format(sQuantity), "small_1")
        MovePos(0.0, dConfigReport['page_section_sep'])
        fXStartPos = GetPos()[0]
        fYStartPos = GetPos()[1]
        fXDisp = dConfigReport['page_char_advance']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Nome", "small_1_bold", dConfigReport['width_name'], dConfigReport['row_height'], "C", "LTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_name']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("DN", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_date']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Id.", "small_2_bold", dConfigReport['width_age'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_age']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("#INTID1#", "small_2_bold", dConfigReport['width_record'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_record']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("#STATEID1#", "small_2_bold", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_state_id_1']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Departmento", "small_2_bold", dConfigReport['width_department'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_department']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Colheita", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TBR", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_date']
        SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
        fXStartPos = GetPos()[0]
        fYStartPos = GetPos()[1]
      fXDisp = dConfigReport['page_char_advance']
      sName = str(dRow['Name']) if str(dRow['Name']).upper() != "NULL" else ""
      sBirthday = str(dRow['Birthday']) if str(dRow['Birthday']).upper() != "NULL" else ""
      sAge = str(dRow['Age']) if str(dRow['Age']).upper() != "NULL" else ""
      sRecord = str(dRow['Record']) if str(dRow['Record']).upper() != "NULL" else ""
      sStateID1 = str(dRow['StateID1']) if str(dRow['StateID1']).upper() != "NULL" else ""
      sDepartmentTitle = str(dRow['DepartmentTitle']) if str(dRow['DepartmentTitle']).upper() != "NULL" else ""
      sDepartmentTitle = str(dRow['TestDepartment']) if str(dRow['TestDepartment']).upper() != "NULL" else ""
      sSampleDate = str(dRow['SampleDate']) if str(dRow['SampleDate']).upper() != "NULL" else ""
      SetPos(fXStartPos + fXDisp, fYStartPos)
      uNameWidth = GetStringWidth(sName, "small_1_bold")
      DrawCell(sName if uNameWidth < dConfigReport['width_name'] else sName.split()[0] + " " + sName.split()[1] + " " + sName.split()[-1], "small_1_bold", dConfigReport['width_name'], dConfigReport['row_height'], "L", "LTB")
      fXDisp = fXDisp + dConfigReport['width_name']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sBirthday, "small_2", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB")
      fXDisp = fXDisp + dConfigReport['width_date']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sAge, "small_1", dConfigReport['width_age'], dConfigReport['row_height'], "C", "TB")
      fXDisp = fXDisp + dConfigReport['width_age']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sRecord, "small_2", dConfigReport['width_record'], dConfigReport['row_height'], "C", "TB")
      fXDisp = fXDisp + dConfigReport['width_record']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sStateID1, "small_2", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "TB")
      fXDisp = fXDisp + dConfigReport['width_state_id_1']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      uDepartmentWidth = GetStringWidth(dRow['DepartmentTitle'], "small_2")
      DrawCell(sDepartmentTitle if uDepartmentWidth < dConfigReport['width_department_tolerance'] else sDepartmentTitle[0:20], "small_2", dConfigReport['width_department'], dConfigReport['row_height'], "C", "TB")
      fXDisp = fXDisp + dConfigReport['width_department']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sSampleDate, "small_2", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TBR")
      fXDisp = fXDisp + dConfigReport['width_date']
      SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
  oReport.output(sFilename)
  return (sFilename, sFilenameTitle)

def CreatePDFReport_COVIDPatients(sFolder, sDatetimeStart = None, sDatetimeEnd = None):
  global dConfig, dConfigReport, oReport, sDestinationFolder
  if sDatetimeStart:
    oDatetimeStart = datetime.datetime.strptime(sDatetimeStart, "%Y-%m-%d %H:%M:%S")
  else:
    oDatetimeStart = datetime.datetime.strptime("2020-01-01 00:00:00", "%Y-%m-%d %H:%M:%S")
  if sDatetimeEnd:
    oDatetimeEnd = datetime.datetime.strptime(sDatetimeEnd, "%Y-%m-%d %H:%M:%S")
  else:
    oDatetimeEnd = datetime.datetime.now()
  sDatetimeDBStart = oDatetimeStart.strftime("%Y-%m-%d %H:%M:%S")
  sDatetimeFilenameStart = oDatetimeStart.strftime(dConfig['report_3'][0]['datetime_filename'][0]['_text'])
  sDatetimeReportStart = oDatetimeStart.strftime(dConfig['report_3'][0]['datetime_report'][0]['_text'])
  sDatetimeDBEnd = oDatetimeEnd.strftime("%Y-%m-%d %H:%M:%S")
  sDatetimeFilenameEnd = oDatetimeEnd.strftime(dConfig['report_3'][0]['datetime_filename'][0]['_text'])
  sDatetimeReportEnd = oDatetimeEnd.strftime(dConfig['report_3'][0]['datetime_report'][0]['_text'])
  sDatetimeFilenameEmailCurr = datetime.datetime.now().strftime(dConfig['report_3'][0]['datetime_filename_email_curr'][0]['_text'])
  sDatetimeFilenameCurr = datetime.datetime.now().strftime(dConfig['report_3'][0]['datetime_filename_curr'][0]['_text'])
  sDatetimeReportCurr = datetime.datetime.now().strftime(dConfig['report_3'][0]['datetime_report_curr'][0]['_text'])
  sFilename = ("{0}\\{1}\\" + dConfig['report_3'][0]['filename_format'][0]['_text']).format(sDestinationFolder, sFolder, sDatetimeFilenameCurr, sDatetimeFilenameStart, sDatetimeFilenameEnd)
  sFilenameTitle = dConfig['report_3'][0]['filename_email_format'][0]['_text'].format(sDatetimeFilenameEmailCurr)
  sQuery = """
    SELECT
      A2.name AS Name,
      DATE_FORMAT(A2.birthday, '%d/%m/%Y') AS Birthday,
      TIMESTAMPDIFF(YEAR, A2.birthday, CURDATE()) AS Age,
      A2.record AS Record,
      A2.state_id_1 AS StateID1,
      A2.patient_status AS Status,
      C.title AS DepartmentTitle,
      DATE_FORMAT(A3.sample_date, '%d/%m/%Y') AS LastPositiveSampleDate,
      DATE_FORMAT(A3.result_datetime, '%d/%m/%Y') AS LastPositiveResultDate,
      DATE_FORMAT(B2.sample_date, '%d/%m/%Y') AS LastSampleDate,
      DATE_FORMAT(B2.result_datetime, '%d/%m/%Y') AS LastResultDate,
      B2.result_code AS LastResult,
      B2.department AS TestDepartment,
      B2.address_1 AS Address1,
      B2.address_2 AS Address2,
      DATE_FORMAT(D2.sample_date, '%d/%m/%Y') AS FirstPositiveSampleDate,
      DATE_FORMAT(D2.result_datetime, '%d/%m/%Y') AS FirstPositiveResultDate
    FROM
    (
      SELECT
        patient_id AS PatientID,
        MAX(sample_id_2) AS TestSample
      FROM tests
      WHERE
        test_code = 'sarscov2'
        AND result_code = 'positive'
        AND status <> 2
        AND result_datetime >= '{0}' AND result_datetime <= '{1}'
      GROUP BY patient_id
    ) AS A1
    LEFT JOIN
    (
      SELECT
        patient_id AS PatientID,
        MAX(sample_id_2) AS TestSample
        FROM tests
        WHERE
          test_code = 'sarscov2'
          AND result_code IS NOT NULL
          AND result_code != 'notest'
          AND result_code != ''
          AND status <> 2
          AND result_datetime >= '{0}' AND result_datetime <= '{1}'
        GROUP BY patient_id
    ) AS B1
      ON B1.PatientID = A1.PatientID
    LEFT JOIN
    (
      SELECT
        A.PatientID AS PatientID,
        MIN(C.sample_id_2) AS TestSample
      FROM
      (
        SELECT
          patient_id AS PatientID,
          MAX(sample_id_2) AS TestSample
        FROM tests
        WHERE
          test_code = 'sarscov2'
          AND result_code = 'positive'
          AND status <> 2
          AND result_datetime >= '{0}' AND result_datetime <= '{1}'
        GROUP BY patient_id
      ) AS A
      LEFT JOIN tests AS B
        ON B.sample_id_2 = A.TestSample
        AND B.status <> 2
      LEFT JOIN tests AS C
        ON C.test_code = 'sarscov2'
        AND C.result_code = 'positive'
        AND C.patient_id = A.PatientID
        AND C.status <> 2
        AND C.result_datetime >= '{0}' AND C.result_datetime <= '{1}'
        AND DATEDIFF(B.result_datetime, C.result_datetime) <= 180
      GROUP BY A.PatientID
    ) AS D1
      ON D1.PatientID = A1.PatientID
    LEFT JOIN patients AS A2
      ON A2.id = A1.PatientID
    LEFT JOIN tests AS A3
      ON A3.sample_id_2 = A1.TestSample
      AND A3.status <> 2
    LEFT JOIN tests AS B2
      ON B2.sample_id_2 = B1.TestSample
      AND B2.status <> 2
    LEFT JOIN tests AS D2
      ON D2.sample_id_2 = D1.TestSample
      AND D2.status <> 2
    LEFT JOIN departments AS C
      ON C.id = A2.department
    ORDER BY D2.result_datetime DESC
    """.format(sDatetimeDBStart, sDatetimeDBEnd);
  dDBQueryRes = RunMySql(sQuery, True)
  StartReport()
  AddPage()
  SetPos(0,0)
  WriteLine("#INSTITUTION# | #DEPARTMENT#", "header_1")
  WriteLine(dConfig['report_3'][0]['title'][0]['_text'], "title_1")
  WriteLine("Data e hora: {}".format(sDatetimeReportCurr), "small_1")
  MovePos(0.0, dConfigReport['page_header_sep'])
  WriteLine("Doentes ativos para COVID-19", "section_1")
  WriteLine("Período: {} a {}".format(sDatetimeReportStart, sDatetimeReportEnd), "small_1")
  uQuantityNonResolved = 0
  uQuantityResolved = 0
  oNowDate = datetime.datetime.now()
  for dRow in dDBQueryRes:
    oLastResultDate = datetime.datetime.strptime(str(dRow['LastResultDate']), "%d/%m/%Y")
    if str(dRow['LastResult']).lower() != "negative" and (oNowDate - oLastResultDate).days <= 20:
      uQuantityNonResolved = uQuantityNonResolved + 1
    else:
      uQuantityResolved = uQuantityResolved + 1
  WriteLine("Quantidade: {}".format(str(uQuantityNonResolved)), "small_1")
  MovePos(0.0, dConfigReport['page_section_sep'])
  if dDBQueryRes and uQuantityNonResolved > 0:
    fXStartPos = GetPos()[0]
    fYStartPos = GetPos()[1]
    fXDisp = dConfigReport['page_char_advance']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Nome", "small_1_bold", dConfigReport['width_name_2'], dConfigReport['row_height'], "C", "LTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_name_2']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("DN", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_date']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Id.", "small_2_bold", dConfigReport['width_age'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_age']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("#INTID1#", "small_2_bold", dConfigReport['width_record'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_record']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("#STATEID1#", "small_2_bold", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_state_id_1']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Departmento", "small_2_bold", dConfigReport['width_department'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_department']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Prim. positivo", "small_3_bold", dConfigReport['width_date_2'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_date_2']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Últ. positivo", "small_3_bold", dConfigReport['width_date_2'], dConfigReport['row_height'], "C", "TBR", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_date_2']
    SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
    for dRow in dDBQueryRes:
      oLastResultDate = datetime.datetime.strptime(str(dRow['LastResultDate']), "%d/%m/%Y")
      if str(dRow['LastResult']).lower() == "negative" or (oNowDate - oLastResultDate).days > 20:
        continue
      fXStartPos = GetPos()[0]
      fYStartPos = GetPos()[1]
      if fYStartPos + dConfigReport['reserved_height_1'] > dConfigReport['max_y_pos']:
        AddPage()
        SetPos(0,0)
        WriteLine("#INSTITUTION# | #DEPARTMENT#", "header_1")
        WriteLine(dConfig['report_3'][0]['title'][0]['_text'], "title_1")
        WriteLine("Data e hora: {}".format(sDatetimeReportCurr), "small_1")
        MovePos(0.0, dConfigReport['page_header_sep'])
        WriteLine("Doentes ativos para COVID-19", "section_1")
        WriteLine("Período: {} a {}".format(sDatetimeReportStart, sDatetimeReportEnd), "small_1")
        WriteLine("Quantidade: {} (continuação da página anterior)".format(str(uQuantityNonResolved)), "small_1")
        MovePos(0.0, dConfigReport['page_section_sep'])
        fXStartPos = GetPos()[0]
        fYStartPos = GetPos()[1]
        fXDisp = dConfigReport['page_char_advance']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Nome", "small_1_bold", dConfigReport['width_name_2'], dConfigReport['row_height'], "C", "LTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_name_2']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("DN", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_date']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Id.", "small_2_bold", dConfigReport['width_age'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_age']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("#INTID1#", "small_2_bold", dConfigReport['width_record'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_record']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("#STATEID1#", "small_2_bold", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_state_id_1']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Departmento", "small_2_bold", dConfigReport['width_department'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_department']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Prim. positivo", "small_3_bold", dConfigReport['width_date_2'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_date_2']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Últ. positivo", "small_3_bold", dConfigReport['width_date_2'], dConfigReport['row_height'], "C", "TBR", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_date_2']
        SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
        fXStartPos = GetPos()[0]
        fYStartPos = GetPos()[1]
      fXDisp = dConfigReport['page_char_advance']
      fXDispStart = fXDisp
      sName = str(dRow['Name']) if str(dRow['Name']).upper() != "NULL" else ""
      sBirthday = str(dRow['Birthday']) if str(dRow['Birthday']).upper() != "NULL" else ""
      sAge = str(dRow['Age']) if str(dRow['Age']).upper() != "NULL" else ""
      sRecord = str(dRow['Record']) if str(dRow['Record']).upper() != "NULL" else ""
      sStateID1 = str(dRow['StateID1']) if str(dRow['StateID1']).upper() != "NULL" else ""
      sDepartmentTitle = str(dRow['DepartmentTitle']) if str(dRow['DepartmentTitle']).upper() != "NULL" else ""
      sDepartmentTitle = str(dRow['TestDepartment']) if str(dRow['TestDepartment']).upper() != "NULL" else ""
      sAddress1 = str(dRow['Address1']) if str(dRow['Address1']).upper() != "NULL" else ""
      sAddress2 = str(dRow['Address2']) if str(dRow['Address2']).upper() != "NULL" else ""
      sStatus = GetStatus(dRow['Status']) if str(dRow['Status']).upper() != "NULL" else ""
      sFirstPositiveResultDate = str(dRow['FirstPositiveResultDate']) if str(dRow['FirstPositiveResultDate']).upper() != "NULL" else ""
      sLastPositiveResultDate = str(dRow['LastPositiveResultDate']) if str(dRow['LastPositiveResultDate']).upper() != "NULL" else ""
      sLastResultDate = str(dRow['LastResultDate']) if str(dRow['LastResultDate']).upper() != "NULL" else ""
      sLastResult = GetResult(str(dRow['LastResult']))
      SetPos(fXStartPos + fXDisp, fYStartPos)
      uNameWidth = GetStringWidth(sName, "small_1_bold")
      DrawCell(sName if uNameWidth < dConfigReport['width_name_2'] else sName.split()[0] + " " + sName.split()[1] + " " + sName.split()[-1], "small_1_bold", dConfigReport['width_name_2'], dConfigReport['row_height'], "L", "LT")
      fXDisp = fXDisp + dConfigReport['width_name_2']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sBirthday, "small_2", dConfigReport['width_date'], dConfigReport['row_height'], "C", "T")
      fXDisp = fXDisp + dConfigReport['width_date']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sAge, "small_1", dConfigReport['width_age'], dConfigReport['row_height'], "C", "T")
      fXDisp = fXDisp + dConfigReport['width_age']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sRecord, "small_2", dConfigReport['width_record'], dConfigReport['row_height'], "C", "T")
      fXDisp = fXDisp + dConfigReport['width_record']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sStateID1, "small_2", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "T")
      fXDisp = fXDisp + dConfigReport['width_state_id_1']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      uDepartmentWidth = GetStringWidth(dRow['DepartmentTitle'], "small_2")
      DrawCell(sDepartmentTitle if uDepartmentWidth < dConfigReport['width_department_tolerance'] else sDepartmentTitle[0:20], "small_2", dConfigReport['width_department'], dConfigReport['row_height'], "C", "T")
      fXDisp = fXDisp + dConfigReport['width_department']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sFirstPositiveResultDate, "small_3", dConfigReport['width_date_2'], dConfigReport['row_height'], "C", "T", bFill = True, tuFillColor = dConfigReport['cell_warning_fill_color'])
      fXDisp = fXDisp + dConfigReport['width_date_2']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sLastPositiveResultDate, "small_3", dConfigReport['width_date_2'], dConfigReport['row_height'], "C", "TR", bFill = True, tuFillColor = dConfigReport['cell_warning_fill_color'])
      fXDisp = fXDisp + dConfigReport['width_date_2']
      fXDispEnd = fXDisp
      fYStartPos = fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp']
      fXDisp = fXDispStart
      SetPos(fXStartPos + fXDispStart, fYStartPos)
      DrawCell("Morada", "small_2", dConfigReport['width_caption_1'], dConfigReport['row_height'], "C", "LB")
      fXDisp = fXDisp + dConfigReport['width_caption_1']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      uCellWidth = fXDispEnd - fXDisp
      sAddress = "{} | {}".format(sAddress2, sAddress1)
      uAddressWidth = GetStringWidth(sAddress, "small_1")
      DrawCell(sAddress if uAddressWidth < uCellWidth - 5 else sAddress[0:dConfigReport['tolerance_address']], "small_1", uCellWidth, dConfigReport['row_height'], "L", "BR")
      SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
  MovePos(0.0, dConfigReport['page_header_sep'])
  fXStartPos = GetPos()[0]
  fYStartPos = GetPos()[1]
  if fYStartPos + dConfigReport['reserved_height_2'] > dConfigReport['max_y_pos']:
    AddPage()
    SetPos(0,0)
    WriteLine("#INSTITUTION# | #DEPARTMENT#", "header_1")
    WriteLine(dConfig['report_3'][0]['title'][0]['_text'], "title_1")
    WriteLine("Data e hora: {}".format(sDatetimeReportCurr), "small_1")
  MovePos(0.0, dConfigReport['page_header_sep'])
  WriteLine("Utentes recuperados ou sem reavaliação nos últimos 20 dias", "section_1")
  WriteLine("Período: {} a {}".format(sDatetimeReportStart, sDatetimeReportEnd), "small_1")
  WriteLine("Quantidade: {}".format(str(uQuantityResolved)), "small_1")
  MovePos(0.0, dConfigReport['page_section_sep'])
  if dDBQueryRes and uQuantityResolved > 0:
    fXStartPos = GetPos()[0]
    fYStartPos = GetPos()[1]
    fXDisp = dConfigReport['page_char_advance']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Nome", "small_1_bold", dConfigReport['width_name'], dConfigReport['row_height'], "C", "LTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_name']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("DN", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_date']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Id.", "small_2_bold", dConfigReport['width_age'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_age']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("#INTID1#", "small_2_bold", dConfigReport['width_record'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_record']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("#STATEID1#", "small_2_bold", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_state_id_1']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Departmento", "small_2_bold", dConfigReport['width_department'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_department']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Prim. pos.", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_date']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Últ. pos.", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_date']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Data últ. res.", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_date']
    SetPos(fXStartPos + fXDisp, fYStartPos)
    DrawCell("Últ. res.", "small_2_bold", dConfigReport['width_result_2'], dConfigReport['row_height'], "C", "TBR", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
    fXDisp = fXDisp + dConfigReport['width_result_2']
    SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
    for dRow in dDBQueryRes:
      oLastResultDate = datetime.datetime.strptime(str(dRow['LastResultDate']), "%d/%m/%Y")
      if str(dRow['LastResult']).lower() != "negative" and (oNowDate - oLastResultDate).days <= 20:
        continue
      fXStartPos = GetPos()[0]
      fYStartPos = GetPos()[1]
      if fYStartPos + dConfigReport['reserved_height_1'] > dConfigReport['max_y_pos']:
        AddPage()
        SetPos(0,0)
        WriteLine("#INSTITUTION# | #DEPARTMENT#", "header_1")
        WriteLine(dConfig['report_3'][0]['title'][0]['_text'], "title_1")
        WriteLine("Data e hora: {}".format(sDatetimeReportCurr), "small_1")
        MovePos(0.0, dConfigReport['page_header_sep'])
        WriteLine("Utentes recuperados ou sem reavaliação nos últimos 20 dias", "section_1")
        WriteLine("Período: {} a {}".format(sDatetimeReportStart, sDatetimeReportEnd), "small_1")
        WriteLine("Quantidade: {} (continuação da página anterior)".format(str(uQuantityResolved)), "small_1")
        MovePos(0.0, dConfigReport['page_section_sep'])
        fXStartPos = GetPos()[0]
        fYStartPos = GetPos()[1]
        fXDisp = dConfigReport['page_char_advance']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Nome", "small_1_bold", dConfigReport['width_name'], dConfigReport['row_height'], "C", "LTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_name']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("DN", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_date']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Id.", "small_2_bold", dConfigReport['width_age'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_age']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("#INTID1#", "small_2_bold", dConfigReport['width_record'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_record']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("#STATEID1#", "small_2_bold", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_state_id_1']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Departmento", "small_2_bold", dConfigReport['width_department'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_department']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Prim. pos.", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_date']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Últ. pos.", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_date']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Data últ. res.", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_date']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Últ. res.", "small_2_bold", dConfigReport['width_result_2'], dConfigReport['row_height'], "C", "TBR", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_result_2']
        SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
        fXStartPos = GetPos()[0]
        fYStartPos = GetPos()[1]
      fXDisp = dConfigReport['page_char_advance']
      sName = str(dRow['Name']) if str(dRow['Name']).upper() != "NULL" else ""
      sBirthday = str(dRow['Birthday']) if str(dRow['Birthday']).upper() != "NULL" else ""
      sAge = str(dRow['Age']) if str(dRow['Age']).upper() != "NULL" else ""
      sRecord = str(dRow['Record']) if str(dRow['Record']).upper() != "NULL" else ""
      sStateID1 = str(dRow['StateID1']) if str(dRow['StateID1']).upper() != "NULL" else ""
      sDepartmentTitle = str(dRow['DepartmentTitle']) if str(dRow['DepartmentTitle']).upper() != "NULL" else ""
      sDepartmentTitle = str(dRow['TestDepartment']) if str(dRow['TestDepartment']).upper() != "NULL" else ""
      sStatus = GetStatus(dRow['Status']) if str(dRow['Status']).upper() != "NULL" else ""
      sFirstPositiveResultDate = str(dRow['FirstPositiveResultDate']) if str(dRow['FirstPositiveResultDate']).upper() != "NULL" else ""
      sLastPositiveResultDate = str(dRow['LastPositiveResultDate']) if str(dRow['LastPositiveResultDate']).upper() != "NULL" else ""
      sLastResultDate = str(dRow['LastResultDate']) if str(dRow['LastResultDate']).upper() != "NULL" else ""
      sLastResult = GetResult(str(dRow['LastResult']))
      SetPos(fXStartPos + fXDisp, fYStartPos)
      uNameWidth = GetStringWidth(sName, "small_1_bold")
      DrawCell(sName if uNameWidth < dConfigReport['width_name'] else sName.split()[0] + " " + sName.split()[1] + " " + sName.split()[-1], "small_1_bold", dConfigReport['width_name'], dConfigReport['row_height'], "L", "LTB")
      fXDisp = fXDisp + dConfigReport['width_name']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sBirthday, "small_2", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB")
      fXDisp = fXDisp + dConfigReport['width_date']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sAge, "small_1", dConfigReport['width_age'], dConfigReport['row_height'], "C", "TB")
      fXDisp = fXDisp + dConfigReport['width_age']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sRecord, "small_2", dConfigReport['width_record'], dConfigReport['row_height'], "C", "TB")
      fXDisp = fXDisp + dConfigReport['width_record']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sStateID1, "small_2", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "TB")
      fXDisp = fXDisp + dConfigReport['width_state_id_1']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      uDepartmentWidth = GetStringWidth(dRow['DepartmentTitle'], "small_2")
      DrawCell(sDepartmentTitle if uDepartmentWidth < dConfigReport['width_department_tolerance'] else sDepartmentTitle[0:20], "small_2", dConfigReport['width_department'], dConfigReport['row_height'], "C", "TB")
      fXDisp = fXDisp + dConfigReport['width_department']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sFirstPositiveResultDate, "small_2", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['cell_warning_fill_color'])
      fXDisp = fXDisp + dConfigReport['width_date']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sLastPositiveResultDate, "small_2", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['cell_warning_fill_color'])
      fXDisp = fXDisp + dConfigReport['width_date']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      DrawCell(sLastResultDate, "small_2", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB")
      fXDisp = fXDisp + dConfigReport['width_date']
      SetPos(fXStartPos + fXDisp, fYStartPos)
      if dRow['LastResult'].lower() == "positive":
        bSelFill = True
        tuSelFillColor = dConfigReport['cell_warning_fill_color']
        sSelStyle = "small_2_bold"
      elif dRow['LastResult'].lower() == "inconclusive":
        bSelFill = True
        tuSelFillColor = dConfigReport['cell_attention_fill_color']
        sSelStyle = "small_2_bold"
      elif dRow['LastResult'].lower() == "error":
        bSelFill = True
        tuSelFillColor = dConfigReport['cell_attention_fill_color']
        sSelStyle = "small_2_bold"
      else:
        bSelFill = False
        tuSelFillColor = None
        sSelStyle = "small_2"
      DrawCell(sLastResult, sSelStyle, dConfigReport['width_result_2'], dConfigReport['row_height'], "C", "TBR", bFill = bSelFill, tuFillColor = tuSelFillColor)
      fXDisp = fXDisp + dConfigReport['width_result_2']
      SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
  oReport.output(sFilename)
  return (sFilename, sFilenameTitle)

def CreatePDFReport_ResultsGroups(sFolder, sDatetimeStart, sDatetimeEnd, lsClasses, lsClassTitles, sReportTitle):
  global dConfig, dConfigReport, oReport, sDestinationFolder, uTestCount
  if sDatetimeStart:
    oDatetimeStart = datetime.datetime.strptime(sDatetimeStart, "%Y-%m-%d %H:%M:%S")
  else:
    oDatetimeStart = (datetime.datetime.now() - datetime.timedelta(days = 1)).replace(hour = 9, minute = 0, second = 0, microsecond = 0)
  if sDatetimeEnd:
    oDatetimeEnd = datetime.datetime.strptime(sDatetimeEnd, "%Y-%m-%d %H:%M:%S")
  else:
    oDatetimeEnd = datetime.datetime.now().replace(hour = 9, minute = 0, second = 0, microsecond = 0)
  sDatetimeDBStart = oDatetimeStart.strftime("%Y-%m-%d %H:%M:%S")
  sDatetimeFilenameStart = oDatetimeStart.strftime(dConfig['report_4'][0]['datetime_filename'][0]['_text'])
  sDatetimeReportStart = oDatetimeStart.strftime(dConfig['report_4'][0]['datetime_report'][0]['_text'])
  sDatetimeDBEnd = oDatetimeEnd.strftime("%Y-%m-%d %H:%M:%S")
  sDatetimeFilenameEnd = oDatetimeEnd.strftime(dConfig['report_4'][0]['datetime_filename'][0]['_text'])
  sDatetimeReportEnd = oDatetimeEnd.strftime(dConfig['report_4'][0]['datetime_report'][0]['_text'])
  sDatetimeFilenameEmailCurr = datetime.datetime.now().strftime(dConfig['report_4'][0]['datetime_filename_email_curr'][0]['_text'])
  sDatetimeFilenameCurr = datetime.datetime.now().strftime(dConfig['report_4'][0]['datetime_filename_curr'][0]['_text'])
  sDatetimeReportCurr = datetime.datetime.now().strftime(dConfig['report_4'][0]['datetime_report_curr'][0]['_text'])
  sQuery = """
    SELECT
      A.id AS TestID,
      B.name AS Name,
      DATE_FORMAT(B.birthday, '%d/%m/%Y') AS Birthday,
      TIMESTAMPDIFF(YEAR, B.birthday, CURDATE()) AS Age,
      A.department AS TestDepartment,
      B.record AS Record,
      B.state_id_1 AS StateID1,
      B.patient_status AS Status,
      C.title AS DepartmentTitle,
      A.address_1 AS Address1,
      A.address_2 AS Address2,
      DATE_FORMAT(A.sample_date, '%d/%m/%Y') AS SampleDate,
      DATE_FORMAT(A.result_datetime, '%d/%m/%Y %H:%i:%s') AS ResultDateTime,
      A.result_code AS Result,
      A.result_comments AS Comments,
      A.patient_id AS PatientID,
      D."group" AS GroupTitle,
      D.class AS GroupClass,
      D.department AS GroupDepartment,
      D.category AS GroupCategory,
      D.location AS GroupLocation,
      DATE_FORMAT(D."date", '%d/%m/%Y') AS GroupDate
    FROM
    (
      SELECT
        TestID AS TestID,
        MAX(GroupID) AS GroupID
      FROM
      (
        SELECT
          A.id AS TestID,
          B.id AS GroupID
        FROM
        (
          SELECT *
          FROM tests
          WHERE 
            status <> 2
            AND ((result_datetime >= '{0}' AND result_datetime <= '{1}') OR (result_code = 'waiting' AND sample_date >= DATE_SUB('{0}', INTERVAL 2 DAY) AND sample_date <= DATE_ADD('{1}', INTERVAL 2 DAY)))
        ) AS A
        LEFT JOIN groups AS B
          ON
            B.status = 1
            AND
            (
              B.sample = A.sample_id
              OR
              B.state_id = A.state_id_1
              OR
              (
                (
                  B.name LIKE CONCAT('%', REPLACE(A.name, ' ', '%'), '%')
                  OR
                  A.name LIKE CONCAT('%', REPLACE(B.name, ' ', '%'), '%')
                )
                AND
                B.birthday = A.birthday
              )
            )
        GROUP BY A.id, B.id
      ) AS A
      GROUP BY TestID
    ) AS Z
    LEFT JOIN tests AS A
      ON A.id = Z.TestID
    LEFT JOIN patients AS B
      ON B.id = A.patient_id
    LEFT JOIN departments AS C
      ON C.id = B.department
    LEFT JOIN groups AS D
      ON D.id = Z.GroupID
    ORDER BY FIELD(A.result_code, 'positive', 'inconclusive', 'error', 'negative', 'waiting', 'notest'), A.name
    """.format(sDatetimeDBStart, sDatetimeDBEnd);
  dDBQueryRes = RunMySql(sQuery, True)
  sQuery2= """
    SELECT
      B.id AS PatientID,
      DATE_FORMAT(C.result_datetime, '%d/%m/%Y') AS ResultDate,
      C.result_code AS Result
    FROM tests AS A
    LEFT JOIN patients AS B
      ON B.id = A.patient_id
    LEFT JOIN tests AS C
      ON C.patient_id = B.id
      AND C.status <> 2
      AND C.result_datetime < A.result_datetime
    WHERE
      A.status <> 2
      AND A.result_code != 'waiting' AND A.result_code != 'notest'
      AND A.result_datetime >= '{}' AND A.result_datetime <= '{}'
    ORDER BY B.id, C.result_datetime DESC
    """.format(sDatetimeDBStart, sDatetimeDBEnd);
  dDBQueryRes2 = RunMySql(sQuery2, True)
  dHistory = dict()
  for dRow in dDBQueryRes2:
    if dRow['ResultDate'] and dRow['ResultDate'].upper() != "NULL":
      if dRow['PatientID'] not in dHistory:
        dHistory[dRow['PatientID']] = dict()
        dHistory[dRow['PatientID']]['results'] = list()
      sResultCode = dRow['Result'].lower()
      sResult = "Erro"
      dResTrans = {
        'negative': "Não detetado",
        'positive': "Detetado",
        'error': "Indeterminado",
        'inconclusive': "Inconclusivo"
        }
      if sResultCode in dResTrans:
        sResult = dResTrans[sResultCode]
      dHistory[dRow['PatientID']]['results'].append({
        'date': dRow['ResultDate'],
        'result': sResult
        })
  for sPatientID, dEntry in dHistory.items():
    sHistory = ""
    for dResult in dEntry['results'][0:5]:
      sHistory = sHistory + "{}{} ({})".format(
        ", " if sHistory != "" else "",
        dResult['result'],
        dResult['date']
        )
    dHistory[sPatientID]['history'] = sHistory
  lsGroups = []
  dGroups = dict()
  for dRow in dDBQueryRes:
    if dRow['GroupClass'] and dRow['GroupClass'] in lsClasses:
      dRow['Group'] = dRow['GroupTitle'] if dRow['GroupTitle'] else ""
      dRow['Department'] = dRow['GroupDepartment'] if dRow['GroupDepartment'] else ""
      dRow['Category'] = dRow['GroupCategory'] if dRow['GroupCategory'] else ""
      dRow['Location'] = dRow['GroupLocation'] if dRow['GroupLocation'] else ""
      dRow['Date'] = dRow['GroupDate'] if dRow['GroupDate'] else ""
      dRow[dRow['GroupClass']] = True
      if dRow['Group'] not in lsGroups:
        lsGroups.append(dRow['Group'])
      dGroups[dRow['TestID']] = dict()
      dGroups[dRow['TestID']]['group'] = dRow['Group']
      dGroups[dRow['TestID']]['department'] = dRow['Department']
      dGroups[dRow['TestID']]['dategory'] = dRow['Category']
      dGroups[dRow['TestID']]['location'] = dRow['Location']
      dGroups[dRow['TestID']]['date'] = dRow['Date']
  if len(lsGroups) == 0:
    return None
  uGroupIndex = 0
  ldRes = list()
  for sCurrGroup in lsGroups:
    sFilename = ("{0}\\{1}\\" + dConfig['report_4'][0]['filename_format'][0]['_text']).format(sDestinationFolder, sFolder, sDatetimeFilenameCurr, StrLatin1(sCurrGroup), sDatetimeFilenameStart, sDatetimeFilenameEnd)
    sFilenameTitle = dConfig['report_4'][0]['filename_email_format'][0]['_text'].format(sDatetimeFilenameEmailCurr, str(uGroupIndex + 1))
    StartReport()
    uClassIndex = 0
    for sClass in lsClasses:
      sClassTitle = lsClassTitles[uClassIndex]
      uClassIndex = uClassIndex + 1
      uTestCount = 0
      uPositiveCount = 0
      uNewPositiveCount = 0
      uNegativeCount = 0
      uInconclusiveCount = 0
      uErrorCount = 0
      uPendingCount = 0
      for dRow in dDBQueryRes:
        if sClass not in dRow or not dRow[sClass] or dRow['Group'] != sCurrGroup or dRow['Result'].lower() == "notest":
          continue
        if dRow['Result'].lower() != "waiting":
          uTestCount = uTestCount + 1
        if dRow['Result'].lower() == "waiting":
          uPendingCount = uPendingCount + 1
        if dRow['Result'].lower() == "positive":
          uPositiveCount = uPositiveCount + 1
          bNewPositive = True
          if dRow['PatientID'] in dHistory:
            for dResult in dHistory[dRow['PatientID']]['results']:
              oHistoryDatetime = datetime.datetime.strptime(dResult['date'], "%d/%m/%Y")
              oCurrDatetime = datetime.datetime.strptime(dRow['ResultDateTime'], "%d/%m/%Y %H:%M:%S")
              if dResult['result'].lower() == "não detetado":
                bNewPositive = True
                break;
              if dResult['result'].lower() == "detetado" and (oCurrDatetime - oHistoryDatetime).days < 90:
                bNewPositive = False
                break;
          if bNewPositive:
            uNewPositiveCount = uNewPositiveCount + 1
          else:
            dHistory[dRow['PatientID']]['old_positive'] = True
        if dRow['Result'].lower() == "negative":
          uNegativeCount = uNegativeCount + 1
        if dRow['Result'].lower() == "inconclusive":
          uInconclusiveCount = uInconclusiveCount + 1
        if dRow['Result'].lower() == "error":
          uErrorCount = uErrorCount + 1
      if uTestCount + uPendingCount == 0:
        continue
      AddPage()
      SetPos(0,0)
      WriteLine("#INSTITUTION# | #DEPARTMENT# | {}".format(sReportTitle), "header_1")
      MovePos(0.0, dConfigReport['page_header_sep'])
      WriteLine(sCurrGroup, "title_1")
      WriteLine(sClassTitle, "title_1")
      WriteLine("Data e hora: {}".format(sDatetimeReportCurr), "small_1")
      MovePos(0.0, dConfigReport['page_header_sep'])
      WriteLine("Análises concluídas", "section_1")
      WriteLine("Período: {} a {}".format(sDatetimeReportStart, sDatetimeReportEnd), "small_1")
      sQuantity = str(len(dDBQueryRes))
      sQuantity = str(uTestCount)
      WriteLine("Quantidade: {}".format(sQuantity), "small_1")
      WriteLine("Detetados de novo: {}".format(str(uNewPositiveCount)), "small_1")
      WriteLine("Detetados (todos): {}".format(str(uPositiveCount)), "small_1")
      WriteLine("Não detetados: {}".format(str(uNegativeCount)), "small_1")
      WriteLine("Inconclusivos: {}".format(str(uInconclusiveCount)), "small_1")
      WriteLine("Indeterminados: {}".format(str(uErrorCount)), "small_1")
      MovePos(0.0, dConfigReport['page_section_sep'])
      if dDBQueryRes and uTestCount > 0:
        fXStartPos = GetPos()[0]
        fYStartPos = GetPos()[1]
        fXDisp = dConfigReport['page_char_advance']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Nome", "small_1_bold", dConfigReport['width_name'], dConfigReport['row_height'], "C", "LTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_name']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("DN", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_date']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Id.", "small_2_bold", dConfigReport['width_age'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_age']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("#INTID1#", "small_2_bold", dConfigReport['width_record'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_record']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("#STATEID1#", "small_2_bold", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_state_id_1']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Departmento", "small_2_bold", dConfigReport['width_department'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_department']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Colheita", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_date']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Resultado", "small_1_bold", dConfigReport['width_result'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_result']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("DH do resultado", "small_2_bold", dConfigReport['width_datetime'], dConfigReport['row_height'], "C", "RTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_datetime']
        SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
        for dRow in dDBQueryRes:
          if sClass not in dRow or not dRow[sClass] or dRow['Group'] != sCurrGroup or dRow['Result'].lower() == "waiting" or dRow['Result'].lower() == "notest":
            continue
          if dRow['Comments'] and dRow['Comments'].strip() != "" and dRow['Comments'].strip().lower() != "null":
            sComments = dRow['Comments'].strip()
          else:
            sComments = None
          if sComments and 'excluded_comments_search' in dConfig:
            for dXMLElement in dConfig['excluded_comments_search'][0]['search']:
              if re.match(sComments, dXMLElement['_text'], flags = re.IGNORECASE):
                sComments = None
                break
          fXStartPos = GetPos()[0]
          fYStartPos = GetPos()[1]
          if fYStartPos + dConfigReport['reserved_height_1'] > dConfigReport['max_y_pos']:
            AddPage()
            SetPos(0,0)
            WriteLine("#INSTITUTION# | #DEPARTMENT# | {}".format(sReportTitle), "header_1")
            MovePos(0.0, dConfigReport['page_header_sep'])
            WriteLine(sCurrGroup, "title_1")
            WriteLine(sClassTitle, "title_1")
            WriteLine("Data e hora: {}".format(sDatetimeReportCurr), "small_1")
            MovePos(0.0, dConfigReport['page_header_sep'])
            WriteLine("Análises concluídas", "section_1")
            WriteLine("Período: {} a {}".format(sDatetimeReportStart, sDatetimeReportEnd), "small_1")
            sQuantity = str(uTestCount)
            WriteLine("Quantidade: {} (continuação da página anterior)".format(sQuantity), "small_1")
            MovePos(0.0, dConfigReport['page_section_sep'])
            fXStartPos = GetPos()[0]
            fYStartPos = GetPos()[1]
            fXDisp = dConfigReport['page_char_advance']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("Nome", "small_1_bold", dConfigReport['width_name'], dConfigReport['row_height'], "C", "LTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_name']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("DN", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_date']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("Id.", "small_2_bold", dConfigReport['width_age'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_age']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("#INTID1#", "small_2_bold", dConfigReport['width_record'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_record']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("#STATEID1#", "small_2_bold", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_state_id_1']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("Departmento", "small_2_bold", dConfigReport['width_department'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_department']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("Colheita", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_date']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("Resultado", "small_1_bold", dConfigReport['width_result'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_result']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("DH do resultado", "small_2_bold", dConfigReport['width_datetime'], dConfigReport['row_height'], "C", "RTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_datetime']
            SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
            fXStartPos = GetPos()[0]
            fYStartPos = GetPos()[1]
          fXDisp = dConfigReport['page_char_advance']
          fXDispStart = fXDisp
          sName = str(dRow['Name']) if str(dRow['Name']).upper() != "NULL" else ""
          sBirthday = str(dRow['Birthday']) if str(dRow['Birthday']).upper() != "NULL" else ""
          sAge = str(dRow['Age']) if str(dRow['Age']).upper() != "NULL" else ""
          sRecord = str(dRow['Record']) if str(dRow['Record']).upper() != "NULL" else ""
          sStateID1 = str(dRow['StateID1']) if str(dRow['StateID1']).upper() != "NULL" else ""
          sDepartmentTitle = str(dRow['DepartmentTitle']) if str(dRow['DepartmentTitle']).upper() != "NULL" else ""
          sDepartmentTitle = str(dRow['TestDepartment']) if str(dRow['TestDepartment']).upper() != "NULL" else ""
          sSampleDate = str(dRow['SampleDate']) if str(dRow['SampleDate']).upper() != "NULL" else ""
          sResult = GetResult(str(dRow['Result']))
          sResultDateTime = str(dRow['ResultDateTime']) if str(dRow['ResultDateTime']).upper() != "NULL" else ""
          sAddress1 = str(dRow['Address1']) if str(dRow['Address1']).upper() != "NULL" else ""
          sAddress2 = str(dRow['Address2']) if str(dRow['Address2']).upper() != "NULL" else ""
          sGroup = str(dRow['Group']) if str(dRow['Group']).upper() != "NULL" else ""
          sDepartment = str(dRow['Department']) if str(dRow['Department']).upper() != "NULL" else ""
          sCategory = str(dRow['Category']) if str(dRow['Category']).upper() != "NULL" else ""
          sLocation = str(dRow['Location']) if str(dRow['Location']).upper() != "NULL" else ""
          SetPos(fXStartPos + fXDisp, fYStartPos)
          uNameWidth = GetStringWidth(sName, "small_1_bold")
          DrawCell(sName if uNameWidth < dConfigReport['width_name'] else sName.split()[0] + " " + sName.split()[1] + " " + sName.split()[-1], "small_1_bold", dConfigReport['width_name'], dConfigReport['row_height'], "L", "LT")
          fXDisp = fXDisp + dConfigReport['width_name']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          DrawCell(sBirthday, "small_2", dConfigReport['width_date'], dConfigReport['row_height'], "C", "T")
          fXDisp = fXDisp + dConfigReport['width_date']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          DrawCell(sAge, "small_1", dConfigReport['width_age'], dConfigReport['row_height'], "C", "T")
          fXDisp = fXDisp + dConfigReport['width_age']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          DrawCell(sRecord, "small_2", dConfigReport['width_record'], dConfigReport['row_height'], "C", "T")
          fXDisp = fXDisp + dConfigReport['width_record']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          DrawCell(sStateID1, "small_2", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "T")
          fXDisp = fXDisp + dConfigReport['width_state_id_1']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          uDepartmentWidth = GetStringWidth(dRow['DepartmentTitle'], "small_2")
          DrawCell(sDepartmentTitle if uDepartmentWidth < dConfigReport['width_department_tolerance'] else sDepartmentTitle[0:20], "small_2", dConfigReport['width_department'], dConfigReport['row_height'], "C", "T")
          fXDisp = fXDisp + dConfigReport['width_department']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          DrawCell(sSampleDate, "small_2", dConfigReport['width_date'], dConfigReport['row_height'], "C", "T")
          fXDisp = fXDisp + dConfigReport['width_date']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          if dRow['Result'].lower() == "positive":
            if dRow['PatientID'] in dHistory and 'old_positive' in dHistory[dRow['PatientID']] and dHistory[dRow['PatientID']]['old_positive']:
              bSelFill = True
              tuSelFillColor = dConfigReport['cell_attention_fill_color']
              sSelStyle = "small_1_bold"
            else:
              bSelFill = True
              tuSelFillColor = dConfigReport['cell_warning_fill_color']
              sSelStyle = "small_1_bold"
          elif dRow['Result'].lower() == "inconclusive":
            bSelFill = True
            tuSelFillColor = dConfigReport['cell_attention_fill_color']
            sSelStyle = "small_1"
          else:
            bSelFill = False
            tuSelFillColor = None
            sSelStyle = "small_1"
          if sComments:
            sResult = "{}*".format(sResult)
          DrawCell(sResult, sSelStyle, dConfigReport['width_result'], dConfigReport['row_height'], "C", "T", bFill = bSelFill, tuFillColor = tuSelFillColor)
          fXDisp = fXDisp + dConfigReport['width_result']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          DrawCell(sResultDateTime, "small_2", dConfigReport['width_datetime'], dConfigReport['row_height'], "C", "RT")
          fXDisp = fXDisp + dConfigReport['width_datetime']
          fXDispEnd = fXDisp
          fYStartPos = fYStartPos + dConfigReport['row_height'] * 1
          fXDisp = fXDispStart
          SetPos(fXStartPos + fXDispStart, fYStartPos)
          DrawCell(sClassTitle, "small_2", dConfigReport['width_caption_1'], dConfigReport['row_height'], "C", "L")
          fXDisp = fXDisp + dConfigReport['width_caption_1']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          uCellWidth = fXDispEnd - fXDisp
          if sComments:
            sComments = "Comentário: {}".format(sComments)
            uCommentsWidth = GetStringWidth(sComments, "small_1_bold_red")
            DrawCell(sComments if uCommentsWidth < uCellWidth - 5 else sComments[0:dConfigReport['tolerance_comments_2']], "small_1_bold_red", uCellWidth, dConfigReport['row_height'], "L", "R")
          else:
            sInfo = "Grupo: {} | Departamento: {} | Categoria: {} | Localização: {}".format(sGroup, sDepartment, sCategory, sLocation)
            uInfoWidth = GetStringWidth(sInfo, "small_1")
            DrawCell(sInfo if uInfoWidth < uCellWidth - 5 else sInfo[0:dConfigReport['tolerance_info']], "small_1", uCellWidth, dConfigReport['row_height'], "L", "R")
          fYStartPos = fYStartPos + dConfigReport['row_height'] * 1
          fXDisp = fXDispStart
          SetPos(fXStartPos + fXDispStart, fYStartPos)
          DrawCell("Histórico", "small_2", dConfigReport['width_caption_1'], dConfigReport['row_height'], "C", "LB")
          fXDisp = fXDisp + dConfigReport['width_caption_1']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          if dRow['PatientID'] in dHistory:
            for dEntry in dHistory[dRow['PatientID']]['results'][0:5]:
              sEntry = "{} ({})".format(dEntry['result'], dEntry['date'])
              if dEntry['result'].lower() == "detetado":
                bSelFill = True
                tuSelFillColor = dConfigReport['cell_warning_fill_color']
                sSelStyle = "small_2"
              elif dEntry['result'].lower() == "inconclusivo":
                bSelFill = True
                tuSelFillColor = dConfigReport['cell_attention_fill_color']
                sSelStyle = "small_2"
              else:
                bSelFill = True
                tuSelFillColor = dConfigReport['cell_normal_fill_color']
                sSelStyle = "small_2"
              DrawCell(sEntry, sSelStyle, dConfigReport['width_result_3'], dConfigReport['row_height'], "C", "B", bFill = bSelFill, tuFillColor = tuSelFillColor)
              fXDisp = fXDisp + dConfigReport['width_result_3']
              SetPos(fXStartPos + fXDisp, fYStartPos)
              DrawCell("", "small_1", dConfigReport['width_history_sep'], dConfigReport['row_height'], "L", "B")
              fXDisp = fXDisp + dConfigReport['width_history_sep']
              SetPos(fXStartPos + fXDisp, fYStartPos)
          uCellWidth = fXDispEnd - fXDisp
          DrawCell("", "small_1", uCellWidth, dConfigReport['row_height'], "L", "BR")
          SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
      uPendingCount = 0
      if dDBQueryRes and len(dDBQueryRes) > 0:
        for dRow in dDBQueryRes:
          if sClass not in dRow or not dRow[sClass] or dRow['Group'] != sCurrGroup:
            continue
          if dRow['Result'] == "waiting":
            uPendingCount = uPendingCount + 1
      if uPendingCount > 0:
        fXStartPos = GetPos()[0]
        fYStartPos = GetPos()[1]
        if fYStartPos + dConfigReport['reserved_height_2'] + dConfigReport['page_header_sep'] > dConfigReport['max_y_pos']:
          AddPage()
          SetPos(0,0)
          WriteLine("#INSTITUTION# | #DEPARTMENT# | {}".format(sReportTitle), "header_1")
          MovePos(0.0, dConfigReport['page_header_sep'])
          WriteLine(sCurrGroup, "title_1")
          WriteLine(sClassTitle, "title_1")
          WriteLine("Data e hora: {}".format(sDatetimeReportCurr), "small_1")
        MovePos(0.0, dConfigReport['page_header_sep'])
        WriteLine("Resultados pendentes", "section_1")
        sQuantity = str(uPendingCount)
        WriteLine("Quantidade: {}".format(sQuantity), "small_1")
        MovePos(0.0, dConfigReport['page_section_sep'])
        fXStartPos = GetPos()[0]
        fYStartPos = GetPos()[1]
        fXDisp = dConfigReport['page_char_advance']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Nome", "small_1_bold", dConfigReport['width_name'], dConfigReport['row_height'], "C", "LTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_name']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("DN", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_date']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Id.", "small_2_bold", dConfigReport['width_age'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_age']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("#INTID1#", "small_2_bold", dConfigReport['width_record'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_record']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("#STATEID1#", "small_2_bold", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_state_id_1']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Departmento", "small_2_bold", dConfigReport['width_department'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_department']
        SetPos(fXStartPos + fXDisp, fYStartPos)
        DrawCell("Colheita", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TBR", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
        fXDisp = fXDisp + dConfigReport['width_date']
        SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
        for dRow in dDBQueryRes:
          if sClass not in dRow or not dRow[sClass] or dRow['Group'] != sCurrGroup or dRow['Result'] != "waiting":
            continue
          fXStartPos = GetPos()[0]
          fYStartPos = GetPos()[1]
          if fYStartPos + dConfigReport['reserved_height_1'] > dConfigReport['max_y_pos']:
            AddPage()
            SetPos(0,0)
            WriteLine("#INSTITUTION# | #DEPARTMENT# | {}".format(sReportTitle), "header_1")
            MovePos(0.0, dConfigReport['page_header_sep'])
            WriteLine(sCurrGroup, "title_1")
            WriteLine(sClassTitle, "title_1")
            WriteLine("Data e hora: {}".format(sDatetimeReportCurr), "small_1")
            MovePos(0.0, dConfigReport['page_header_sep'])
            WriteLine("Resultados pendentes", "section_1")
            sQuantity = str(uPendingCount)
            WriteLine("Quantidade: {} (continuação da página anterior)".format(sQuantity), "small_1")
            MovePos(0.0, dConfigReport['page_section_sep'])
            fXStartPos = GetPos()[0]
            fYStartPos = GetPos()[1]
            fXDisp = dConfigReport['page_char_advance']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("Nome", "small_1_bold", dConfigReport['width_name'], dConfigReport['row_height'], "C", "LTB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_name']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("DN", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_date']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("Id.", "small_2_bold", dConfigReport['width_age'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_age']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("#INTID1#", "small_2_bold", dConfigReport['width_record'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_record']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("#STATEID1#", "small_2_bold", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_state_id_1']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("Departmento", "small_2_bold", dConfigReport['width_department'], dConfigReport['row_height'], "C", "TB", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_department']
            SetPos(fXStartPos + fXDisp, fYStartPos)
            DrawCell("Colheita", "small_2_bold", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TBR", bFill = True, tuFillColor = dConfigReport['row_header_fill_color'])
            fXDisp = fXDisp + dConfigReport['width_date']
            SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
            fXStartPos = GetPos()[0]
            fYStartPos = GetPos()[1]
          fXDisp = dConfigReport['page_char_advance']
          sName = str(dRow['Name']) if str(dRow['Name']).upper() != "NULL" else ""
          sBirthday = str(dRow['Birthday']) if str(dRow['Birthday']).upper() != "NULL" else ""
          sAge = str(dRow['Age']) if str(dRow['Age']).upper() != "NULL" else ""
          sRecord = str(dRow['Record']) if str(dRow['Record']).upper() != "NULL" else ""
          sStateID1 = str(dRow['StateID1']) if str(dRow['StateID1']).upper() != "NULL" else ""
          sDepartmentTitle = str(dRow['DepartmentTitle']) if str(dRow['DepartmentTitle']).upper() != "NULL" else ""
          sDepartmentTitle = str(dRow['TestDepartment']) if str(dRow['TestDepartment']).upper() != "NULL" else ""
          sSampleDate = str(dRow['SampleDate']) if str(dRow['SampleDate']).upper() != "NULL" else ""
          SetPos(fXStartPos + fXDisp, fYStartPos)
          uNameWidth = GetStringWidth(sName, "small_1_bold")
          DrawCell(sName if uNameWidth < dConfigReport['width_name'] else sName.split()[0] + " " + sName.split()[1] + " " + sName.split()[-1], "small_1_bold", dConfigReport['width_name'], dConfigReport['row_height'], "L", "LTB")
          fXDisp = fXDisp + dConfigReport['width_name']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          DrawCell(sBirthday, "small_2", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TB")
          fXDisp = fXDisp + dConfigReport['width_date']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          DrawCell(sAge, "small_1", dConfigReport['width_age'], dConfigReport['row_height'], "C", "TB")
          fXDisp = fXDisp + dConfigReport['width_age']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          DrawCell(sRecord, "small_2", dConfigReport['width_record'], dConfigReport['row_height'], "C", "TB")
          fXDisp = fXDisp + dConfigReport['width_record']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          DrawCell(sStateID1, "small_2", dConfigReport['width_state_id_1'], dConfigReport['row_height'], "C", "TB")
          fXDisp = fXDisp + dConfigReport['width_state_id_1']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          uDepartmentWidth = GetStringWidth(dRow['DepartmentTitle'], "small_2")
          DrawCell(sDepartmentTitle if uDepartmentWidth < dConfigReport['width_department_tolerance'] else sDepartmentTitle[0:20], "small_2", dConfigReport['width_department'], dConfigReport['row_height'], "C", "TB")
          fXDisp = fXDisp + dConfigReport['width_department']
          SetPos(fXStartPos + fXDisp, fYStartPos)
          DrawCell(sSampleDate, "small_2", dConfigReport['width_date'], dConfigReport['row_height'], "C", "TBR")
          fXDisp = fXDisp + dConfigReport['width_date']
          SetPos(fXStartPos, fYStartPos + dConfigReport['row_height'] * 1 + dConfigReport['row_disp'])
    uGroupIndex = uGroupIndex + 1
    oReport.output(sFilename)
    ldRes.append({
      'filename': sFilename,
      'filename_email': sFilenameTitle,
      'title': sCurrGroup
      })
  return ldRes

def GetSocialInstitutionCounts(sDatetimeStart, sDatetimeEnd, lsInstitutionClasses):
  lsVoid = ["", "NULL", "null", "VOID", "void", None]
  if sDatetimeStart:
    oDatetimeStart = datetime.datetime.strptime(sDatetimeStart, "%Y-%m-%d %H:%M:%S")
  else:
    oDatetimeStart = (datetime.datetime.now() - datetime.timedelta(days = 1)).replace(hour = 9, minute = 0, second = 0, microsecond = 0)
  if sDatetimeEnd:
    oDatetimeEnd = datetime.datetime.strptime(sDatetimeEnd, "%Y-%m-%d %H:%M:%S")
  else:
    oDatetimeEnd = datetime.datetime.now().replace(hour = 9, minute = 0, second = 0, microsecond = 0)
  sDatetimeDBStart = oDatetimeStart.strftime("%Y-%m-%d %H:%M:%S")
  sDatetimeDBEnd = oDatetimeEnd.strftime("%Y-%m-%d %H:%M:%S")
  sQuery = """
    SELECT
      A2.name AS Name,
      TIMESTAMPDIFF(YEAR, A2.birthday, CURDATE()) AS Age,
      A3.result_code AS LastResult,
      A3.method_code AS LastMethod,
      DATE_FORMAT(A3.sample_date, '%Y-%m-%d') AS LastSampleDate,
      DATE_FORMAT(A3.result_datetime, '%Y-%m-%d %H:%i:%S') AS LastResultDatetime,
      B2.result_code AS LastPositiveResult,
      B2.method_code AS LastPositiveMethod,
      DATE_FORMAT(B2.sample_date, '%Y-%m-%d') AS LastPositiveSampleDate,
      DATE_FORMAT(B2.result_datetime, '%Y-%m-%d %H:%i:%S') AS LastPositiveResultDatetime,
      C2.result_code AS LastNegative,
      C2.method_code AS LastNegativeMethod,
      DATE_FORMAT(C2.sample_date, '%Y-%m-%d') AS LastNegativeSampleDate,
      DATE_FORMAT(C2.result_datetime, '%Y-%m-%d %H:%i:%S') AS LastNegativeResultDatetime,
      D2."group" AS GroupTitle,
      D2.social_institution AS GroupSocialInstitution,
      D2.institution_class AS GroupInstitutionClass,
      D2.class AS GroupClass,
      D2.department AS GroupDepartment,
      D2.category AS GroupCategory,
      D2.location AS GroupLocation,
      DATE_FORMAT(D2."date", '%d/%m/%Y') AS GroupDate
    FROM
    (
      SELECT
        patient_id AS PatientID,
        MAX(sample_id_2) AS TestSample
        FROM tests
        WHERE
          test_code = 'sarscov2'
          AND result_code IS NOT NULL
          AND result_code != 'notest'
          AND result_code != ''
          AND status <> 2
          AND result_datetime >= '{0}' AND result_datetime <= '{1}'
        GROUP BY patient_id
    ) AS A
    LEFT JOIN
    (
      SELECT
        patient_id AS PatientID,
        MAX(sample_id_2) AS TestSample
      FROM tests
      WHERE
        test_code = 'sarscov2'
        AND result_code = 'positive'
        AND status <> 2
        AND result_datetime >= '{0}' AND result_datetime <= '{1}'
      GROUP BY patient_id
    ) AS B
      ON B.PatientID = A.PatientID
    LEFT JOIN
    (
      SELECT
        patient_id AS PatientID,
        MAX(sample_id_2) AS TestSample
      FROM tests
      WHERE
        test_code = 'sarscov2'
        AND result_code = 'negative'
        AND status <> 2
        AND result_datetime >= '{0}' AND result_datetime <= '{1}'
      GROUP BY patient_id
    ) AS C
      ON C.PatientID = A.PatientID
    LEFT JOIN patients AS A2
      ON A2.id = A.PatientID
    LEFT JOIN tests AS A3
      ON A3.sample_id_2 = A.TestSample
    LEFT JOIN tests AS B2
      ON B2.sample_id_2 = B.TestSample
    LEFT JOIN tests AS C2
      ON C2.sample_id_2 = C.TestSample
    LEFT JOIN
    (
      SELECT
        A.id AS TestID,
        MAX(B.id) AS GroupID
      FROM 
      (
        SELECT *
        FROM tests
        WHERE
          status <> 2
          AND result_datetime >= '{0}' AND result_datetime <= '{1}'
      ) AS A
      LEFT JOIN groups AS B
        ON
          (B.state_id = A.state_id_1)
          OR (
            (B.name LIKE CONCAT('%', REPLACE(A.name, ' ', '%'), '%')
            OR A.name LIKE CONCAT('%', REPLACE(B.name, ' ', '%'), '%'))
            AND B.birthday = A.birthday)
      GROUP BY A.id, B.id
    ) AS D
      ON D.TestID = A3.id
    LEFT JOIN groups AS D2
      ON D2.id = D.GroupID
    ORDER BY D2."group" ASC
    """.format(sDatetimeDBStart, sDatetimeDBEnd)
  dDBQueryRes = RunMySql(sQuery, True)
  dRes = dict()
  for dRow in dDBQueryRes:
    if dRow['GroupSocialInstitution'] == "1" or dRow['GroupInstitutionClass'] in lsInstitutionClasses:
      sGroup = dRow['GroupTitle'] if dRow['GroupTitle'] not in lsVoid else "unknown"
      sClass = dRow['GroupClass'] if dRow['GroupClass'] not in lsVoid else "unknown"
      sResult = dRow['LastResult'] if dRow['LastResult'] not in lsVoid else "error"
      sMethod = dRow['LastMethod'] if dRow['LastMethod'] not in lsVoid else "error"
      oLastResultDatetime = datetime.datetime.strptime(dRow['LastResultDatetime'], "%Y-%m-%d %H:%M:%S")
      if dRow['LastPositiveResult'] not in lsVoid:
        sResult = dRow['LastPositiveResult']
        sMethod = dRow['LastPositiveMethod']
      if sGroup not in dRes:
        dRes[sGroup] = dict()
      if "datetime" not in dRes[sGroup]:
        dRes[sGroup]["datetime"] = oLastResultDatetime
      else:
        if oLastResultDatetime < dRes[sGroup]["datetime"]:
          dRes[sGroup]["datetime"] = oLastResultDatetime
      if sClass not in dRes[sGroup]:
        dRes[sGroup][sClass] = dict()
      if sResult not in dRes[sGroup][sClass]:
        dRes[sGroup][sClass][sResult] = dict()
      if sMethod not in dRes[sGroup][sClass][sResult]:
        dRes[sGroup][sClass][sResult][sMethod] = 1
      else:
        dRes[sGroup][sClass][sResult][sMethod] = dRes[sGroup][sClass][sResult][sMethod] + 1
  if len(dRes) == 0:
    return None
  else:
    return dRes

def CreatePDFReport_Summary_1(sFolder, dData):
  global dConfig, dConfigReport, oReport, sDestinationFolder
  sStart = dData['start'].strftime(dConfig['summary_1'][0]['datetime_report'][0]['_text'])
  sEnd = dData['end'].strftime(dConfig['summary_1'][0]['datetime_report'][0]['_text'])
  sDate = dData['date'].strftime(dConfig['summary_1'][0]['date_report'][0]['_text'])
  sDatetimeFilenameCurr = datetime.datetime.now().strftime(dConfig['summary_1'][0]['datetime_filename_curr'][0]['_text'])
  sDatetimeReportCurr = datetime.datetime.now().strftime(dConfig['summary_1'][0]['datetime_report_curr'][0]['_text'])
  sFilename = ("{0}\\{1}\\" + dConfig['summary_1'][0]['filename_format'][0]['_text']).format(sDestinationFolder, sFolder, sDatetimeFilenameCurr)
  StartReport()
  AddPage()
  SetPos(0,0)
  WriteLine("#INSTITUTION# | #DEPARTMENT#", "header_1")
  WriteLine(dConfig['summary_1'][0]['title'][0]['_text'], "title_1")
  WriteLine("Período: {} a {}".format(sStart, sEnd), "small_1")
  WriteLine("Data e hora do ficheiro: {}".format(sDatetimeReportCurr), "small_2")
  MovePos(0.0, dConfigReport['page_header_sep_summary_1'])
  fXStartPos = GetPos()[0]
  fYStartPos = GetPos()[1]
  fXDisp = dConfigReport['page_char_advance']
  fYDisp = 0 
  SetPos(fXStartPos + fXDisp , fYStartPos + fYDisp)
  DrawCell("Laboratório/entidade", "small_1", dConfigReport['width_summary_1_label'], dConfigReport['row_height_summary_1'], "L", "LTB", bFill = False)
  fXDisp = fXDisp + dConfigReport['width_summary_1_label']
  SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
  DrawCell(dData['department'], "small_1_bold", dConfigReport['width_summary_1_content'], dConfigReport['row_height_summary_1'], "L", "TBR", bFill = False)
  fXDisp = dConfigReport['page_char_advance']
  fYDisp = fYDisp + dConfigReport['row_height_summary_1']
  SetPos(fXStartPos + fXDisp , fYStartPos + fYDisp)
  DrawCell("Contacto de email", "small_1", dConfigReport['width_summary_1_label'], dConfigReport['row_height_summary_1'], "L", "LTB", bFill = False)
  fXDisp = fXDisp + dConfigReport['width_summary_1_label']
  SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
  DrawCell(dData['email'], "small_1_bold", dConfigReport['width_summary_1_content'], dConfigReport['row_height_summary_1'], "L", "TBR", bFill = False)
  fXDisp = dConfigReport['page_char_advance']
  fYDisp = fYDisp + dConfigReport['row_height_summary_1']
  SetPos(fXStartPos + fXDisp , fYStartPos + fYDisp)
  DrawCell("Data a que se refere o reporte", "small_1", dConfigReport['width_summary_1_label'], dConfigReport['row_height_summary_1'], "L", "LTB", bFill = False)
  fXDisp = fXDisp + dConfigReport['width_summary_1_label']
  SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
  DrawCell(sDate, "small_1_bold", dConfigReport['width_summary_1_content'], dConfigReport['row_height_summary_1'], "L", "TBR", bFill = False)
  fXDisp = dConfigReport['page_char_advance']
  fYDisp = fYDisp + dConfigReport['row_height_summary_1']
  SetPos(fXStartPos + fXDisp , fYStartPos + fYDisp)
  DrawCell("N.º de amostras processadas por RT-PCR no dia de reporte", "small_1", dConfigReport['width_summary_1_label'], dConfigReport['row_height_summary_1'], "L", "LTB", bFill = False)
  fXDisp = fXDisp + dConfigReport['width_summary_1_label']
  SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
  DrawCell(str(dData['rtpcr_total']), "small_1_bold", dConfigReport['width_summary_1_content'], dConfigReport['row_height_summary_1'], "L", "TBR", bFill = False)
  fXDisp = dConfigReport['page_char_advance']
  fYDisp = fYDisp + dConfigReport['row_height_summary_1']
  SetPos(fXStartPos + fXDisp , fYStartPos + fYDisp)
  DrawCell("N.º de amostras positivas por RT-PCR no dia de reporte", "small_1", dConfigReport['width_summary_1_label'], dConfigReport['row_height_summary_1'], "L", "LTB", bFill = False)
  fXDisp = fXDisp + dConfigReport['width_summary_1_label']
  SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
  DrawCell(str(dData['rtpcr_pos']), "small_1_bold", dConfigReport['width_summary_1_content'], dConfigReport['row_height_summary_1'], "L", "TBR", bFill = False)
  fXDisp = dConfigReport['page_char_advance']
  fYDisp = fYDisp + dConfigReport['row_height_summary_1']
  SetPos(fXStartPos + fXDisp , fYStartPos + fYDisp)
  DrawCell("N.º de amostras processadas por TRAg no dia de reporte", "small_1", dConfigReport['width_summary_1_label'], dConfigReport['row_height_summary_1'], "L", "LTB", bFill = False)
  fXDisp = fXDisp + dConfigReport['width_summary_1_label']
  SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
  DrawCell(str(dData['ag_total']), "small_1_bold", dConfigReport['width_summary_1_content'], dConfigReport['row_height_summary_1'], "L", "TBR", bFill = False)
  fXDisp = dConfigReport['page_char_advance']
  fYDisp = fYDisp + dConfigReport['row_height_summary_1']
  SetPos(fXStartPos + fXDisp , fYStartPos + fYDisp)
  DrawCell("N.º de amostras positivas por TRAg no dia de reporte", "small_1", dConfigReport['width_summary_1_label'], dConfigReport['row_height_summary_1'], "L", "LTB", bFill = False)
  fXDisp = fXDisp + dConfigReport['width_summary_1_label']
  SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
  DrawCell(str(dData['ag_pos']), "small_1_bold", dConfigReport['width_summary_1_content'], dConfigReport['row_height_summary_1'], "L", "TBR", bFill = False)
  fXDisp = dConfigReport['page_char_advance']
  fYDisp = fYDisp + dConfigReport['row_height_summary_1']
  SetPos(fXStartPos + fXDisp , fYStartPos + fYDisp)
  DrawCell("Observações", "small_1", dConfigReport['width_summary_1_label'], dConfigReport['row_height_summary_1'], "L", "LTB", bFill = False)
  fXDisp = fXDisp + dConfigReport['width_summary_1_label']
  SetPos(fXStartPos + fXDisp, fYStartPos + fYDisp)
  DrawCell(dData['notes'], "small_2_bold", dConfigReport['width_summary_1_content'], dConfigReport['row_height_summary_1'], "L", "TBR", bFill = False)
  fXDisp = dConfigReport['page_char_advance']
  fYDisp = fYDisp + dConfigReport['row_height_summary_1']
  SetPos(fXStartPos + fXDisp , fYStartPos + fYDisp)
  oReport.output(sFilename)
  return sFilename

# ------------------
# XLS report methods
# ------------------

def CreateXLSReport_Results(sFolder, sDatetimeStart = None, sDatetimeEnd = None):
  global dConfig, dConfigReport, sDestinationFolder
  if sDatetimeStart and sDatetimeEnd:
    sDatetimeDBStart = datetime.datetime.strptime(sDatetimeStart, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d %H:%M:%S")
    sDatetimeDBEnd = datetime.datetime.strptime(sDatetimeEnd, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d %H:%M:%S")
    sDatetimeFilenameStart = datetime.datetime.strptime(sDatetimeStart, "%Y-%m-%d %H:%M:%S").strftime(dConfig['report_2'][0]['datetime_filename'][0])
    sDatetimeFilenameEnd = datetime.datetime.strptime(sDatetimeEnd, "%Y-%m-%d %H:%M:%S").strftime(dConfig['report_2'][0]['datetime_filename'][0]['_text'])
    sDatetimeFilenameEmailStart = datetime.datetime.strptime(sDatetimeStart, "%Y-%m-%d %H:%M:%S").strftime(dConfig['report_2'][0]['datetime_filename_email'][0]['_text'])
    sDatetimeFilenameEmailEnd = datetime.datetime.strptime(sDatetimeEnd, "%Y-%m-%d %H:%M:%S").strftime(dConfig['report_2'][0]['datetime_filename_email'][0]['_text'])
  else:
    sDatetimeDBStart = (datetime.datetime.now() - datetime.timedelta(days = 1)).strftime("%Y-%m-%d") + " 09:00"
    sDatetimeDBEnd = datetime.datetime.now().strftime("%Y-%m-%d") + " 09:00"
    sDatetimeFilenameStart = (datetime.datetime.now() - datetime.timedelta(days = 1)).replace(hour = 9, minute = 0, second = 0, microsecond = 0).strftime(dConfig['report_2'][0]['datetime_filename'][0]['_text'])
    sDatetimeFilenameEnd = datetime.datetime.now().replace(hour = 9, minute = 0, second = 0, microsecond = 0).strftime(dConfig['report_2'][0]['datetime_filename'][0]['_text'])
    sDatetimeFilenameEmailStart = (datetime.datetime.now() - datetime.timedelta(days = 1)).replace(hour = 9, minute = 0, second = 0, microsecond = 0).strftime(dConfig['report_2'][0]['datetime_filename_email'][0]['_text'])
    sDatetimeFilenameEmailEnd = datetime.datetime.now().replace(hour = 9, minute = 0, second = 0, microsecond = 0).strftime(dConfig['report_2'][0]['datetime_filename_email'][0]['_text'])
  sDatetimeFilenameCurr = datetime.datetime.now().strftime(dConfig['report_2'][0]['datetime_filename_curr'][0]['_text'])
  sDatetimeFilenameEmailCurr = datetime.datetime.now().strftime(dConfig['report_2'][0]['datetime_filename_email_curr'][0]['_text'])
  sXLSFilename = ("{0}\\{1}\\" + dConfig['report_2'][0]['filename_format'][0]['_text']).format(sDestinationFolder, sFolder, sDatetimeFilenameCurr, sDatetimeFilenameStart, sDatetimeFilenameEnd)
  sXLSFileTitle = dConfig['report_2'][0]['filename_email_format'][0]['_text'].format(sDatetimeFilenameEmailCurr)
  sQuery = """
    SELECT
      B.name AS Name,
      DATE_FORMAT(B.birthday, '%d/%m/%Y') AS Birthday,
      TIMESTAMPDIFF(YEAR, B.birthday, CURDATE()) AS Age,
      B.record AS Record,
      B.state_id_1 AS StateID1,
      B.patient_status AS Status,
      C.title AS DepartmentTitle,
      A.department AS TestDepartment,
      DATE_FORMAT(A.sample_date, '%d/%m/%Y') AS SampleDate,
      DATE_FORMAT(A.result_datetime, '%d/%m/%Y %H:%i:%s') AS ResultDateTime,
      A.result_code AS Result
    FROM tests AS A
    LEFT JOIN patients AS B
      ON B.id = A.patient_id
    LEFT JOIN departments AS C
      ON C.id = B.department
    WHERE
      A.status <> 2
      AND A.result_code != 'waiting' AND A.result_code != 'notest'
      AND A.result_datetime >= '{}' AND A.result_datetime <= '{}'
    ORDER BY FIELD(A.result_code, 'positive', 'inconclusive', 'error', 'negative', 'waiting', 'notest'), A.name
    """.format(sDatetimeDBStart, sDatetimeDBEnd);
  dDBQueryRes = RunMySql(sQuery, True)
  dStyle1 = xlwt.easyxf('font: name Calibri, height 220; align: horiz left, vert center')
  dStyle2 = xlwt.easyxf('font: name Calibri, height 220; align: horiz center, vert center')
  oWB = xlwt.Workbook()
  oWS = oWB.add_sheet('Resultados')
  oWS.write(0, 0, "Nome", dStyle1)
  oWS.write(0, 1, "Data de nascimento", dStyle2)
  oWS.write(0, 2, "#INTID1#", dStyle2)
  oWS.write(0, 3, "#STATEID1#", dStyle2)
  oWS.write(0, 4, "Departmento", dStyle2)
  oWS.write(0, 5, "Data de amostra", dStyle2)
  oWS.write(0, 6, "Resultado", dStyle2)
  oWS.write(0, 7, "Data e hora de resultado", dStyle2)
  oWS.col(0).width = dConfigReport['xls_width_0']
  oWS.col(1).width = dConfigReport['xls_width_1']
  oWS.col(2).width = dConfigReport['xls_width_2']
  oWS.col(3).width = dConfigReport['xls_width_3']
  oWS.col(4).width = dConfigReport['xls_width_4']
  oWS.col(5).width = dConfigReport['xls_width_5']
  oWS.col(6).width = dConfigReport['xls_width_6']
  oWS.col(7).width = dConfigReport['xls_width_7']
  uRow = 1
  if dDBQueryRes and len(dDBQueryRes) > 0:
    for dRow in dDBQueryRes:
      oWS.write(uRow, 0, dRow['Name'], dStyle1)
      oWS.write(uRow, 1, dRow['Birthday'], dStyle2)
      oWS.write(uRow, 2, GetText(dRow['Record']), dStyle2)
      oWS.write(uRow, 3, GetText(dRow['StateID1']), dStyle2)
      oWS.write(uRow, 4, GetText(dRow['TestDepartment']), dStyle2)
      oWS.write(uRow, 5, GetText(dRow['SampleDate']), dStyle2)
      oWS.write(uRow, 6, GetResult(dRow['Result']), dStyle2)
      oWS.write(uRow, 7, GetText(dRow['ResultDateTime']), dStyle2)
      uRow = uRow + 1
  oWB.save(sXLSFilename)
  return (sXLSFilename, sXLSFileTitle)

def CreateXLSReport_SocialInstitutions(sDateStart, sDateEnd, sSignature):
  global lsSocialInstitutionClasses
  global dConfig
  global sDestinationFolder
  global sDateFormat, sDateTimeFormat
  global sConfigFolder
  oDate = datetime.datetime.now()
  oDatetimeStart = datetime.datetime.strptime("{} 00:00:00".format(sDateStart), sDateTimeFormat)
  oDatetimeEnd = datetime.datetime.strptime("{} 23:59:59".format(sDateEnd), sDateTimeFormat)
  sDatetimeStart = oDatetimeStart.strftime(sDateTimeFormat)
  sDatetimeEnd = oDatetimeEnd.strftime(sDateTimeFormat)
  sFileDatetime = oDate.strftime(dConfig['report_7'][0]['datetime_file'][0]['_text'])
  sFileDatetimeStart = oDate.strftime(dConfig['report_7'][0]['datetime'][0]['_text'])
  sFileDatetimeEnd = oDate.strftime(dConfig['report_7'][0]['datetime'][0]['_text'])
  sFileDatetimeEmail = oDate.strftime(dConfig['report_7'][0]['datetime_file_email'][0]['_text'])
  sFileDatetimeStartEmail = oDate.strftime(dConfig['report_7'][0]['datetime_email'][0]['_text'])
  sFileDatetimeEndEmail = oDate.strftime(dConfig['report_7'][0]['datetime_email'][0]['_text'])
  sFolder = "Instituições sociais"
  uCountCutoff = int(dConfig['report_7'][0]['count_cutoff'][0]['_text'])
  dSocialInstitutionCounts = GetSocialInstitutionCounts(sDatetimeStart, sDatetimeEnd, lsSocialInstitutionClasses)
  sXLSFilename = ("{0}\\{1}\\" + dConfig['report_7'][0]['filename_format'][0]['_text']).format(sDestinationFolder, sFolder, sFileDatetime, sFileDatetimeStart, sFileDatetimeEnd, "xlsx")
  sXLSFilenameEmail = (dConfig['report_7'][0]['filename_format_email'][0]['_text']).format(sFileDatetimeEmail, sFileDatetimeStartEmail, sFileDatetimeEndEmail, "xlsx")
  sTextFilename = ("{0}\\{1}\\" + dConfig['report_7'][0]['filename_format'][0]['_text']).format(sDestinationFolder, sFolder, sFileDatetime, sFileDatetimeStart, sFileDatetimeEnd, "txt")
  oFile = open(sTextFilename, "w")
  for sGroup, dGroup in dSocialInstitutionCounts.items():
    uCollaboratorPositive = 0
    uCollaboratorNegative = 0
    uPatientPositive = 0
    uPatientNegative = 0
    if "collaborator" in dGroup:
      if "positive" in dGroup['collaborator']:
        for sMethod, uCount in dGroup['collaborator']['positive'].items():
          uCollaboratorPositive = uCollaboratorPositive + uCount
      if "negative" in dGroup['collaborator']:
        for sMethod, uCount in dGroup['collaborator']['negative'].items():
          uCollaboratorNegative = uCollaboratorNegative + uCount
    if "patient" in dGroup:
      if "positive" in dGroup['patient']:
        for sMethod, uCount in dGroup['patient']['positive'].items():
          uPatientPositive = uPatientPositive + uCount
      if "negative" in dGroup['patient']:
        for sMethod, uCount in dGroup['patient']['negative'].items():
          uPatientNegative = uPatientNegative + uCount
    oFile.write("{}, {}, {}, {}, {}\n".format(sGroup, str(uCollaboratorPositive), str(uCollaboratorNegative), str(uPatientPositive), str(uPatientNegative)))
  oFile.close()
  oWB = load_workbook("{}\\{}".format(sConfigFolder, "si_template.xlsx"))
  oWS = oWB['RASTREIO SARS CoV-2']
  oWS['B3'] = "#INSTITUTION#"
  oWS['B4'] = "DATA DE PREENCHIMENTO: {}".format(datetime.datetime.now().strftime("%Y-%m-%d"))
  oWS['B5'] = "NOME DE QUEM PREENCHEU: {}".format(sSignature)
  uRow = 9
  bRows = False
  for sGroup, dGroup in dSocialInstitutionCounts.items():
    dData = dict()
    dData['collaborator'] = dict()
    dData['patient'] = dict()
    dData['collaborator']['positive'] = dict()
    dData['collaborator']['negative'] = dict()
    dData['collaborator']['inconclusive'] = dict()
    dData['patient']['positive'] = dict()
    dData['patient']['negative'] = dict()
    dData['patient']['inconclusive'] = dict()
    dData['collaborator']['positive']['agicg'] = 0
    dData['collaborator']['negative']['agicg'] = 0
    dData['collaborator']['inconclusive']['agicg'] = 0
    dData['collaborator']['positive']['rtpcr'] = 0
    dData['collaborator']['negative']['rtpcr'] = 0
    dData['collaborator']['inconclusive']['rtpcr'] = 0
    dData['patient']['positive']['agicg'] = 0
    dData['patient']['negative']['agicg'] = 0
    dData['patient']['inconclusive']['agicg'] = 0
    dData['patient']['positive']['rtpcr'] = 0
    dData['patient']['negative']['rtpcr'] = 0
    dData['patient']['inconclusive']['rtpcr'] = 0
    for sKey1 in dData:
      if sKey1 in dGroup:
        for sKey2 in dData[sKey1]:
          if sKey2 in dGroup[sKey1]:
            for sKey3 in dData[sKey1][sKey2]:
              if sKey3 in dGroup[sKey1][sKey2]:
                dData[sKey1][sKey2][sKey3] = dGroup[sKey1][sKey2][sKey3]
    uCollaboratorPositive = dData['collaborator']['positive']['agicg'] + dData['collaborator']['positive']['rtpcr']
    uCollaboratorNegative = dData['collaborator']['negative']['agicg'] + dData['collaborator']['negative']['rtpcr']
    uCollaboratorInconclusive = dData['collaborator']['inconclusive']['agicg'] + dData['collaborator']['inconclusive']['rtpcr']
    uPatientPositive = dData['patient']['positive']['agicg'] + dData['patient']['positive']['rtpcr']
    uPatientNegative = dData['patient']['negative']['agicg'] + dData['patient']['negative']['rtpcr']
    uPatientInconclusive = dData['patient']['inconclusive']['agicg'] + dData['patient']['inconclusive']['rtpcr']
    uCount = uCollaboratorPositive + uCollaboratorNegative + uCollaboratorInconclusive + uPatientPositive + uPatientNegative + uPatientInconclusive
    if uCount >= uCountCutoff:
      oWS['A{}'.format(str(uRow))] = sGroup
      oWS['B{}'.format(str(uRow))] = dGroup['datetime'].strftime("%Y-%m-%d")
      oWS['C{}'.format(str(uRow))] = uCollaboratorPositive + uCollaboratorNegative + uCollaboratorInconclusive
      oWS['D{}'.format(str(uRow))] = uPatientPositive + uPatientNegative + uPatientInconclusive
      oWS['E{}'.format(str(uRow))] = dData['patient']['positive']['agicg'] + dData['collaborator']['positive']['agicg']
      oWS['F{}'.format(str(uRow))] = dData['patient']['positive']['rtpcr'] + dData['collaborator']['positive']['rtpcr']
      oWS['G{}'.format(str(uRow))] = dData['patient']['negative']['agicg'] + dData['collaborator']['negative']['agicg']
      oWS['H{}'.format(str(uRow))] = dData['patient']['negative']['rtpcr'] + dData['collaborator']['negative']['rtpcr']
      oWS['I{}'.format(str(uRow))] = dData['patient']['inconclusive']['agicg'] + dData['collaborator']['inconclusive']['agicg']
      oWS['J{}'.format(str(uRow))] = dData['patient']['inconclusive']['rtpcr'] + dData['collaborator']['inconclusive']['rtpcr']
      oWS['M{}'.format(str(uRow))] = "#INSTITUTION#"
      uRow = uRow + 1
      bRows = True
  oWB.save(sXLSFilename)
  return (sXLSFilename, sXLSFilenameEmail, sTextFilename) if bRows else None

# --------------------
# Email report methods
# --------------------

def EmailStockPrepareData(bEmailCancel = False):
  global dConfig, dEmailStockReportInputs, sConfigFolder, dCounts
  dData = dict()
  dData['date'] = datetime.datetime.now().strftime(dConfig['stock_email_date_format'][0]['_text'])
  print("")
  print("Dados para reporte de stocks:\n- C: Cancelar atualizacao de dados e envio\n- V: Voltar ao item anterior\n- R: Reiniciar")
  for dMethod in dConfig['methods'][0]['method']:
    uLastCount = dCounts[dMethod['id']]['_total'] if dMethod['id'] in dCounts else 0
    print("- Testes realizados por {}: {}".format(dMethod['_text'], str(uLastCount)))
  dOldData = listools.LoadDictSimple("{}\\stock_data.csv".format(sConfigFolder), "param", "value")
  bCancel = False
  bDataInput = True
  lsInputKeys = list(dEmailStockReportInputs.keys())
  uCurr = 0
  while bDataInput:
    bReset = False
    for sKey in lsInputKeys[uCurr:]:
      sCaption = RemoveAccents(dEmailStockReportInputs[sKey][0])
      bInput = True
      uNewValue = int(dOldData[sKey])
      if dEmailStockReportInputs[sKey][1] and dEmailStockReportInputs[sKey][1] in dCounts:
        uNewValue = uNewValue - round(dCounts[dEmailStockReportInputs[sKey][1]]['_total'] * dEmailStockReportInputs[sKey][2])
      while bInput:
        print("")
        sInput = input("- {}\n  ({} -> {}): ".format(sCaption, str(int(dOldData[sKey])), str(uNewValue))).strip()
        if sInput.lower() == "r":
          uCurr = 0
          bReset = True
          bInput = False
        elif sInput.lower() == "v":
          uCurr = uCurr - 1 if uCurr > 0 else 0
          bReset = True
          bInput = False
        elif sInput.lower() == "c":
          bReset = True
          bCancel = True
          bInput = False
        elif sInput.lower() == "":
          dData[sKey] = uNewValue
          if dData[sKey] < 0:
            dData[sKey]  = 0
          bInput = False
        else:
          bInput = False
          try:
            dData[sKey] = int(eval(sInput))
            if dData[sKey] < 0:
              raise
          except:
            print("- Erro: Introduza um valor numerico valido.")
            bInput = True
      if bReset or bCancel:
        break
      else:
        uCurr = uCurr + 1
    if not bReset or bCancel:
      bDataInput = False
  if bCancel:
    return None
  dData['email_1'] = 0
  if not bEmailCancel:
    bInput = True
    print("")
    while bInput:
      sInput = input("- Reportar stocks a farmacia (s/n): ").strip().lower()
      if sInput == "s":
        dData['email_1'] = 1
        bInput = False
      elif sInput == "n":
        dData['email_1'] = 0
        bInput = False
      else:
        continue
  dData['email_2'] = 0
  if not bEmailCancel:
    bInput = True
    while bInput:
      sInput = input("- Reportar stocks a entidade (s/n): ").strip().lower()
      if sInput == "s":
        dData['email_2'] = 1
        bInput = False
      elif sInput == "n":
        dData['email_2'] = 0
        bInput = False
      else:
        continue
  tDataTable = listools.GetTableFromDictSimple(dData, "param", "value")
  listools.SaveTable(tDataTable, "{}\\stock_data.csv".format(sConfigFolder))
  return dData

def EmailStockReport_1(sSender, lsTo, lsCc, sSignature, dData):
  global oSMTPServer, dConfig
  if not dData:
    return
  oMessage = email.mime.multipart.MIMEMultipart('alternative')
  oMessage["From"] = sSender
  oMessage["To"] = ",".join(lsTo)
  oMessage["Cc"] = ",".join(lsCc)
  oMessage["subject"] = "SARS-CoV-2, Reporte de stock, {}".format(dData['date'])
  sCounts = ""
  for dStock in dConfig['stock_references'][0]['stock']:
    if 'acquired' in dStock and dStock['acquired'].lower() == "true":
      sAdd = "{}{} -\n  - Unidades em stock: {};\n  - Pedidos de compra: {}.\n\n".format(
        dStock['_text'],
        " (referência {})".format(dStock['reference']) if 'reference' in dStock else "",
        str(dData["{}_stock".format(dStock['id'])]),
        str(dData["{}_requested".format(dStock['id'])])
        )
      sCounts = sCounts + sAdd
  sBody = "{}.\n\nEnvio informações relativas ao stock de consumíveis do #DEPARTMENT# para o diagnóstico da COVID-19.\n\nDia: {}.\n\n{}Agradeço a atenção dispensada.\n\nCom os melhores cumprimentos,\n\n{}.".format(
    GetGreeting(),
    dData['date'],
    sCounts,
    sSignature
    )
  sHTMLBody = """\
  <!DOCTYPE html>
  <html>
    <head>
      <meta charset="UTF-8">
      <style>
        p {{
          margin-bottom: 1em;
        }}
        body {{
          font-family:sans-serif;
          font-size: 1em;
        }}
        table {{
          border-collapse: collapse;
          margin-top: 2em;
          margin-bottom: 2em;
        }}
        td, th {{
          border-style: solid;
          border-width: 1px;
          padding: 5px;
          margin: 0px;
          vertical-align: middle;
          text-align: center;
        }}
      </style>
    </head>
    <body>
      <p>{}.</p>
      <p>Envio a tabela seguinte com informações relativas ao stock de consumíveis do #DEPARTMENT# para o diagnóstico da COVID-19.</p>
      <table>
        <thead>
          <tr>
            <th colspan = "3">Dia: {}</th>
          </tr>
          <tr>
            <th></th>
            <th>Unidades em stock</th>
            <th>Pedidos de compra</th>
          </tr>
        </thead>
        <tbody>
  """.format(
    GetGreeting(),
    dData['date'])
  for dStock in dConfig['stock_references'][0]['stock']:
    if 'acquired' in dStock and dStock['acquired'].lower() == "true":
      sRow = """\
              <tr>
                <td style = "text-align: left;">{}{}</td>
                <td>{}</td>
                <td>{}</td>
              </tr>
              """.format(
        dStock['_text'],
        " (referência {})".format(dStock['reference']) if 'reference' in dStock else "",
        dData["{}_stock".format(dStock['id'])],
        dData["{}_requested".format(dStock['id'])]
        )
      sHTMLBody = sHTMLBody + sRow
  sHTMLBody = sHTMLBody + """\
        </tbody>
      </table>
      <p>Agradeço a atenção dispensada.</p>
      <p>Com os melhores cumprimentos,</p>
      <p>{}.</p>
    </body>
  </html> 
  """.format(sSignature)
  oPart1 = email.mime.text.MIMEText(sBody, 'plain')
  oPart2 = email.mime.text.MIMEText(sHTMLBody, 'html')
  oMessage.attach(oPart1)
  oMessage.attach(oPart2)
  sMessage = oMessage.as_string()
  oSMTPServer.sendmail(sSender, lsTo + lsCc, sMessage)

def EmailStockReport_2(sSender, lsTo, lsCc, sSignature, dData):
  global oSMTPServer, dConfig
  if not dData:
    return
  oMessage = email.mime.multipart.MIMEMultipart('alternative')
  oMessage["From"] = sSender
  oMessage["To"] = ",".join(lsTo)
  oMessage["Cc"] = ",".join(lsCc)
  oMessage["subject"] = "Avaliação da capacitação para testes COVID-19, {}".format(dData['date'])
  sBody = "{}, {}.\n\nEnviamos informações relativas ao stock de consumíveis e à capacitação para o diagnóstico da COVID-19.\n\nDia: {}.\n\n1. Stock de reagentes de extração: {} testes;\n2. Stock de reagentes de amplificação: {} testes;\n3. Capacitação: {} testes/dia.\n\nAgradeço a atenção dispensada.\n\nCom os melhores cumprimentos,\n\n{}.".format(
    GetGreeting(),
    dConfig['report_6'][0]['recipient'][0]['_text'],
    dData['date'],
    str(dData['auto_extraction_stock']),
    str(dData['pcr_reactions_stock'] + dData['pcr_closed_stock'] + dData['pcr_panel_stock']),
    str(dData['capacity']),
    sSignature
    )
  sHTMLBody = """\
  <!DOCTYPE html>
  <html>
    <head>
      <meta charset="UTF-8">
      <style>
        p {{
          margin-bottom: 1em;
        }}
        body {{
          font-family:sans-serif;
          font-size: 1em;
        }}
        table {{
          border-collapse: collapse;
          margin-top: 2em;
          margin-bottom: 2em;
        }}
        td, th {{
          border-style: solid;
          border-width: 1px;
          padding: 5px;
          margin: 0px;
          text-align: center;
          vertical-align: middle;
        }}
      </style>
    </head>
    <body>
      <p>{}, {}.</p>
      <p>Enviamos a tabela seguinte com informações relativas ao stock de consumíveis e à capacitação para o diagnóstico da COVID-19.</p>
      <table>
        <thead>
          <tr>
            <th colspan = "3">Laboratório: {}</th>
          </tr>
          <tr>
            <th colspan = "3">Dia: {}</th>
          </tr>
          <tr>
            <th>Capacidade máxima atual<br>(n.º máximo de testes/dia)</th>
            <th>Stock atual de reagentes de extração<br>(n.º de testes)</th>
            <th>Stock atual de reagentes de amplificação<br>(n.º de testes)</th>
          </tr>
        </thead>
        <tbody>
          <tr>
            <td>{}</td>
            <td>{}</td>
            <td>{}</td>
          </tr>
        </tbody>
      </table>
      <p>Agradeço a atenção dispensada.</p>
      <p>Com os melhores cumprimentos,</p>
      <p>{}.</p>
    </body>
  </html> 
  """.format(
    GetGreeting(),
    dConfig['report_6'][0]['recipient'][0]['_text'],
    "#INSTITUTION#",
    dData['date'],
    str(dData['capacity']),
    str(dData['auto_extraction_stock']),
    str(dData['pcr_reactions_stock'] + dData['pcr_closed_stock'] + dData['pcr_panel_stock']),
    sSignature
  )
  oPart1 = email.mime.text.MIMEText(sBody, 'plain')
  oPart2 = email.mime.text.MIMEText(sHTMLBody, 'html')
  oMessage.attach(oPart1)
  oMessage.attach(oPart2)
  sMessage = oMessage.as_string()
  oSMTPServer.sendmail(sSender, lsTo + lsCc, sMessage)

# ------------
# Main routine
# ------------

bExit = False

try:
  LoadConfigXML()
except Exception as dError:
  Log("")
  Log("Erro: Nao foi possivel carregar o ficheiro de configuracao XML.")
  Log("Mensagem de erro:\n" + str(dError))
  Log("")
  bExit = True
if bExit:
  LogSave()
  sys.exit()

#UpdateDBHCWorkers()
#UpdateDBPatientsInst()
#sys.exit()

print("")
Log("#INSTITUTION# | #DEPARTMENT#")
Log("Ferramenta para relatorios da plataforma COVID-DX-DB")
Log("----------------------------------------------------")
Log("")
Log("Data e hora: " + datetime.datetime.now().strftime(sTimestampFormat))

bProcess = True
bCancel = False
bEmail = False
print("")
while bProcess:
  sSignature = input("Nome ou assinatura do remetente (\"C\" para cancelar o envio por email):\n  ").strip()
  if sSignature.lower() == "c":
    bCancel = True
    bProcess = False
  elif sSignature.lower() == "":
    continue
  else:
    bProcess = False

try:
  MySqlCursorStart()
except Exception as dError:
  Log("")
  Log("Erro: Nao foi possivel ligar ao servidor de BD.")
  Log("Mensagem de erro:\n" + str(dError))
  Log("")
  bExit = True
if bExit:
  LogSave()
  sys.exit()

uHour = datetime.datetime.now().hour

print("")
bInput = True
bYes = False
while bInput:
  sInput = input("Criar relatorio de resultados em PDF (s/n)? ").strip().lower()
  if sInput not in ["s", "n"]:
    continue
  if sInput == "s":
    bYes = True
  bInput = False
bCreate = False
bReset = False
if bYes:
  print("Dados para o relatorio com resultados em PDF:")
  print("- Indique a data e a hora no formato 'aaaa-mm-dd hh:mm'")
  print("- Pressione <ENTER> sem preencher para usar o modo automatico")
  print("- Indique 'c' para cancelar e 'r' para reiniciar")
  bDataInput = True
  while bDataInput:
    bReset = False
    bProcess = True
    while bProcess:
      sDatetimeStartInput = input("- Data e hora inicial: ").strip()
      if sDatetimeStartInput == "":
        uHour = datetime.datetime.now().hour
        if uHour >= 12:
          oDatetimeStartReport1 = datetime.datetime.now().replace(hour = 9, minute = 0, second = 0, microsecond = 0)
        else:
          oDatetimeStartReport1 = (datetime.datetime.now() - datetime.timedelta(days = 1)).replace(hour = 9, minute = 0, second = 0, microsecond = 0)
        bProcess = False
      elif sDatetimeStartInput.lower() == "c":
        bDataInput = False
        bProcess = False
      elif sDatetimeStartInput.lower() == "r":
        bReset = True
        bProcess = False
      else:
        try:
          oDatetimeStartReport1 = datetime.datetime.strptime(sDatetimeStartInput, "%Y-%m-%d %H:%M")
        except:
          print("- Erro: Indique a data e hora num formato valido.")
          bProcess = True
        else:
          bProcess = False
    if not bDataInput:
      break
    if bReset:
      continue
    bProcess = True
    while bProcess:
      sDatetimeEndInput = input("- Data e hora final: ").strip()
      if sDatetimeEndInput == "":
        uHour = datetime.datetime.now().hour
        if uHour < 12 and uHour >= 9:
          oDatetimeEndReport1 = datetime.datetime.now().replace(hour = 9, minute = 0, second = 0, microsecond = 0)
        else:
          oDatetimeEndReport1 = datetime.datetime.now()
        bProcess = False
      elif sDatetimeEndInput.lower() == "c":
        bDataInput = False
        bProcess = False
      elif sDatetimeEndInput.lower() == "r":
        bReset = True
        bProcess = False
      else:
        try:
          oDatetimeEndReport1 = datetime.datetime.strptime(sDatetimeEndInput, "%Y-%m-%d %H:%M")
        except:
          print("- Erro: Indique a data e hora num formato valido.")
          bProcess = True
        else:
          bProcess = False
    if not bDataInput:
      break
    if bReset:
      continue
    if (oDatetimeEndReport1 - oDatetimeStartReport1).days < 0:
      print("- Erro: A data inicial tem que ser anterior a data final.")
      bDataInput = True
    else:
      bDataInput = False
      bCreate = True
if bCreate:
  if not bCancel:
    for dGroup in dConfig['report_1'][0]['group']:
      bProcess = True
      while bProcess:
        sInput = input("- Enviar email ao grupo {} (s/n): ".format(dGroup['caption'])).strip()
        if sInput.lower() == "s":
          dGroup['email'] = True
          bEmail = True
          bProcess = False
        elif sInput.lower() == "n":
          dGroup['email'] = False
          bProcess = False
        else:
          bProcess = True
  sDatetimeStartReport1 = oDatetimeStartReport1.strftime(dConfig['report_1'][0]['datetime_email'][0]['_text'])
  sDatetimeEndReport1 = oDatetimeEndReport1.strftime(dConfig['report_1'][0]['datetime_email'][0]['_text'])
  for dGroup in dConfig['report_1'][0]['group']:
    try:
      Log("")
      Log("A criar PDF com resultados...\n- Data e hora inicial: {}\n- Data e hora final: {}\n- Grupo: {}".format(sDatetimeStartReport1, sDatetimeEndReport1, dGroup['caption']))
      lsFilter = []
      for dFilter in dGroup['filter_department']:
        lsFilter.append(dFilter['_text'])
      tsPDF_1_Filenames = CreatePDFReport_Results(
        "results",
        oDatetimeStartReport1.strftime("%Y-%m-%d %H:%M:%S"),
        oDatetimeEndReport1.strftime("%Y-%m-%d %H:%M:%S"),
        dGroup['caption'],
        lsFilter
        )
      dGroup['filenames'] = tsPDF_1_Filenames
    except Exception as dError:
      Log("")
      Log("Erro: Nao foi possivel criar o ficheiro PDF.")
      Log("Mensagem de erro:\n" + str(dError))
      Log("")
      bExit = True
    if bExit:
      MySqlCursorClose()
      LogSave()
      sys.exit()

print("")
bInput = True
bYes = False
while bInput:
  sInput = input("Criar relatorios em PDF com resultados por grupos (s/n)? ").strip().lower()
  if sInput not in ["s", "n"]:
    continue
  if sInput == "s":
    bYes = True
  bInput = False
bCreate = False
bReset = False
if bYes:
  print("Dados para os relatorios em PDF com resultados por grupos:")
  print("- Indique a data e a hora no formato 'aaaa-mm-dd hh:mm'")
  print("- Pressione <ENTER> sem preencher para usar o modo automatico")
  print("- Indique 'c' para cancelar e 'r' para reiniciar")
  bDataInput = True
  while bDataInput:
    bReset = False
    bProcess = True
    while bProcess:
      sDatetimeStartInput = input("- Data e hora inicial: ").strip()
      if sDatetimeStartInput == "":
        uHour = datetime.datetime.now().hour
        if uHour >= 12:
          oDatetimeStartReport3 = datetime.datetime.now().replace(hour = 9, minute = 0, second = 0, microsecond = 0)
        else:
          oDatetimeStartReport3 = (datetime.datetime.now() - datetime.timedelta(days = 1)).replace(hour = 9, minute = 0, second = 0, microsecond = 0)
        bProcess = False
      elif sDatetimeStartInput.lower() == "c":
        bProcess = False
        bDataInput = False
      elif sDatetimeStartInput.lower() == "r":
        bProcess = False
        bReset = True
      else:
        try:
          oDatetimeStartReport3 = datetime.datetime.strptime(sDatetimeStartInput, "%Y-%m-%d %H:%M")
        except:
          print("- Erro: Indique a data e hora num formato valido.")
          bProcess = True
        else:
          bProcess = False
    if not bDataInput:
      break
    if bReset:
      continue
    bProcess = True
    while bProcess:
      sDatetimeEndInput = input("- Data e hora final: ").strip()
      if sDatetimeEndInput == "":
        uHour = datetime.datetime.now().hour
        if uHour < 12 and uHour >= 9:
          oDatetimeEndReport3 = datetime.datetime.now().replace(hour = 9, minute = 0, second = 0, microsecond = 0)
        else:
          oDatetimeEndReport3 = datetime.datetime.now()
        bProcess = False
      elif sDatetimeEndInput.lower() == "c":
        bProcess = False
        bDataInput = False
      elif sDatetimeEndInput.lower() == "r":
        bProcess = False
        bReset = True
      else:
        try:
          oDatetimeEndReport3 = datetime.datetime.strptime(sDatetimeEndInput, "%Y-%m-%d %H:%M")
        except:
          print("- Erro: Indique a data e hora num formato valido.")
          bProcess = True
        else:
          bProcess = False
    if not bDataInput:
      break
    if bReset:
      continue
    if (oDatetimeEndReport3 - oDatetimeStartReport3).days < 0:
      print("- Erro: A data inicial tem que ser anterior a data final.")
      bDataInput = True
    else:
      bDataInput = False
      bCreate = True
if bCreate:
  sDatetimeStartReport3 = oDatetimeStartReport3.strftime(dConfig['report_4'][0]['datetime_email'][0]['_text'])
  sDatetimeEndReport3 = oDatetimeEndReport3.strftime(dConfig['report_4'][0]['datetime_email'][0]['_text'])
  if not bCancel:
    for dGroup in dConfig['report_4'][0]['group']:
      bProcess = True
      while bProcess:
        sInput = input("- Enviar email ao grupo {} (s/n): ".format(dGroup['caption'])).strip()
        if sInput.lower() == "s":
          dGroup['email'] = True
          bEmail = True
          bProcess = False
        elif sInput.lower() == "n":
          dGroup['email'] = False
          bProcess = False
        else:
          bProcess = True
  try:
    Log("")
    Log("A criar PDF com resultados por grupos...\n- Data e hora inicial: {}\n- Data e hora final: {}".format(sDatetimeStartReport3, sDatetimeEndReport3))
    sDateTimeFolder = datetime.datetime.now().strftime("%Y-%m-%d %H_%M_%S")
    os.mkdir("{}\\groups\\{}".format(sDestinationFolder, sDateTimeFolder))
    tsPDF_4_Filenames = CreatePDFReport_ResultsGroups(
      "groups\\{}".format(sDateTimeFolder),
      oDatetimeStartReport3.strftime("%Y-%m-%d %H:%M:%S"),
      oDatetimeEndReport3.strftime("%Y-%m-%d %H:%M:%S"),
      ["collaborator", "patient", "internal", "contact"],
      ["Colaboradores", "Utentes", "Internos", "Contactos"],
      "Resultados por grupos para SARS-CoV-2"
      )
  except Exception as dError:
    Log("")
    Log("Erro: Nao foi possivel criar o ficheiro PDF.")
    Log("Mensagem de erro:\n" + str(dError))
    Log("")
    bExit = True
  if bExit:
    MySqlCursorClose()
    LogSave()
    sys.exit()
  for dGroup in dConfig['report_4'][0]['group']:
    dGroup['filenames_1'] = tsPDF_4_Filenames

try:
  Log("")
  Log("A criar PDF com resultados do dia anterior, das 00:00 as 24:00...")
  sDatetimeStart = (datetime.datetime.now() - datetime.timedelta(days = 1)).strftime("%Y-%m-%d") + " 00:00:00"
  sDatetimeEnd = (datetime.datetime.now() - datetime.timedelta(days = 1)).strftime("%Y-%m-%d") + " 23:59:59"
  tsPDF_2_Filenames = CreatePDFReport_Results("#INSTITUTION#", sDatetimeStart, sDatetimeEnd, "#INSTITUTION#", None)
  dSummaryData = dict()
  dSummaryData['start'] = datetime.datetime.strptime(sDatetimeStart, "%Y-%m-%d %H:%M:%S")
  dSummaryData['end'] = datetime.datetime.strptime(sDatetimeEnd, "%Y-%m-%d %H:%M:%S")
  dSummaryData['date'] = datetime.datetime.strptime(sDatetimeStart, "%Y-%m-%d %H:%M:%S")
  dSummaryData['department'] = dConfig['summary_1'][0]['department'][0]['_text']
  dSummaryData['email'] = dConfig['summary_1'][0]['email'][0]['_text']
  dSummaryData['rtpcr_total'] = dCounts['rtpcr']['_total'] if 'rtpcr' in dCounts and '_total' in dCounts['rtpcr'] else 0
  dSummaryData['rtpcr_pos'] = dCounts['rtpcr']['positive'] if 'rtpcr' in dCounts and 'positive' in dCounts['rtpcr'] else 0
  dSummaryData['ag_total'] = dCounts['agicg']['_total'] if 'agicg' in dCounts and '_total' in dCounts['agicg'] else 0
  dSummaryData['ag_pos'] = dCounts['agicg']['positive'] if 'agicg' in dCounts and 'positive' in dCounts['agicg'] else 0
  if dCounts['_new_positive'] == 1:
    dSummaryData['notes'] = "Das amostras positivas, uma era de um utente que testou positivo de novo neste laboratório."
  elif dCounts['_new_positive'] > 1:
    dSummaryData['notes'] = "Das amostras positivas, {} eram de utentes que testaram positivo de novo neste laboratório.".format(str(dCounts['_new_positive']))
  else:
    dSummaryData['notes'] = ""
except Exception as dError:
  Log("")
  Log("Erro: Nao foi possivel criar o ficheiro PDF.")
  Log("Mensagem de erro:\n" + str(dError))
  Log("")
  bExit = True
if bExit:
  MySqlCursorClose()
  LogSave()
  sys.exit()

print("")
bInput = True
bYes = False
while bInput:
  sInput = input("Criar relatorio em PDF de utentes COVID-19 (s/n)? ").strip().lower()
  if sInput not in ["s", "n"]:
    continue
  if sInput == "s":
    bYes = True
  bInput = False
if bYes:
  if not bCancel:
    print("Dados para o relatorio PDF de utentes COVID-19:")
    for dGroup in dConfig['report_3'][0]['group']:
      bProcess = True
      while bProcess:
        sInput = input("- Enviar email ao grupo {} (s/n): ".format(dGroup['caption'])).strip()
        if sInput.lower() == "s":
          dGroup['email'] = True
          bEmail = True
          bProcess = False
        elif sInput.lower() == "n":
          dGroup['email'] = False
          bProcess = False
        else:
          bProcess = True
  try:
    Log("")
    Log("A criar relatorio PDF de utentes COVID-19...")
    tsPDF_3_Filenames = CreatePDFReport_COVIDPatients("covid19_patients", None, None)
  except Exception as dError:
    Log("")
    Log("Erro: Nao foi possivel criar o ficheiro PDF.")
    Log("Mensagem de erro:\n" + str(dError))
    Log("")
    bExit = True
  if bExit:
    MySqlCursorClose()
    LogSave()
    sys.exit()
  for dGroup in dConfig['report_3'][0]['group']:
    dGroup['filenames'] = tsPDF_3_Filenames

print("")
bInput = True
bYes = False
while bInput:
  sInput = input("Criar ficheiro XLS com resultados (s/n)? ").strip().lower()
  if sInput not in ["s", "n"]:
    continue
  if sInput == "s":
    bYes = True
  bInput = False
if bYes:
  if not bCancel:
    print("Dados para o ficheiro XLS com resultados:")
    for dGroup in dConfig['report_2'][0]['group']:
      bProcess = True
      while bProcess:
        sInput = input("- Enviar email ao grupo {} (s/n): ".format(dGroup['caption'])).strip()
        if sInput.lower() == "s":
          dGroup['email'] = True
          bEmail = True
          bProcess = False
        elif sInput.lower() == "n":
          dGroup['email'] = False
          bProcess = False
        else:
          bProcess = True
  oDatetimeStartReport2 = (datetime.datetime.now() - datetime.timedelta(days = 1)).replace(hour = 9, minute = 0, second = 0, microsecond = 0)
  oDatetimeEndReport2 = datetime.datetime.now().replace(hour = 9, minute = 0, second = 0, microsecond = 0)
  sDatetimeStartReport2 = oDatetimeStartReport2.strftime(dConfig['report_2'][0]['datetime_email'][0]['_text'])
  sDatetimeEndReport2 = oDatetimeEndReport2.strftime(dConfig['report_2'][0]['datetime_email'][0]['_text'])
  try:
    Log("")
    Log("A criar ficheiro XLS com resultados...\n- Inicio: dia anterior, as 09:00\n- Fim: dia atual, as 09:00")
    tsXLS_1_Filenames = CreateXLSReport_Results("results", None, None)
  except Exception as dError:
    Log("")
    Log("Erro: Nao foi possivel criar o ficheiro XLS.")
    Log("Mensagem de erro:\n" + str(dError))
    Log("")
    bExit = True
  if bExit:
    MySqlCursorClose()
    LogSave()
    sys.exit()
  for dGroup in dConfig['report_2'][0]['group']:
    dGroup['filenames'] = tsXLS_1_Filenames

print("")
bInput = True
bYes = False
while bInput:
  sInput = input("Criar resumo de testes em instituicoes sociais (s/n)? ").strip().lower()
  if sInput not in ["s", "n"]:
    continue
  if sInput == "s":
    bYes = True
  bInput = False
bCreate = False
bReset = False
if bYes:
  print("Dados para o resumo de testes em instituicoes sociais:")
  print("- Indique a data no formato 'aaaa-mm-dd'")
  print("- Pressione <ENTER> sem preencher para usar o modo automatico")
  print("- Indique 'c' para cancelar e 'r' para reiniciar")
  bDataInput = True
  while bDataInput:
    bReset = False
    bProcess = True
    bAutoMode = False
    while bProcess:
      sDatetimeStartInput = input("- Data inicial: ").strip()
      if sDatetimeStartInput == "":
        oDatetimeCurr = datetime.datetime.today() - datetime.timedelta(days = 3)
        bFound = False
        while not bFound:
          uWeekDay = oDatetimeCurr.weekday()
          if uWeekDay == 0:
            bFound = True
          else:
            oDatetimeCurr = oDatetimeCurr - datetime.timedelta(days = 1)
        oDatetimeStartReport7 = oDatetimeCurr
        oDatetimeEndReport7 = oDatetimeCurr + datetime.timedelta(days = 6)
        bAutoMode = True
        bProcess = False
      elif sDatetimeStartInput.lower() == "c":
        bProcess = False
        bDataInput = False
      elif sDatetimeStartInput.lower() == "r":
        bProcess = False
        bReset = True
      else:
        try:
          oDatetimeStartReport7 = datetime.datetime.strptime(sDatetimeStartInput, "%Y-%m-%d")
        except:
          print("- Erro: Indique a data num formato valido.")
          bProcess = True
        else:
          bProcess = False
    if not bDataInput:
      break
    if bReset:
      continue
    bProcess = True
    if bAutoMode:
      bProcess = False
    while bProcess:
      sDatetimeEndInput = input("- Data final: ").strip()
      if sDatetimeEndInput == "":
        bProcess = False
        bDataInput = False
      elif sDatetimeEndInput.lower() == "c":
        bProcess = False
        bDataInput = False
      elif sDatetimeEndInput.lower() == "r":
        bProcess = False
        bReset = True
      else:
        try:
          oDatetimeEndReport7 = datetime.datetime.strptime(sDatetimeEndInput, "%Y-%m-%d")
        except:
          print("- Erro: Indique a data num formato valido.")
          bProcess = True
        else:
          bProcess = False
    if not bDataInput:
      break
    if bReset:
      continue
    if (oDatetimeEndReport7 - oDatetimeStartReport7).days < 0:
      print("- Erro: A data inicial tem que ser anterior a data final.")
      bDataInput = True
    else:
      bDataInput = False
      bCreate = True
    if bCancel and not bDataInput:
      bDataInput = True
      bCreate = False
      bProcess = True
      while bProcess:
        sSignature = input("- Nome ou assinatura do remetente: ").strip()
        if sSignature.lower() == "c":
          bDataInput = False
          bProcess = False
        elif sSignature.lower() == "r":
          bProcess = False
          bReset = True
        elif sSignature.lower() == "":
          continue
        else:
          bProcess = False
          bDataInput = False
          bCreate = True
      if not bDataInput:
        break
      if bReset:
        continue
if bCreate:
  sDatetimeStartReport7 = oDatetimeStartReport7.strftime(sDateFormat)
  sDatetimeEndReport7 = oDatetimeEndReport7.strftime(sDateFormat)
  dDatetimeReport7 = datetime.datetime.now()
  if not bCancel:
    for dGroup in dConfig['report_7'][0]['group']:
      bProcess = True
      while bProcess:
        sInput = input("- Enviar email ao grupo {} (s/n): ".format(dGroup['caption'])).strip()
        if sInput.lower() == "s":
          dGroup['email'] = True
          bEmail = True
          bProcess = False
        elif sInput.lower() == "n":
          dGroup['email'] = False
          bProcess = False
        else:
          bProcess = True
  try:
    Log("")
    Log("A criar resumo de testes em instituicoes sociais...\n- Data inicial: {}\n- Data final: {}".format(sDatetimeStartReport7, sDatetimeEndReport7))
    tsReport7Filenames = CreateXLSReport_SocialInstitutions(sDatetimeStartReport7, sDatetimeEndReport7, sSignature)
  except Exception as dError:
    Log("")
    Log("Erro: Nao foi possivel criar o resumos de testes em instituicoes sociais.")
    Log("Mensagem de erro:\n" + str(dError))
    Log("")
    bExit = True
  if bExit:
    MySqlCursorClose()
    LogSave()
    sys.exit()
  for dGroup in dConfig['report_7'][0]['group']:
    dGroup['filenames'] = tsReport7Filenames
    if not tsReport7Filenames:
      dGroup['email'] = False

MySqlCursorClose()

dStockData = None
print("")
bInput = True
bYes = False
while bInput:
  sInput = input("Atualizar dados dos stocks (s/n)? ").strip().lower()
  if sInput not in ["s", "n"]:
    continue
  if sInput == "s":
    bYes = True
  bInput = False
if bYes:
  dStockData = EmailStockPrepareData(bCancel)
if dStockData and (dStockData['email_1'] or dStockData['email_2']):
  bEmail = True
if not bCancel and bEmail:
  try:
    Log("")
    Log("A ligar ao servidor SMTP de email...")
    EmailLogin(sSMTPServer, uSMTPPort, sSMTPSender, sSMTPPassword)
  except Exception as dError:
    Log("")
    Log("Erro: Nao foi possivel ligar ao servidor SMTP de email.")
    Log("Mensagem de erro:\n" + str(dError))
    Log("")
    bExit = True
  if bExit:
    LogSave()
    sys.exit()
  if dStockData and dStockData['email_1']:
    try:
      Log("")
      Log("A enviar o email de reporte dos stocks a farmacia...")
      lsTo = []
      for dTo in dConfig['report_5'][0]['to']:
        if "_text" in dTo and dTo['_text'] != "":
          lsTo.append(dTo['_text'])
      lsCc = []
      for dCc in dConfig['report_5'][0]['cc']:
        if "_text" in dCc and dCc['_text'] != "":
          lsCc.append(dCc['_text'])
      EmailStockReport_1(sSMTPSender, lsTo, lsCc, sSignature, dStockData)
    except Exception as dError:
      Log("")
      Log("Erro: Nao foi possivel enviar o email.")
      Log("Mensagem de erro:\n" + str(dError))
      Log("")
      bExit = True
    if bExit:
      EmailLogout()
      LogSave()
      sys.exit()
  if dStockData and dStockData['email_2']:
    try:
      Log("")
      Log("A enviar o email de reporte dos stocks para #INSTITUTION#...")
      lsTo = []
      for dTo in dConfig['report_6'][0]['to']:
        if "_text" in dTo and dTo['_text'] != "":
          lsTo.append(dTo['_text'])
      lsCc = []
      for dCc in dConfig['report_6'][0]['cc']:
        if "_text" in dCc and dCc['_text'] != "":
          lsCc.append(dCc['_text'])
      EmailStockReport_2(sSMTPSender, lsTo, lsCc, sSignature, dStockData)
    except Exception as dError:
      Log("")
      Log("Erro: Nao foi possivel enviar o email.")
      Log("Mensagem de erro:\n" + str(dError))
      Log("")
      bExit = True
    if bExit:
      EmailLogout()
      LogSave()
      sys.exit()
  for dGroup in dConfig['report_1'][0]['group']:
    if "email" in dGroup and dGroup['email']:
      try:
        Log("")
        Log("A enviar resultados em PDF por email...\n- Grupo: {}".format(dGroup['caption']))
        lsTo = []
        for dTo in dGroup['to']:
          if "_text" in dTo and dTo['_text'] != "":
            lsTo.append(dTo['_text'])
        lsCc = []
        for dCc in dGroup['cc']:
          if "_text" in dCc and dCc['_text'] != "":
            lsCc.append(dCc['_text'])
        sSubject = dGroup['subject'][0]['_text']
        sBody = eval("\"{}\"".format(dGroup['body'][0]['_text']))
        EmailSendWithFile(
          sSMTPSender,
          lsTo,
          lsCc,
          sSubject,
          sBody.format(GetGreeting(), sDatetimeStartReport1, sDatetimeEndReport1, sSignature),
          dGroup['filenames'][0],
          dGroup['filenames'][1]
        )
        Log("")
        Log("Email com resultados em PDF enviado para:\n- {}".format("\n- ".join(lsTo + lsCc)))
      except Exception as dError:
        Log("")
        Log("Erro: Nao foi possivel enviar o email.")
        Log("Mensagem de erro:\n" + str(dError))
        Log("")
        bExit = True
      if bExit:
        EmailLogout()
        LogSave()
        sys.exit()
  for dGroup in dConfig['report_2'][0]['group']:
    if "email" in dGroup and dGroup['email']:
      try:
        Log("")
        Log("A enviar resultados em XLS por email...\n- Grupo: {}".format(dGroup['caption']))
        lsTo = []
        for dTo in dGroup['to']:
          if "_text" in dTo and dTo['_text'] != "":
            lsTo.append(dTo['_text'])
        lsCc = []
        for dCc in dGroup['cc']:
          if "_text" in dCc and dCc['_text'] != "":
            lsCc.append(dCc['_text'])
        sSubject = dGroup['subject'][0]['_text']
        sBody = eval("\"{}\"".format(dGroup['body'][0]['_text']))
        EmailSendWithFile(
          sSMTPSender,
          lsTo,
          lsCc,
          sSubject,
          sBody.format(GetGreeting(), sDatetimeStartReport2, sDatetimeEndReport2, sSignature),
          dGroup['filenames'][0],
          dGroup['filenames'][1]
        )
        Log("")
        Log("Email com resultados em XLS enviado para:\n- {}".format("\n- ".join(lsTo + lsCc)))
      except Exception as dError:
        Log("")
        Log("Erro: Nao foi possivel enviar o email.")
        Log("Mensagem de erro:\n" + str(dError))
        Log("")
        bExit = True
      if bExit:
        EmailLogout()
        LogSave()
        sys.exit()
  for dGroup in dConfig['report_3'][0]['group']:
    if "email" in dGroup and dGroup['email']:
      try:
        Log("")
        Log("A enviar relatorio PDF de utentes COVID-19 por email...\n- Grupo: {}".format(dGroup['caption']))
        lsTo = []
        for dTo in dGroup['to']:
          if "_text" in dTo and dTo['_text'] != "":
            lsTo.append(dTo['_text'])
        lsCc = []
        for dCc in dGroup['cc']:
          if "_text" in dCc and dCc['_text'] != "":
            lsCc.append(dCc['_text'])
        sSubject = dGroup['subject'][0]['_text']
        sBody = eval("\"{}\"".format(dGroup['body'][0]['_text']))
        EmailSendWithFile(
          sSMTPSender,
          lsTo,
          lsCc,
          sSubject,
          sBody.format(GetGreeting(), sSignature),
          dGroup['filenames'][0],
          dGroup['filenames'][1]
        )
        Log("")
        Log("Email com relatorio PDF de utentes COVID-19 enviado para:\n- {}".format("\n- ".join(lsTo + lsCc)))
      except Exception as dError:
        Log("")
        Log("Erro: Nao foi possivel enviar o email.")
        Log("Mensagem de erro:\n" + str(dError))
        Log("")
        bExit = True
      if bExit:
        EmailLogout()
        LogSave()
        sys.exit()
  for dGroup in dConfig['report_4'][0]['group']:
    if "email" in dGroup and dGroup['email'] and "filenames_1" in dGroup and dGroup['filenames_1']:
      try:
        Log("")
        Log("A enviar resultados em PDF de grupos por email...\n- Grupo: {}".format(dGroup['caption']))
        lsTo = []
        for dTo in dGroup['to']:
          if "_text" in dTo and dTo['_text'] != "":
            lsTo.append(dTo['_text'])
        lsCc = []
        for dCc in dGroup['cc']:
          if "_text" in dCc and dCc['_text'] != "":
            lsCc.append(dCc['_text'])
        sSubject = dGroup['subject'][0]['_text']
        sBody = eval("\"{}\"".format(dGroup['body'][0]['_text']))
        sContents = ""
        ltsFiles = list()
        uIndex = 0
        for dFile in dGroup['filenames_1']:
          ltsFiles.append((dFile['filename'], dFile['filename_email']))
          sContents = sContents + "\n  - Grupo " + str(uIndex + 1) + ": " + dFile['title']
          uIndex = uIndex + 1
        EmailSendWithMultipleFiles(
          sSMTPSender,
          lsTo,
          lsCc,
          sSubject.format("grupos"),
          sBody.format(GetGreeting(), sDatetimeStartReport3, sDatetimeEndReport3, sContents, sSignature),
          ltsFiles
        )
        Log("")
        Log("Email com resultados em PDF de grupos enviado para:\n- {}".format("\n- ".join(lsTo + lsCc)))
      except Exception as dError:
        Log("")
        Log("Erro: Nao foi possivel enviar o email.")
        Log("Mensagem de erro:\n" + str(dError))
        Log("")
        bExit = True
      if bExit:
        EmailLogout()
        LogSave()
        sys.exit()
  for dGroup in dConfig['report_7'][0]['group']:
    if "email" in dGroup and dGroup['email']:
      try:
        Log("")
        Log("A enviar resumo de testes de rastreio em instituicoes sociais por email...\n- Grupo: {}".format(dGroup['caption']))
        lsTo = []
        for dTo in dGroup['to']:
          if "_text" in dTo and dTo['_text'] != "":
            lsTo.append(dTo['_text'])
        lsCc = []
        for dCc in dGroup['cc']:
          if "_text" in dCc and dCc['_text'] != "":
            lsCc.append(dCc['_text'])
        sSubject = dGroup['subject'][0]['_text']
        sBody = eval("\"{}\"".format(dGroup['body'][0]['_text']))
        EmailSendWithFile(
          sSMTPSender,
          lsTo,
          lsCc,
          sSubject.format(
            oDatetimeStartReport7.strftime(dConfig['report_7'][0]['datetime'][0]['_text']),
            oDatetimeEndReport7.strftime(dConfig['report_7'][0]['datetime'][0]['_text'])
          ),
          sBody.format(
            GetGreeting(),
            oDatetimeStartReport7.strftime(dConfig['report_7'][0]['datetime'][0]['_text']),
            oDatetimeEndReport7.strftime(dConfig['report_7'][0]['datetime'][0]['_text']),
            sSignature
          ),
          dGroup['filenames'][0],
          dGroup['filenames'][1]
        )
        Log("")
        Log("Email com resumo de testes de rastreio em instituicoes sociais enviado para:\n- {}".format("\n- ".join(lsTo + lsCc)))
      except Exception as dError:
        Log("")
        Log("Erro: Nao foi possivel enviar o email.")
        Log("Mensagem de erro:\n" + str(dError))
        Log("")
        bExit = True
      if bExit:
        EmailLogout()
        LogSave()
        sys.exit()
  EmailLogout()
try:
  Log("")
  Log("A abrir relatorios e sitios web...")
  if 'website' in dConfig:
    for dWebsite in dConfig['website']:
      bInput = True
      while bInput:
        sInput = input("- Abrir {} (s/n)? ".format(dWebsite['caption'])).strip()
        if sInput.lower() == "s":
          sSummary1FileName = CreatePDFReport_Summary_1("#INSTITUTION#", dSummaryData)
          os.startfile(tsPDF_2_Filenames[0])
          os.startfile(sSummary1FileName)
          os.startfile(dWebsite['_text'])
          bInput = False
        elif sInput.lower() == "n":
          bInput = False
        else:
          bInput = True
  bInput = True
  while bInput:
    sInput = input("- Abrir a pasta dos relatorios (s/n)? ").strip()
    if sInput.lower() == "s":
      os.startfile(sDestinationFolder)
      bInput = False
    elif sInput.lower() == "n":
      bInput = False
    else:
      bInput = True

except Exception as dError:
  Log("")
  Log("Erro: Nao foi possivel abrir um dos ficheiros.")
  Log("Mensagem de erro:\n" + str(dError))
  Log("")
  bExit = True
if bExit:
  LogSave()
  sys.exit()

Log("")
Log("Rotina concluida.")

LogSave()
