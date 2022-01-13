# update.py (utf-8)
# 
# Edited by: RR-DSE
# Timestamp: 22-01-13 15:55:49

# ------------
# Dependencies
# ------------

import unicodedata
import subprocess
import mysql.connector
import re
import listools
import datetime
import sys
import io
import os
import glob
import xlrd
from openpyxl import Workbook, load_workbook
import xml.etree.ElementTree as ElementTree
from copy import copy
import warnings

# ---------
# Constants
# ---------

sConfigFolder = "config"
sConfigXML = "config.xml"

sCode = "utf-8"
sInputDateFormat = "%Y-%m-%d"
sDateFormat = "%Y-%m-%d"
sInputDateTimeFormat = "%Y-%m-%d %H:%M:%S"
sDateTimeFormat = "%Y-%m-%d %H:%M:%S"
sDateFormat = "%Y-%m-%d"
sHourFormat = "%H:%M"

sStandardDepartment = "28"
sStandardStatus = "outpatient"
sUserID = "15"
sFileTimestampFormat = "%Y_%m_%d_%H_%M_%S"
sTimestampFormat = "%Y-%m-%d %H:%M:%S"

sSampleMatch = "\\d{6,7}$"

uResSourceStartLine = 16
uResSourceTrimLines = 0

sSetupFolder = "config"
sSourceFolder = "update"
sLogFolder = "logs"

sTransDepartmentsTable = "{}\\{}".format(sSetupFolder, "trans_departments.csv")
sTransResultsTable = "{}\\{}".format(sSetupFolder, "trans_results.csv")
sTestRelationsTable = "{}\\{}".format(sSetupFolder, "test_relations.csv")
sTransGroupClassesTable = "{}\\{}".format(sSetupFolder, "trans_group_classes.csv")
sSelResultsTable = "{}\\{}".format(sSetupFolder, "sel_results.csv")
sSelWorklistTable = "{}\\{}".format(sSetupFolder, "sel_worklist.csv")

sSourceRes = "{}\\{}".format(sSourceFolder, "resultados.tsv")
sSourceWorklist = "{}\\{}".format(sSourceFolder, "lista_trabalho.tsv")

sSampleRE = r"\d{6,7}"

uEmailsFileStartLine = 16
uEmailsFileTrimLines = 0
ltEmailsSelectors = [
  ("NC", "sample", listools.C_SELECTOR_STR),
  ("Email pessoal", "email_patient", listools.C_SELECTOR_STR),
  ("Email clínico", "email_clinician", listools.C_SELECTOR_STR),
  ("Email institucional", "email_manager", listools.C_SELECTOR_STR),
  ("Email de destino", "email_unknown", listools.C_SELECTOR_STR),
  ("Resultado enviado por email", "sent", listools.C_SELECTOR_STR)
  ]
  
sEmailRE = r"^[A-Za-z0-9\.\+_-]+@[A-Za-z0-9\._-]+\.[a-zA-Z]*$"
sNameCleanRE = "d'|\'|\"|\\^|\\.|\\s+de\\s+|\\s+do\\s+|\\s+da\\s+|\\s+dos\\s+|\\s+das\\s+|\\s+des\\s+"
sEmailCleanRE = "\'|\"|\\x1F|\\x1C|\\x1D|\\x1E|~|\\^|ª|º|\\\\|/"

# --------------
# Global objects
# --------------

class Error(Exception):
  pass

dConfig = None

sHost = None
sUser = None
sPassword = None
sDatabase = None

dTransDepartments = listools.LoadTable(sTransDepartmentsTable)
dTransResults = listools.LoadTable(sTransResultsTable)
dTransGroupClasses = listools.LoadDictList(sTransGroupClassesTable, "context")
dTestRelations = listools.LoadDict(sTestRelationsTable, "test")

lsTests = list()
for sTest in dTestRelations:
 if sTest not in lsTests:
   lsTests.append(sTest)

dMySqlDB = None
dMySqlCursor = None

dMemLog = io.StringIO(newline=None)
dMemErrorLog = io.StringIO(newline=None)
dMemUpdateLog = io.StringIO(newline=None)
dMemInsertLog = io.StringIO(newline=None)

uErrorCount = 0
    
lsSocialInstitutionClasses = list()

# -----------------
# Auxiliary methods
# -----------------

def SearchDepartmentAndStatus(sSource):
  global dTransDepartments
  sDepartment = None
  sStatus = None
  if not sSource or sSource == "":
    return (sDepartment, sStatus)
  for dRow in dTransDepartments[1]:
    sSearch = sSource
    sMatch = dRow["source"]
    if sSearch == None:
      sSearch = ""
    if sMatch == None:
      sMatch = ""
    if re.search(sMatch, sSearch, re.IGNORECASE):
      sDepartment = dRow["department"]
      sStatus = dRow["status"]
      break
  return (sDepartment, sStatus)

def SearchResult(sResult):
  global dTransResults
  sRes = None
  if not sResult or sResult == "":
    return sRes
  for dRow in dTransResults[1]:
    sSearch = sResult
    sMatch = dRow["source"]
    if sSearch == None:
      sSearch = ""
    if sMatch == None:
      sMatch = ""
    if re.search(sMatch, sSearch, re.IGNORECASE):
      sRes = dRow["result"]
      break
  return sRes

def SearchGroupClass(sSource):
  global dTransGroupClasses
  sRes = None
  if not sSource:
    return sRes
  for dRow in dTransGroupClasses['group']:
    sSearch = sSource
    sMatch = dRow["source"]
    if sSearch == None:
      sSearch = ""
    if sMatch == None:
      sMatch = ""
    if re.search(sMatch, sSearch, re.IGNORECASE):
      sRes = dRow["result"]
      break
  return sRes

def SearchGroupPatientClass(sSource):
  global dTransGroupClasses
  sRes = None
  if not sSource:
    return sRes
  for dRow in dTransGroupClasses['patient']:
    sSearch = sSource
    sMatch = dRow["source"]
    if sSearch == None:
      sSearch = ""
    if sMatch == None:
      sMatch = ""
    if re.search(sMatch, sSearch, re.IGNORECASE):
      sRes = dRow["result"]
      break
  return sRes

def GetGender(sSource):
  if sSource == "M" or sSource == "m":
    return "1"
  if sSource == "F" or sSource == "f":
    return "2"
  return "0"

def GetDate(sSource = None):
  global sDateFormat, sInputDateFormat
  if not sSource or sSource == "":
    return datetime.datetime.now().strftime(sDateFormat)
  else:
    return datetime.datetime.strptime(sSource.replace("/", "-").replace("\\", "-"), sInputDateFormat).strftime(sDateFormat)

def GetDateTime(sSource = None):
  global sDateTimeFormat, sInputDateTimeFormat
  if not sSource or sSource == "":
    return datetime.datetime.now().strftime(sDateTimeFormat)
  else:
    return datetime.datetime.strptime(sSource.replace("/", "-").replace("\\", "-"), sInputDateTimeFormat).strftime(sDateTimeFormat)

def GetDateDelta(sFirst, sSecond):
  global sDateFormat
  dFirst = datetime.datetime.strptime(sFirst, sDateFormat)
  dSecond = datetime.datetime.strptime(sSecond, sDateFormat)
  return (dSecond-dFirst).days

def GetDateTimeDelta(sFirst, sSecond):
  global sDateTimeFormat
  dFirst = datetime.datetime.strptime(sFirst, sDateTimeFormat)
  dSecond = datetime.datetime.strptime(sSecond, sDateTimeFormat)
  return (dSecond-dFirst).seconds / 86400

def GetRecord(sRecord):
  if not sRecord or sRecord == "":
    return "NULL"
  return sRecord

def GetText(sText):
  if not sText or sText == "":
    return "NULL"
  return sText

def GetStateID1(sStateID1, sStateID1Type):
  if not sStateID1 or sStateID1 == "":
    return "NULL"
  if re.search(".*SERV.*NAC.*SA.*", sStateID1Type, re.IGNORECASE):
    return sStateID1
  return "NULL"

def GetDepartment(sDepartment, sRecord):
  global sStandardDepartment
  if not sRecord or sRecord == "" or sRecord == "NULL":
    return sStandardDepartment
  if not sDepartment or sDepartment == "" or sDepartment == "NULL":
    return sStandardDepartment
  dDepartmentAndStatus = SearchDepartmentAndStatus(sDepartment)
  if not dDepartmentAndStatus[0]:
    return "NULL"
  return dDepartmentAndStatus[0]

def GetStatus(sDepartment, sRecord):
  global sStandardStatus
  if not sRecord or sRecord == "" or sRecord == "NULL":
    return sStandardStatus
  if not sDepartment or sDepartment == "" or sDepartment == "NULL":
    return sStandardStatus
  dDepartmentAndStatus = SearchDepartmentAndStatus(sDepartment)
  if not dDepartmentAndStatus[1]:
    return "unknown"
  return dDepartmentAndStatus[1]

def GetName(sSource):
  sRes = sSource.upper()
  return sRes

def GetResult(sSource):
  sRes = SearchResult(sSource.strip())
  return sRes

def CheckPatientUpdate(
  dRecord,
  sRecord,
  sStateID1,
  sGender,
  sStatus,
  sStatusDateTime,
  sDepartment,
  sAdmittanceDate):
  sRes = dict()
  if dRecord['gender'] == "0" and sGender != "0":
    sRes['gender'] = sGender
  if (dRecord['record'].upper() == "NULL" or dRecord['record'] == "") and (sRecord and sRecord != "" and sRecord.upper() != "NULL"):
    sRes['record'] = GetDBString(sRecord)
  if (dRecord['state_id_1'].upper() == "NULL" or dRecord['state_id_1'] == "") and (sStateID1 and sStateID1 != "" and sStateID1.upper() != "NULL"):
    sRes['state_id_1'] = GetDBString(sStateID1)
  if sStatus != "unknown":
    if dRecord['patient_status'] == "inpatient" or \
       dRecord['patient_status'] == "outpatient" or \
       dRecord['patient_status'] == "outpatientquarant" or \
       dRecord['patient_status'] == "outpatientresolved" or \
       dRecord['patient_status'] == "discharged":
      if dRecord['patient_status_datetime'] == "NULL" or GetDateTimeDelta(sStatusDateTime, GetDateTime(dRecord['patient_status_datetime'])) < 0:
        sRes['patient_status'] = GetDBString(sStatus)
        sRes['patient_status_datetime'] = GetDBString(sStatusDateTime)
  if sDepartment != "NULL" and sAdmittanceDate != "NULL":
    if dRecord['department'] != "NULL" and dRecord['admittance_date'] != "NULL":
      if GetDateDelta(sAdmittanceDate, GetDate(dRecord['admittance_date'])) < 0:
        sRes['department'] = sDepartment
        sRes['admittance_date'] = GetDBString(sAdmittanceDate)
    else:
      sRes['department'] = sDepartment
      sRes['admittance_date'] = GetDBString(sAdmittanceDate)
  return sRes

def RemoveAccents(sInput):
  sNFKD = unicodedata.normalize('NFKD', sInput)
  sRes = u"".join([c for c in sNFKD if not unicodedata.combining(c)])
  return sRes

def CleanStrName(sSrc):
  global sNameCleanRE
  if not sSrc:
    return None
  sRes = re.sub(sNameCleanRE, " ", str(sSrc), flags = re.IGNORECASE)
  sRes = RemoveAccents(sRes)
  sRes = re.sub(r"^\s+|\s+$", "", sRes)
  sRes = re.sub(r"\s+", " ", sRes)
  sRes = sRes.upper()
  return sRes

def CleanStrEmail(sSrc):
  global sEmailCleanRE
  if not sSrc:
    return None
  sRes = re.sub(sEmailCleanRE, "", str(sSrc), flags = re.IGNORECASE)
  sRes = RemoveAccents(sRes)
  sRes = re.sub(r"^\s+|\s+$", "", sRes)
  sRes = re.sub(r"\s+", " ", sRes)
  sRes = sRes.lower()
  return sRes

def CleanStrGroup(sSrc):
  if not sSrc:
    return None
  sRes = re.sub(r"^\s+|\s+$", "", str(sSrc))
  sRes = re.sub(r"\s+", " ", sRes)
  return sRes

# -----------
# XML methods
# -----------

def XMLToDict(oXML, bRoot = True):
  if bRoot:
    return {oXML.tag : XMLToDict(oXML, False)}
  dRes = copy(oXML.attrib)
  if oXML.text:
    dRes["_text"] = oXML.text
  for oElement in oXML.findall("./*"):
    if oElement.tag not in dRes:
      dRes[oElement.tag] = []
    dRes[oElement.tag].append(XMLToDict(oElement, False))
  return dRes

# --------------
# Config methods
# --------------
 
def LoadConfigXML():
  global \
    sConfigFolder,\
    sConfigXML,\
    dConfig,\
    sHost,\
    sUser,\
    sPassword,\
    sDatabase,\
    lsSocialInstitutionClasses
  oFile = open("{}\\{}".format(sConfigFolder, sConfigXML), "r", encoding = "utf-8")
  sFile = oFile.read()
  oFile.close()
  oXML = ElementTree.fromstring(sFile)
  dConfig = XMLToDict(oXML)['config']
  sHost = dConfig['host'][0]['_text']
  sUser = dConfig['user'][0]['_text']
  sPassword = dConfig['password'][0]['_text']
  sDatabase = dConfig['database'][0]['_text']
  for dClass in dConfig['social_institution_classes'][0]['class']:
    lsSocialInstitutionClasses.append(dClass['_text'])
  return dConfig

# ----------------
# Database methods
# ----------------

def MySqlCursorStart():
  global dMySqlDB, dMySqlCursor
  global sHost, sUser, sPassword, sDatabase
  dMySqlDB = mysql.connector.connect(host = sHost, user = sUser, passwd = sPassword, database = sDatabase)
  dMySqlCursor = dMySqlDB.cursor(dictionary = True)
  return

def MySqlCursorClose():
  global dMySqlDB, dMySqlCursor
  dMySqlCursor.close()
  dMySqlDB.close()
  return

def RunMySqlCursor(sSQLQuery, bFetch = False):
  global dMySqlCursor, sDateFormat, sDateTimeFormat
  dRes = list()
  dMySqlCursor.execute(sSQLQuery)
  if dMySqlCursor.rowcount == 0 or not bFetch:
    return dRes
  else:
    dRes = dMySqlCursor.fetchall()
    for dRow in dRes:
      for sKey in dRow:
        if isinstance(dRow[sKey], int):
          dRow[sKey] = str(dRow[sKey])
        elif isinstance(dRow[sKey], float):
          dRow[sKey] = str(float(dRow[sKey]))
        elif isinstance(dRow[sKey], datetime.datetime):
          dRow[sKey] = dRow[sKey].strftime(sDateTimeFormat)
        elif isinstance(dRow[sKey], datetime.date):
          dRow[sKey] = dRow[sKey].strftime(sDateFormat)
        elif dRow[sKey] == None:
          dRow[sKey] = "NULL"
        else:
          dRow[sKey] = str(dRow[sKey])
  return dRes

def RunMySqlBatch(sSQLQuery, bFetch = False):
  global sHost, sUser, sPassword, sDatabase, sCode
  sCmd = "mysql -h {0} -u {1} -p{2} {3} -B -e \"{4}\"".format(sHost, sUser, sPassword, sDatabase, sSQLQuery)
  lsLines = subprocess.check_output(sCmd, stderr=subprocess.DEVNULL).decode(sCode).splitlines()
  if len(lsLines) == 0:
    return list()
  lsFields = lsLines[0].split("\t")
  ldRes = list()
  for sLine in lsLines[1:]:
    lsLine = sLine.split("\t")
    ldNew = dict()
    uIndex = 0
    for sField in lsLine:
      ldNew[lsFields[uIndex]] = sField
      uIndex = uIndex + 1
    ldRes.append(ldNew)
  return ldRes

def RunMySql(sSQLQuery, bFetch = False):
  #return RunMySqlBatch(sSQLQuery, bFetch)
  return RunMySqlCursor(sSQLQuery, bFetch)

def StartTransactionMySql():
  return
  global dMySqlDB
  if dMySqlDB:
    dMySqlDB.start_transaction()

def LockTablesMySql(lsTables):
  global dMySqlDB
  if not lsTables or len(lsTables) == 0:
    return
  sQuery = "LOCK TABLE {} WRITE;".format(" WRITE, ".join(lsTables))
  if dMySqlDB:
    return RunMySql(sQuery, False)

def UnlockTablesMySql():
  global dMySqlDB
  sQuery = "UNLOCK TABLES;"
  if dMySqlDB:
    return RunMySql(sQuery, False)

def CommitMySql():
  global dMySqlDB
  if dMySqlDB:
    dMySqlDB.commit()

def SearchDBPatient(sRecord, sName, sBirthday):
  dRes = None
  if sRecord and sRecord.upper() != "NULL" and sRecord != "":
    sQuery = "SELECT * FROM patients WHERE record = '{}' AND status <> 2".format(sRecord)
    dDBQueryRes = RunMySql(sQuery, True)
    if len(dDBQueryRes) == 0:
      sQuery = "SELECT * FROM patients WHERE UPPER(name) = '{}' AND birthday = '{}' AND status <> 2".format(sName.upper(), sBirthday)
      dDBQueryRes = RunMySql(sQuery, True)
      if len(dDBQueryRes) == 0:
        return None
      elif len(dDBQueryRes) > 1:
        return "ErrorMultipleResults"
      else:
        if not (dDBQueryRes[0]['record'].upper() == "NULL" or dDBQueryRes[0]['record'] == ""):
          return "ErrorIncorrectData"
        else:
          dDBQueryRes[0]['curr_record'] = sRecord
          dRes = dDBQueryRes[0]
    elif len(dDBQueryRes) > 1:
      return "ErrorMultipleResults"
    else:
      if sName.upper() == dDBQueryRes[0]['name'].upper() and sBirthday == dDBQueryRes[0]['birthday']:
        dRes = dDBQueryRes[0]
      else:
        return "ErrorIncorrectData"
  else:
    sQuery = "SELECT * FROM patients WHERE UPPER(name) = '{}' AND birthday = '{}' AND status <> 2".format(sName.upper(), sBirthday)
    dDBQueryRes = RunMySql(sQuery, True)
    if len(dDBQueryRes) == 0:
      return None
    elif len(dDBQueryRes) > 1:
      return "ErrorMultipleResults"
    else:
      dRes = dDBQueryRes[0]
  return dRes

def SearchDBTest(sSample):
  dRes = None
  if sSample and sSample.upper() != "NULL" and sSample != "":
    sQuery = "SELECT * FROM tests WHERE sample_id = '{}' AND status <> 2".format(sSample)
    dDBQueryRes = RunMySql(sQuery, True)
    if len(dDBQueryRes) == 0:
      return None
    elif len(dDBQueryRes) > 1:
      return "ErrorMultipleResults"
    else:
      dRes = dDBQueryRes[0]
  return dRes

def GetDBString(sValue):
  if sValue == "NULL" or sValue == "null" or not sValue:
    return "NULL"
  else:
    sValue = re.sub("'", "''", sValue)
    return "'{}'".format(sValue)

def GetDBInt(vValue):
  if not vValue:
    return "NULL"
  else:
    return "{}".format(int(vValue))

def InsertDBPatient(
  sRecord,
  sName,
  sBirthday,
  sGender,
  sStateID1,
  sStatus,
  sStatusDateTime,
  sDepartment,
  sAdmittanceDate,
  sFloor,
  sRoom,
  sBed):
  global sUserID, dMySqlDB
  sQuery = """
    INSERT INTO patients VALUES (
      NULL, 
      {}, 
      {}, 
      {}, 
      {}, 
      {},
      {},
      {},
      {},
      {},
      {},
      {},
      {},
      NULL,
      NULL,
      NULL,
      NULL,
      1,
      NULL,
      {},
      NOW(),
      {},
      NOW()
    )""".format(
      GetDBString(sName),
      GetDBString(sBirthday),
      sGender,
      GetDBString(sRecord),
      GetDBString(sStateID1),
      GetDBString(sStatus),
      GetDBString(sStatusDateTime),
      sDepartment,
      GetDBString(sAdmittanceDate),
      GetDBString(sFloor),
      GetDBString(sRoom),
      GetDBString(sBed),
      sUserID,
      sUserID)
  RunMySql(sQuery, False)
  if dMySqlDB:
    dMySqlDB.commit()
  return SearchDBPatient(sRecord, sName, sBirthday)
    
def InsertDBTest(
  sSample,
  sAccession,
  sName,
  sBirthday,
  sGender,
  sRecord,
  sEpisode,
  sStateID1,
  sPatientID, 
  sDepartment,
  sDepartmentCode,
  sDepartmentID,
  sTestCode,
  sMethodCode,
  sSampleDate,
  sResult,
  sResultComments,
  sResultDateTime,
  sTecValidator,
  sValidationDatetime,
  sBioValidator,
  sPrescriber,
  sAddress1,
  sAddress2):
  global sUserID, dMySqlDB
  sQuery = """
    INSERT INTO tests VALUES (
      NULL, 
      {}, 
      {}, 
      {}, 
      {}, 
      {}, 
      {},
      {},
      {},
      {},
      {},
      {},
      {},
      {},
      {},
      {},
      {},
      {},
      {},
      {},
      {},
      {},
      {},
      NULL,
      NULL,
      {},
      {},
      NULL,
      NULL,
      NULL,
      {},
      1,
      NULL,
      {},
      NOW(),
      {},
      NOW()
    )""".format(
      GetDBString(sName),
      GetDBString(sBirthday),
      sGender,
      GetDBString(sRecord),
      GetDBString(sEpisode),
      GetDBString(sStateID1),
      GetDBString(sPatientID),
      GetDBString(sDepartment),
      GetDBString(sDepartmentCode),
      GetDBString(sDepartmentID),
      GetDBString(sTestCode),
      GetDBString(sMethodCode),
      GetDBString(sSample),
      GetDBInt(sSample),
      GetDBString(sAccession),
      GetDBString(sSampleDate),
      GetDBString(sResult),
      GetDBString(sResultComments),
      GetDBString(sResultDateTime),
      GetDBString(sTecValidator),
      GetDBString(sValidationDatetime),
      GetDBString(sBioValidator),
      GetDBString(sAddress1 if len(sAddress1) < 256 else sAddress1[0:255]),
      GetDBString(sAddress2 if len(sAddress2) < 64 else sAddress2[0:63]),
      GetDBString(sPrescriber),
      sUserID,
      sUserID)
  RunMySql(sQuery, False)
  if dMySqlDB:
    dMySqlDB.commit()
  return 

def UpdateDBPatient(sID, dData):
  global sUserID, dMySqlDB
  sSetData = ""
  for sDataField, sValue in dData.items():
    if sSetData != "":
      sSetData = sSetData + ", "
    sSetData = sSetData + sDataField + " = " + sValue
  if sSetData == "":
    return
  sQuery = "INSERT audit_patients SELECT * FROM patients WHERE id={}".format(sID)
  RunMySql(sQuery, False)
  if dMySqlDB:
    dMySqlDB.commit()
  sQuery = "UPDATE patients SET {}, mod_user = {}, mod_datetime = NOW() WHERE id = {}".format(sSetData, sUserID, sID)
  RunMySql(sQuery, False)
  if dMySqlDB:
    dMySqlDB.commit()
  return 

def UpdateDBTest(
  dRow,
  sSample,
  sAccession,
  sName,
  sBirthday,
  sGender,
  sRecord,
  sEpisode,
  sStateID1,
  sPatientID, 
  sDepartment,
  sDepartmentCode,
  sDepartmentID,
  sTestCode,
  sMethodCode,
  sSampleDate,
  sResult,
  sResultComments,
  sResultDateTime,
  sTecValidator,
  sValidationDatetime,
  sBioValidator,
  sPrescriber,
  sAddress1,
  sAddress2):
  global sUserID, dMySqlDB
  bUpdate = False
  if sAccession != dRow['accession'][0:6]:
    bUpdate = True
  if sRecord != dRow['record']:
    bUpdate = True
  if sEpisode != dRow['episode']:
    bUpdate = True
  if sName != dRow['name']:
    bUpdate = True
  if sBirthday != dRow['birthday']:
    bUpdate = True
  if sGender != dRow['gender']:
    bUpdate = True
  if sStateID1 != dRow['state_id_1']:
    bUpdate = True
  if sPatientID != dRow['patient_id']:
    bUpdate = True
  if sDepartment != dRow['department']:
    bUpdate = True
  if sDepartmentCode != dRow['department_code']:
    bUpdate = True
  if sDepartmentID != dRow['department_id']:
    bUpdate = True
  if sTestCode != dRow['test_code']:
    bUpdate = True
  if sMethodCode != dRow['method_code']:
    bUpdate = True
  if sSampleDate != dRow['sample_date']:
    bUpdate = True
  if sResult != dRow['result_code']:
    bUpdate = True
  if sResultComments != dRow['result_comments']:
    bUpdate = True
  if sResultDateTime != dRow['result_datetime']:
    bUpdate = True
  if sTecValidator != dRow['tec_validator']:
    bUpdate = True
  if sValidationDatetime != dRow['validation_datetime']:
    bUpdate = True
  if sBioValidator != dRow['bio_validator']:
    bUpdate = True
  if sPrescriber != dRow['prescriber']:
    bUpdate = True
  if sAddress1 and sAddress1 != "" and sAddress1.upper() != "NULL" and sAddress1 != dRow['address_1']:
    bUpdate = True
    sUsedAddress1 = sAddress1
  else:
    sUsedAddress1 = dRow['address_1']
  if sAddress2 and sAddress2 != "" and sAddress2.upper() != "NULL" and sAddress2 != dRow['address_2']:
    bUpdate = True
    sUsedAddress2 = sAddress2
  else:
    sUsedAddress2 = dRow['address_2']
  if bUpdate:
    sQuery = "INSERT audit_tests SELECT * FROM tests WHERE id={}".format(dRow['id'])
    RunMySql(sQuery, False)
    if dMySqlDB:
      dMySqlDB.commit()
    sQuery = """
      UPDATE tests SET
        accession = {},
        record = {},
        episode = {},
        name = {},
        birthday = {},
        gender = {},
        state_id_1 = {},
        patient_id = {},
        department = {},
        department_code = {},
        department_id = {},
        test_code = {},
        method_code = {},
        sample_date = {},
        result_code = {},
        result_comments = {},
        result_datetime = {},
        tec_validator = {},
        validation_datetime = {},
        bio_validator = {},
        prescriber = {},
        address_1 = {},
        address_2 = {},
        mod_user = {},
        mod_datetime = NOW()
      WHERE id = {}
      """.format(
        GetDBString(sAccession),
        GetDBString(sRecord),
        GetDBString(sEpisode),
        GetDBString(sName),
        GetDBString(sBirthday),
        sGender,
        GetDBString(sStateID1),
        GetDBString(sPatientID),
        GetDBString(sDepartment),
        GetDBString(sDepartmentCode),
        GetDBString(sDepartmentID),
        GetDBString(sTestCode),
        GetDBString(sMethodCode),
        GetDBString(sSampleDate),
        GetDBString(sResult),
        GetDBString(sResultComments),
        GetDBString(sResultDateTime),
        GetDBString(sTecValidator),
        GetDBString(sValidationDatetime),
        GetDBString(sBioValidator),
        GetDBString(sPrescriber),
        GetDBString(sUsedAddress1 if len(sUsedAddress1) < 256 else sUsedAddress1[0:255]),
        GetDBString(sUsedAddress2 if len(sUsedAddress2) < 64 else sUsedAddress2[0:63]),
        sUserID,
        dRow['id']
        )
    RunMySql(sQuery, False)
    if dMySqlDB:
      dMySqlDB.commit()
  return bUpdate

def ResetDBGroups():
  try:
    MySqlCursorStart()
  except Exception as dError:
    Log(1, "")
    Log(1, "Erro: Nao foi possivel ligar ao servidor de BD.")
    Log(1, "Mensagem de erro:\n" + str(dError))
    Log(1, "")
    raise dError
  try:
    sQuery = """
      DROP TABLE IF EXISTS groups
      """
    RunMySql(sQuery, False)
    sQuery = """
      CREATE TABLE groups
      (
        id SERIAL,
        name CHAR(128) NOT NULL,
        birthday DATE NULL,
        sample CHAR(32) UNIQUE NULL,
        state_id CHAR(32) NULL,
        "group" CHAR(128) NOT NULL,
        class ENUM('patient', 'collaborator', 'internal', 'contact') NOT NULL,
        department CHAR(128) NULL,
        category CHAR(128) NULL,
        location CHAR(128) NULL,
        social_institution TINYINT(1) NULL,
        institution_class ENUM('nursing_home', 'school', 'children_home', 'healthcare_unit', 'other') NULL,
        "date" date NOT NULL,
        source CHAR(32) NULL,
        status TINYINT(1) UNSIGNED NOT NULL,
        "datetime" DATETIME NOT NULL,
        user BIGINT(20) UNSIGNED NOT NULL
      )
      """
    RunMySql(sQuery, False)
    CommitMySql()
  except Exception as dError:
    Log(1, "")
    Log(1, "Erro: Ocorreu um erro no acesso a tabela 'groups' da base de dados.")
    Log(1, "Mensagem de erro:\n" + str(dError))
    Log(1, "")
    MySqlCursorClose()
    raise dError
  MySqlCursorClose()

def ResetDBTransmissions():
  dTable = None
  try:
    MySqlCursorStart()
  except Exception as dError:
    Log(1, "")
    Log(1, "Erro: Nao foi possivel ligar ao servidor de base de dados.")
    Log(1, "Mensagem de erro:\n" + str(dError))
    Log(1, "")
    raise dError
  try:
    sQuery = """
      DROP TABLE IF EXISTS transmissions
      """
    RunMySql(sQuery, False)
    sQuery = """
      DROP TABLE IF EXISTS audit_transmissions
      """
    RunMySql(sQuery, False)
    sQuery = """
      CREATE TABLE transmissions
      (
        id SERIAL,
        sample_id CHAR(32) NOT NULL,
        class ENUM('email', 'phone', 'print', 'internal', 'unknown') NOT NULL,
        description VARCHAR(256) NOT NULL,
        recipient_class ENUM('patient','clinician','manager','caregiver','family','unknown') NOT NULL,
        recipient_description VARCHAR(256) NOT NULL,
        "datetime" DATETIME NOT NULL,
        status ENUM('ok', 'pending', 'error', 'canceled') NOT NULL,
        add_user BIGINT(20) UNSIGNED NOT NULL,
        add_datetime DATETIME NOT NULL,
        mod_user BIGINT(20) UNSIGNED NOT NULL,
        mod_datetime DATETIME NOT NULL
      )
      """
    RunMySql(sQuery, False)
    sQuery = """
      CREATE TABLE audit_transmissions
      (
        id BIGINT(20) UNSIGNED NOT NULL,
        sample_id CHAR(32) NOT NULL,
        class ENUM('email', 'phone', 'print', 'internal', 'unknown') NOT NULL,
        description VARCHAR(256) NOT NULL,
        recipient_class ENUM('patient','clinician','manager','caregiver','family','unknown') NOT NULL,
        recipient_description VARCHAR(256) NOT NULL,
        "datetime" DATETIME NOT NULL,
        status ENUM('ok', 'pending', 'error', 'canceled') NOT NULL,
        add_user BIGINT(20) UNSIGNED NOT NULL,
        add_datetime DATETIME NOT NULL,
        mod_user BIGINT(20) UNSIGNED NOT NULL,
        mod_datetime DATETIME NOT NULL
      )
      """
    RunMySql(sQuery, False)
    CommitMySql()
  except Exception as dError:
    Log(1, "")
    Log(1, "Erro: Ocorreu um erro no acesso a tabela 'transmissions' da base de dados.")
    Log(1, "Mensagem de erro:\n" + str(dError))
    Log(1, "")
    MySqlCursorClose()
    raise dError
  MySqlCursorClose()

def ResetDBEmails():
  dTable = None
  try:
    MySqlCursorStart()
  except Exception as dError:
    Log(1, "")
    Log(1, "Erro: Nao foi possivel ligar ao servidor de base de dados.")
    Log(1, "Mensagem de erro:\n" + str(dError))
    Log(1, "")
    raise dError
  try:
    sQuery = """
      DROP TABLE IF EXISTS emails
      """
    RunMySql(sQuery, False)
    sQuery = """
      CREATE TABLE emails
      (
        id SERIAL,
        name CHAR(128) NOT NULL,
        birthday DATE NULL,
        state_id CHAR(32) NULL,
        email CHAR(128) NOT NULL,
        class ENUM('patient','clinician','manager','caregiver','family') NULL,
        date DATE NOT NULL,
        "datetime" DATETIME NOT NULL,
        user BIGINT(20) UNSIGNED NOT NULL
      )
      """
    RunMySql(sQuery, False)
    CommitMySql()
  except Exception as dError:
    Log(1, "")
    Log(1, "Erro: Ocorreu um erro no acesso a tabela 'emails' da base de dados.")
    Log(1, "Mensagem de erro:\n" + str(dError))
    Log(1, "")
    MySqlCursorClose()
    raise dError
  MySqlCursorClose()

# -----------
# Log methods
# -----------

def Log(uType, sText):
  global dMemLog, dMemErrorLog, dMemUpdateLog, dMemInsertLog
  global uErrorCount
  dMemLog.write(sText + os.linesep)
  if uType == 0:
    dMemErrorLog.write(sText + os.linesep)
    dMemUpdateLog.write(sText + os.linesep)
    dMemInsertLog.write(sText + os.linesep)
  if uType == 1:
    dMemErrorLog.write(sText + os.linesep)
    uErrorCount = uErrorCount + 1
  if uType == 2:
    dMemUpdateLog.write(sText + os.linesep)
  if uType == 3:
    dMemInsertLog.write(sText + os.linesep)
  print(sText)

def LogSave():
  global dMemLog, dMemErrorLog, dMemUpdateLog, dMemInsertLog, sLogFolder
  dMemLog.seek(0)
  sTimeStamp = datetime.datetime.now().strftime(sFileTimestampFormat)
  dFile = open("{}\\{}_{}".format(sLogFolder, sTimeStamp, "coviddxdb_update.log"), "w", encoding = "utf-8")
  dFile.write(dMemLog.read())
  dFile.close()
  dMemErrorLog.seek(0)
  dFile = open("{}\\{}_{}".format(sLogFolder, sTimeStamp, "coviddxdb_update_errors.log"), "w", encoding = "utf-8")
  dFile.write(dMemErrorLog.read())
  dFile.close()
  dMemUpdateLog.seek(0)
  dFile = open("{}\\{}_{}".format(sLogFolder, sTimeStamp, "coviddxdb_update_changes.log"), "w", encoding = "utf-8")
  dFile.write(dMemUpdateLog.read())
  dFile.close()
  dMemInsertLog.seek(0)
  dFile = open("{}\\{}_{}".format(sLogFolder, sTimeStamp, "coviddxdb_update_new_entries.log"), "w", encoding = "utf-8")
  dFile.write(dMemInsertLog.read())
  dFile.close()

# ---------------
# Process methods
# ---------------

def ProcessPatient(
  sRecord,
  sName,
  sBirthday,
  sGender,
  sStateID1,
  sStatus,
  sStatusDateTime,
  sDepartment,
  sAdmittanceDate,
  sFloor,
  sRoom,
  sBed):
  uProcessType = 0
  sProcessType = ""
  dDBPatient = SearchDBPatient(sRecord, sName, sBirthday)
  if dDBPatient and uProcessType == 0:
    if dDBPatient == "ErrorIncorrectData":
      sProcessType = "Erro: Dados de utente incorretos ou inconsistentes com os da base de dados."
      uProcessType = 2
      dDBPatient = None
    elif dDBPatient == "ErrorMultipleResults":
      sProcessType = "Erro: Multiplos utentes na base de dados."
      uProcessType = 3
      dDBPatient = None
    else:
      dUpdate = CheckPatientUpdate(dDBPatient, sRecord, sStateID1, sGender, sStatus, sStatusDateTime, sDepartment, sAdmittanceDate)
      if len(dUpdate) > 0:
        UpdateDBPatient(dDBPatient['id'], dUpdate)
        sProcessType = "Estado: Registo de utente atualizado."
        uProcessType = 6
      else:
        sProcessType = "Estado: Registo de utente nao modificado."
        uProcessType = 8
  else:
    dDBPatient = InsertDBPatient(sRecord, sName, sBirthday, sGender, sStateID1, sStatus, sStatusDateTime, sDepartment, sAdmittanceDate, sFloor, sRoom, sBed)
    if not dDBPatient or dDBPatient == "ErrorIncorrectData":
      sProcessType = "Erro: Dados de utente incorretos ou inconsistentes com os da base de dados."
      uProcessType = 2
      dDBPatient = None
    elif dDBPatient == "ErrorMultipleResults":
      sProcessType = "Erro: Multiplos utentes na base de dados."
      uProcessType = 3
      dDBPatient = None
    else:
      sProcessType = "Estado: Novo utente introduzido."
      uProcessType = 9
  if uProcessType == 2:
    uLogType = 1
  if uProcessType == 3:
    uLogType = 1
  if uProcessType == 4:
    uLogType = 1
  if uProcessType == 5:
    uLogType = 1
  if uProcessType == 6:
    uLogType = 2
  if uProcessType == 7:
    uLogType = 2
  if uProcessType == 8:
    uLogType = None
  if uProcessType == 9:
    uLogType = 3
  Log(uLogType, "  # Dados do utente do teste processado -")
  if dDBPatient:
    Log(uLogType, "    - ID na base de dados: {}".format(dDBPatient['id']))
    Log(uLogType, "    - Sexo: {}".format(dDBPatient['gender']))
    Log(uLogType, "    - Processo: {}".format(dDBPatient['record']))
    Log(uLogType, "    - Numero SNS: {}".format(dDBPatient['state_id_1']))
    Log(uLogType, "    - Codigo de servico: {}".format(dDBPatient['department']))
    Log(uLogType, "    - Data de admissao: {}".format(dDBPatient['admittance_date']))
  Log(uLogType, "    - " + sProcessType)
  return dDBPatient

def ProcessTest(dRow):
  global dTestRelations
  uProcessType = 0
  sProcessType = ""
  sTestCode = GetText(dRow['test_code']).lower()
  sSample = dRow['sample']
  sAccession = dRow['accession'][0:6]
  sName = GetName(dRow['name'])
  sBirthday = GetDate(dRow['birthday'])
  sGender = GetGender(dRow['gender'])
  sRecord = GetRecord(dRow['record'])
  sEpisode = GetText(dRow['episode'])
  sStateID1 = GetStateID1(dRow['state_id_1'], dRow['state_id_1_type'])
  sDepartment = GetText(dRow['department'])
  sDepartmentCode = GetText(dRow['department_code'])
  sDepartmentID = GetDepartment(sDepartment, sRecord)
  sStatus = GetStatus(sDepartment, sRecord)
  sSampleDate = GetDate(dRow['sample_date'])
  sResult = GetResult(dRow[dTestRelations[sTestCode]['result_field']])
  sResultComments = GetText(dRow[dTestRelations[sTestCode]['comments_field']])
  sMethod = dTestRelations[sTestCode]['method']
  sResultDateTime = GetDateTime(dRow['result_datetime'])
  sTecValidator = GetText(dRow['tec_validator'])
  sValidationDatetime = GetDateTime(dRow['validation_datetime'])
  sBioValidator = GetText(dRow['bio_validator'])
  sPrescriber = GetName(dRow['prescriber'])
  sAddress1 = GetText(dRow['address_1'])
  sAddress2 = GetText(dRow['address_2'])
  if not sResult:
    uProcessType = 4
    sProcessType = "Erro: Nao foi possivel traduzir o resultado."
  sPatientID = "ERRO"
  dDBTest = SearchDBTest(sSample)
  if dDBTest and dDBTest == "ErrorMultipleResults":
    sProcessType = "Erro: Multiplos testes na base de dados com o mesmo identificador."
    uProcessType = 3
    dDBTest = None
  if uProcessType == 0:
    dDBPatient = ProcessPatient(sRecord, sName, sBirthday, sGender, sStateID1, sStatus, sSampleDate + " 00:00:00", sDepartmentID, sSampleDate, "NULL", "NULL", "NULL")
  if dDBTest and not isinstance(dDBTest, str) and sPatientID == "ERRO":
    sPatientID = dDBTest['patient_id']
  if uProcessType == 0 and not dDBPatient:
    sProcessType = "Erro: Nao foi possivel introduzir ou encontrar o utente."
    uProcessType = 2
  if dDBTest and uProcessType == 0:
    if UpdateDBTest(
        dDBTest,
        sSample,
        sAccession,
        sName,
        sBirthday,
        sGender,
        sRecord,
        sEpisode,
        sStateID1,
        dDBPatient['id'],
        sDepartment,
        sDepartmentCode,
        sDepartmentID,
        "sarscov2",
        sMethod,
        sSampleDate,
        sResult,
        sResultComments,
        sResultDateTime,
        sTecValidator,
        sValidationDatetime,
        sBioValidator,
        sPrescriber,
        sAddress1,
        sAddress2):
      sProcessType = "Estado: Teste atualizado."
      uProcessType = 6
    else:
      sProcessType = "Estado: Teste nao modificado."
      uProcessType = 8
  elif uProcessType == 0:
    if sPatientID == "ERRO":
      sPatientID = dDBPatient['id']
    InsertDBTest(
      sSample,
      sAccession,
      sName,
      sBirthday,
      sGender,
      sRecord, 
      sEpisode, 
      sStateID1, 
      dDBPatient['id'], 
      sDepartment, 
      sDepartmentCode, 
      sDepartmentID,
      "sarscov2",
      sMethod,
      sSampleDate,
      sResult,
      sResultComments,
      sResultDateTime,
      sTecValidator,
      sValidationDatetime,
      sBioValidator,
      sPrescriber,
      sAddress1,
      sAddress2
      )
    sProcessType = "Estado: Teste introduzido."
    uProcessType = 9
  uLogType = 0
  if uProcessType == 2:
    uLogType = 1
  if uProcessType == 3:
    uLogType = 1
  if uProcessType == 4:
    uLogType = 1
  if uProcessType == 5:
    uLogType = 1
  if uProcessType == 6:
    uLogType = 2
  if uProcessType == 7:
    uLogType = 2
  if uProcessType == 8:
    uLogType = None
  if uProcessType == 9:
    uLogType = 3
  Log(uLogType, "  # Dados do teste processado -")
  Log(uLogType, "    - Nome: {}".format(sName))
  Log(uLogType, "    - Data de nascimento: {}".format(sBirthday))
  Log(uLogType, "    - Sexo: {} ({})".format(dRow['gender'], sGender))
  Log(uLogType, "    - ID de base de dados do utente: {}".format(sPatientID))
  Log(uLogType, "    - Processo: {}".format(sRecord))
  Log(uLogType, "    - Numero SNS: {}".format(sStateID1))
  Log(uLogType, "    - Servico: {} ({})".format(sDepartment, sDepartmentID))
  Log(uLogType, "    - Amostra: {}".format(sSample))
  Log(uLogType, "    - Data da amostra: {}".format(sSampleDate))
  Log(uLogType, "    - Prescritor: {}".format(sPrescriber))
  Log(uLogType, "    - Resultado: {} ({})".format(dRow[dTestRelations[sTestCode]['result_field']], sResult))
  Log(uLogType, "    - Data e hora de resultado: {}".format(sResultDateTime))
  Log(uLogType, "    - " + sProcessType)
  return

def ProcessTest_2(
  sSample,
  sAccession,
  sName,
  sBirthday,
  sGender,
  sRecord,
  sEpisode,
  sDepartment,
  sSampleDate,
  sPrescriber,
  sMethod,
  sAddress1,
  sAddress2):
  uProcessType = 0
  sProcessType = ""
  sStateID1 = "NULL"
  sDepartmentCode = "NULL"
  sDepartmentID = GetDepartment(sDepartment, sRecord)
  sStatus = GetStatus(sDepartment, sRecord)
  sResult = "waiting"
  sResultComments = "NULL"
  sResultDateTime = "NULL"
  sTecValidator = "NULL"
  sValidationDatetime = "NULL"
  sBioValidator = "NULL"
  sPatientID = "ERRO"
  dDBTest = SearchDBTest(sSample)
  if dDBTest and dDBTest == "ErrorMultipleResults":
    sProcessType = "Erro: Multiplos utentes na base de dados."
    uProcessType = 3
    dDBTest = None
  if uProcessType == 0:
    dDBPatient = ProcessPatient(sRecord, sName, sBirthday, sGender, sStateID1, sStatus, sSampleDate + " 00:00:00", sDepartmentID, sSampleDate, "NULL", "NULL", "NULL")
  if dDBTest and not isinstance(dDBTest, str) and sPatientID == "ERRO":
    sPatientID = dDBTest['patient_id']
  if uProcessType == 0 and not dDBPatient:
    sProcessType = "Erro: Nao foi possivel introduzir ou encontrar o utente."
    uProcessType = 2
  if dDBTest and uProcessType == 0:
    sProcessType = "Estado: Teste ja introduzido. Nao foram efetuadas modificacoes."
    uProcessType = 10
  elif uProcessType == 0:
    if sPatientID == "ERRO":
      sPatientID = dDBPatient['id']
    InsertDBTest(
      sSample,
      sAccession,
      sName,
      sBirthday,
      sGender,
      sRecord,
      sEpisode,
      sStateID1,
      dDBPatient['id'],
      sDepartment,
      sDepartmentCode,
      sDepartmentID,
      "sarscov2",
      sMethod,
      sSampleDate,
      sResult,
      sResultComments,
      sResultDateTime,
      sTecValidator,
      sValidationDatetime,
      sBioValidator,
      sPrescriber,
      sAddress1,
      sAddress2
      )
    sProcessType = "Estado: Teste introduzido."
    uProcessType = 9
  uLogType = 0
  if uProcessType == 2:
    uLogType = 1
  if uProcessType == 3:
    uLogType = 1
  if uProcessType == 4:
    uLogType = 1
  if uProcessType == 5:
    uLogType = 1
  if uProcessType == 6:
    uLogType = 2
  if uProcessType == 7:
    uLogType = 2
  if uProcessType == 8:
    uLogType = None
  if uProcessType == 9:
    uLogType = 3
  if uProcessType == 10:
    uLogType = None
  Log(uLogType, "  # A dados do teste processado -")
  Log(uLogType, "    - Nome: {}".format(sName))
  Log(uLogType, "    - Data de nascimento: {}".format(sBirthday))
  Log(uLogType, "    - Sexo: {}".format(sGender))
  Log(uLogType, "    - ID de base de dados do utente: {}".format(sPatientID))
  Log(uLogType, "    - Processo: {}".format(sRecord))
  Log(uLogType, "    - Servico: {} ({})".format(sDepartment, sDepartmentID))
  Log(uLogType, "    - Amostra: {}".format(sSample))
  Log(uLogType, "    - Data da amostra: {}".format(sSampleDate))
  Log(uLogType, "    - " + sProcessType)
  return

def ProcessGroupData(dRow):
  global sEmailRE, lsSocialInstitutionClasses
  sSample = dRow['sample']
  sDate = GetDate(dRow['sample_date'])
  sName = GetName(dRow['name'])
  sBirthday = GetDate(dRow['birthday'])
  sStateID1 = GetStateID1(dRow['state_id_1'], dRow['state_id_1_type'])
  sGroup = CleanStrGroup(dRow['group'])
  sContextClass = dRow['context_class']
  sGroupClass = SearchGroupClass(sContextClass)
  sGroupPatientClass1 = dRow['group_patient_class_1']
  sGroupPatientClass2 = dRow['group_patient_class_2']
  sPatientClass = SearchGroupPatientClass(sGroupPatientClass1)
  if not sPatientClass:
    sPatientClass = SearchGroupPatientClass(sGroupPatientClass2)
  bSocialInstitution = False
  if sGroupClass and sGroupClass in lsSocialInstitutionClasses:
    bSocialInstitution = True
  sDepartment = GetText(dRow['group_department'])
  sLocation = GetText(dRow['group_location'])
  if sGroup and not sGroupClass:
    raise Error("Categoria de grupo nao definida ou incorreta na amostra {}.".format(sSample))
  if sGroup and not sPatientClass:
    raise Error("Categoria de utente nao definida ou incorreta na amostra {}.".format(sSample))
  if sGroup and sGroupClass and sPatientClass:
    DBInsertGroup(
      sName,
      sBirthday,
      sSample,
      sStateID1,
      sGroup, 
      sPatientClass,
      sDepartment,
      None,
      sLocation,
      bSocialInstitution,
      sGroupClass,
      sDate,
      "lis"
      )
    Log(2, "  # Dados de grupo do teste processado -")
    Log(2, "    - Nome: {}".format(sName))
    Log(2, "    - Data de nascimento: {}".format(sBirthday))
    Log(2, "    - Numero SNS: {}".format(sStateID1))
    Log(2, "    - Instituicao ou grupo: {}".format(sGroup))
    if sDepartment:
      Log(2, "    - Departamento: {}".format(sDepartment))
    if sLocation:
      Log(2, "    - Localizacao: {}".format(sLocation))
    Log(2, "    - Data: {}".format(sDate))

def LoadWorklistFromXLS():
  global sSourceWorklist, sSampleRE
  oWB = xlrd.open_workbook(sSourceFile)
  oSheet = oWB.sheet_by_index(0)
  ldCol1Values = oSheet.col_values(3)
  ldCol1Types = oSheet.col_types(3)
  uCurrRow = 0
  lsKeys = ["sample", "name", "birthday", "gender", "record", "department", "sample_date"]
  ldTable = list()
  for uType in ldCol1Types:
    if uType == 2:
      sValue = str(int(ldCol1Values[uCurrRow]))
      if re.match(sSampleRE, sValue):
        dNew = dict()
        dNew['sample'] = sValue
        dNew['gender'] = GetGender(oSheet.cell_value(uCurrRow, 5).strip())
        if oSheet.cell_type(uCurrRow, 7) == 2 and oSheet.cell_value(uCurrRow, 7) > 0:
          dNew['record'] = str(int(oSheet.cell_value(uCurrRow, 7)))
        else:
          dNew['record'] = "NULL"
        dNew['department'] = oSheet.cell_value(uCurrRow, 9).strip()
        dNew['name'] = GetName(oSheet.cell_value(uCurrRow + 1, 2).strip())
        dNew['birthday'] = GetDate(oSheet.cell_value(uCurrRow + 2, 2))
        dNew['sample_date'] = GetDate(oSheet.cell_value(uCurrRow + 2, 4))
        ldTable.append(dNew)
    uCurrRow = uCurrRow + 1
  return (lsKeys, ldTable)

def XLSXGetString(sValue):
  sRes = str(sValue).strip() if sValue else None
  if sRes:
    sRes = re.sub("\\s+", " ", sRes)
  if sRes == "":
    sRes = None
  return sRes

def XLSXGetStringDate(oValue):
  global sDateFormat
  sRes = oValue.date().strftime(sDateFormat) if oValue else datetime.date.today().strftime(sDateFormat)
  return sRes

def XLSXGetStringTime(oValue):
  global sHourFormat
  sRes = oValue.strftime(sHourFormat) if oValue else datetime.datetime.now().strftime(sHourFormat)
  return sRes

def DBInsertGroup(sName, sBirthday, sSample, sStateID, sGroup, sClass, sDepartment, sCategory, sLocation, bSocialInstitution, sInstitutionClass, sDate, sSource):
  global sUserID
  try:
    if sSample:
      sQuery = """
        SELECT
          B.source AS Source,
          B.status AS Status,
          B."group" AS Title,
          A.id AS ID,
          B.date AS Date
        FROM
        (
          SELECT
            MAX(id) AS id
          FROM groups
          WHERE 
            sample = {}
        ) AS A
        LEFT JOIN groups AS B
          ON B.id = A.id
        """.format(GetDBString(sSample))
    else:
      sQuery = """
        SELECT
          B.source AS Source,
          B.status AS Status,
          B."group" AS Title,
          A.id AS ID,
          B.date AS Date
        FROM
        (
          SELECT
            MAX(id) AS id
          FROM groups
          WHERE 
            (
              (
                state_id = {}
                AND
                sample IS NULL
              )
              OR
              (
                (
                  name LIKE CONCAT('%', REPLACE({}, ' ', '%'), '%')
                  OR
                  {} LIKE CONCAT('%', REPLACE(name, ' ', '%'), '%')
                )
                AND
                birthday = {}
                AND
                sample IS NULL
              )
            )
            AND
            date = {}
        ) AS A
        LEFT JOIN groups AS B
          ON B.id = A.id
        """.format(
          GetDBString(sStateID),
          GetDBString(sName),
          GetDBString(sName),
          GetDBString(sBirthday),
          GetDBString(sDate)
        )
    dGroups = RunMySql(sQuery, True)
    bFound = False
    bUpdate = True
    if len(dGroups) > 0 and dGroups[0]['ID'] and dGroups[0]['ID'].lower() not in [None, "null", ""]:
      bFound = True
    if \
      bFound \
      and sSource == "lis" \
      and (
        dGroups[0]['Source'] == "xlsx"
        or dGroups[0]['Source'] == "transmission"
      ):
      bUpdate = False
    uSocialInstitution = 0
    if bSocialInstitution:
      uSocialInstitution = 1
    if not bUpdate:
      return
    if bFound:
      sID = dGroups[0]['ID']
      sQuery = """
        UPDATE groups
        SET
          name = {},
          birthday = {},
          state_id = {},
          "group" = {},
          class = {},
          department = {},
          category = {},
          location = {},
          social_institution = {},
          institution_class = {},
          date = {},
          source = {},
          status = {},
          "datetime" = NOW(),
          user = {}
        WHERE
          id = {}
      """.format(
        GetDBString(sName),
        GetDBString(sBirthday),
        GetDBString(sStateID),
        GetDBString(sGroup),
        GetDBString(sClass),
        GetDBString(sDepartment),
        GetDBString(sCategory),
        GetDBString(sLocation),
        str(uSocialInstitution),
        GetDBString(sInstitutionClass),
        GetDBString(sDate) if sDate and sDate.upper() != "NULL" and sDate != "" else "NOW()",
        GetDBString(sSource),
        "1",
        sUserID,
        GetDBString(sID)
      )
    else:
      sQuery = """
        INSERT INTO groups VALUES (
          NULL, 
          {}, 
          {},
          {},
          {}, 
          {}, 
          {}, 
          {},
          {},
          {},
          {},
          {},
          {},
          {},
          1,
          NOW(),
          {}
        )
        """.format(
          GetDBString(sName),
          GetDBString(sBirthday),
          GetDBString(sSample),
          GetDBString(sStateID),
          GetDBString(sGroup),
          GetDBString(sClass),
          GetDBString(sDepartment),
          GetDBString(sCategory),
          GetDBString(sLocation),
          str(uSocialInstitution),
          GetDBString(sInstitutionClass),
          GetDBString(sDate) if sDate and sDate.upper() != "NULL" and sDate != "" else "NOW()",
          GetDBString(sSource),
          sUserID
      )
    RunMySql(sQuery, False)
    CommitMySql()
  except Exception as dError:
    Log(1, "")
    Log(1, "Erro: Ocorreu um erro no acesso a tabela 'groups' da base de dados.")
    Log(1, "Mensagem de erro:\n" + str(dError))
    Log(1, "")
    raise dError

def DBInsertEmail(sName, sBirthday, sStateID, sEmail, sClass, sDate):
  global sUserID
  try:
    sQuery = """
      SELECT
        id
      FROM emails
      WHERE 
        (
          state_id = {}
          OR
          (
            (
              name LIKE CONCAT('%', REPLACE({}, ' ', '%'), '%')
              OR
              {} LIKE CONCAT('%', REPLACE(name, ' ', '%'), '%')
            )
            AND
            birthday = {}
          )
        )
        AND
        email = {}
        AND
        date = {}
      """.format(
      GetDBString(sStateID),
      GetDBString(sName),
      GetDBString(sName),
      GetDBString(sBirthday),
      GetDBString(sEmail),
      GetDBString(sDate)
    )
    dEmails = RunMySql(sQuery, True)
    bFound = False
    if len(dEmails) > 0:
      bFound = True
    if bFound:
      sID = dEmails[0]['id']
      sQuery = """
        UPDATE emails
        SET
          name = {},
          birthday = {},
          state_id = {},
          email = {},
          class = {},
          "datetime" = NOW(),
          user = {}
        WHERE
          id = {}
      """.format(
        GetDBString(sName),
        GetDBString(sBirthday),
        GetDBString(sStateID),
        GetDBString(sEmail),
        GetDBString(sClass),
        sUserID,
        GetDBString(sID)
      )
    else:
      sQuery = """
        INSERT INTO emails VALUES (
          NULL, 
          {}, 
          {}, 
          {}, 
          {}, 
          {},
          {},
          NOW(),
          {}
        )
        """.format(
          GetDBString(sName),
          GetDBString(sBirthday),
          GetDBString(sStateID),
          GetDBString(sEmail),
          GetDBString(sClass),
          GetDBString(sDate) if sDate and sDate.upper() != "NULL" and sDate != "" else "NOW()",
          sUserID
      )
    RunMySql(sQuery, False)
    CommitMySql()
  except Exception as dError:
    Log(1, "")
    Log(1, "Erro: Ocorreu um erro no acesso a tabela 'emails' da base de dados.")
    Log(1, "Mensagem de erro:\n" + str(dError))
    Log(1, "")
    raise dError

def DBInsertTransmission(sSample, sClass, sDescription, sRecipientClass, sRecipientDescription, sStatus):
  global sUserID
  try:
    sQuery = """
      SELECT
        id
      FROM transmissions
      WHERE 
        sample_id = {}
        AND description = {}
      """.format(
      GetDBString(sSample),
      GetDBString(sDescription)
    )
    dTransmissions = RunMySql(sQuery, True)
    bFound = False
    if len(dTransmissions) > 0:
      bFound = True
    if bFound:
      sID = dTransmissions[0]['id']
      sQuery = "INSERT audit_transmissions SELECT * FROM transmissions WHERE id={}".format(sID)
      RunMySql(sQuery, False)
      sQuery = """
        UPDATE transmissions
        SET
          class = {},
          recipient_description = {},
          mod_user = {},
          mod_datetime = NOW()
        WHERE
          id = {}
      """.format(
        GetDBString(sClass),
        GetDBString(sRecipientDescription),
        sUserID,
        GetDBString(sID)
      )
    else:
      sQuery = """
        INSERT INTO transmissions VALUES (
          NULL, 
          {}, 
          {}, 
          {}, 
          {}, 
          {},
          NOW(),
          {},
          {},
          NOW(),
          {},
          NOW()
        )
        """.format(
          GetDBString(sSample),
          GetDBString(sClass),
          GetDBString(sDescription),
          GetDBString(sRecipientClass),
          GetDBString(sRecipientDescription),
          GetDBString(sStatus),
          sUserID,
          sUserID
      )
    RunMySql(sQuery, False)
    CommitMySql()
  except Exception as dError:
    Log(1, "")
    Log(1, "Erro: Ocorreu um erro no acesso a tabela 'transmissions' da base de dados.")
    Log(1, "Mensagem de erro:\n" + str(dError))
    Log(1, "")
    raise dError

def LoadWorkSheetFromXLSX(sFile):
  global lsSocialInstitutionClasses, sEmailRE
  Log(0, "")
  Log(0, "A processar o ficheiro \"{}\"...".format(os.path.basename(sFile)))
  oWB = load_workbook(sFile)
  oWS = oWB["Dados"]
  uRow = 2
  try:
    while True:
      try:
        sName = XLSXGetString(oWS.cell(row = uRow, column = 2).value)
        sBirthday = XLSXGetStringDate(oWS.cell(row = uRow, column = 3).value)
        if not sName or not sBirthday:
          break
        sName = CleanStrName(sName)
        sStateID = XLSXGetString(oWS.cell(row = uRow, column = 4).value)
        sGroup = CleanStrGroup(XLSXGetString(oWS.cell(row = uRow, column = 5).value))
        sGroupClass = SearchGroupClass(XLSXGetString(oWS.cell(row = uRow, column = 6).value))
        if not sGroupClass and sGroup:
          raise Error("Categoria do grupo nao definido ou incorreto.")
        bSocialInstitution = False
        if sGroupClass and sGroupClass in lsSocialInstitutionClasses:
          bSocialInstitution = True
        sClass = SearchGroupPatientClass(XLSXGetString(oWS.cell(row = uRow, column = 7).value))
        if sGroup and sGroupClass and not sClass:
          raise Error("Categoria do utente nao definido ou incorreto.")
        sDepartment = XLSXGetString(oWS.cell(row = uRow, column = 8).value)
        sLocation = XLSXGetString(oWS.cell(row = uRow, column = 9).value)
        sCategory = XLSXGetString(oWS.cell(row = uRow, column = 10).value)
        sDate = XLSXGetStringDate(oWS.cell(row = uRow, column = 12).value)
        sHour = XLSXGetStringTime(oWS.cell(row = uRow, column = 13).value)
        sEmailPatient = CleanStrEmail(XLSXGetString(oWS.cell(row = uRow, column = 14).value))
        sEmailClinician = CleanStrEmail(XLSXGetString(oWS.cell(row = uRow, column = 15).value))
        sEmailManager = CleanStrEmail(XLSXGetString(oWS.cell(row = uRow, column = 16).value))
        sPhonePatient = XLSXGetString(oWS.cell(row = uRow, column = 17).value)
        sPhoneClinician = XLSXGetString(oWS.cell(row = uRow, column = 18).value)
        sPhoneManager = XLSXGetString(oWS.cell(row = uRow, column = 19).value)
      except Exception as dError:
        Log(1, "")
        Log(1, "Erro: Ocorreu um erro no processamento de \"{}\" na linha \"{}\". Verifique o ficheiro.".format(os.path.basename(sFile), str(uRow)))
        Log(1, "Mensagem de erro:\n" + str(dError))
        Log(1, "")
        raise
      print("")
      print("A processar {} - {}...".format(str(uRow - 1), sName))
      print("- Data de nascimento: {}".format(sBirthday))
      print("- Numero SNS: {}".format(sStateID))
      if sGroup:
        print("- Instituicao ou grupo: {}".format(sGroup))
        if sDepartment:
          print("- Departamento: {}".format(sDepartment))
        if sLocation:
          print("- Localizacao: {}".format(sLocation))
        if sCategory:
          print("- Categoria: {}".format(sCategory))
      print("- Data: {}".format(sDate))
      if sGroup:
        DBInsertGroup(sName, sBirthday, None, sStateID, sGroup, sClass, sDepartment, sCategory, sLocation, bSocialInstitution, sGroupClass, sDate, "xlsx")
      uRow = uRow + 1
      if sEmailPatient:
        lsEmails = re.split(r"\s+|,+|;+|\|", sEmailPatient)
        for sCurrEmail in lsEmails:
          sEmail = CleanStrEmail(sCurrEmail)
          if re.match(sEmailRE, sEmail):
            DBInsertEmail(sName, sBirthday, sStateID, sEmail, "patient", sDate)
      if sEmailClinician:
        lsEmails = re.split(r"\s+|,+|;+|\|", sEmailClinician)
        for sCurrEmail in lsEmails:
          sEmail = CleanStrEmail(sCurrEmail)
          if re.match(sEmailRE, sEmail):
            DBInsertEmail(sName, sBirthday, sStateID, sEmail, "clinician", sDate)
      if sEmailManager:
        lsEmails = re.split(r"\s+|,+|;+|\|", sEmailManager)
        for sCurrEmail in lsEmails:
          sEmail = CleanStrEmail(sCurrEmail)
          if re.match(sEmailRE, sEmail):
            DBInsertEmail(sName, sBirthday, sStateID, sEmail, "manager", sDate)
  except:
    raise

def LoadEmailsFromCSV(sFile):
  global \
    uEmailsFileStartLine,\
    uEmailsFileTrimLines,\
    ltEmailsSelectors
  sEmailRE = r"^[A-Za-z0-9\.\+_-]+@[A-Za-z0-9\._-]+\.[a-zA-Z]*$"
  Log(0, "")
  Log(0, "A processar dados de email do ficheiro \"{}\"...".format(os.path.basename(sFile)))
  lSourceTable = listools.GetTable(sFile, uEmailsFileStartLine, uEmailsFileTrimLines)
  lEmailsTable = listools.SelectTable(lSourceTable, ltEmailsSelectors)
  LockTablesMySql(["transmissions", "audit_transmissions"])
  StartTransactionMySql()
  for dRow in lEmailsTable[1]:
    sSample = dRow['sample']
    sStatus = "pending"
    if dRow['sent'] and dRow['sent'].strip().lower() == "v":
      sStatus = "ok"
    sEmail = CleanStrEmail(dRow['email_unknown'])
    if sEmail and sEmail != "":
      lsEmails = re.split(r"\s+|,+|;+|\|", sEmail)
      for sCurrEmail in lsEmails:
        sNewEmail = sCurrEmail.strip()
        if re.match(sEmailRE, sNewEmail):
          DBInsertTransmission(sSample, "email", sNewEmail, "unknown", "void", sStatus)
    sEmail = CleanStrEmail(dRow['email_clinician'])
    if sEmail and sEmail != "":
      lsEmails = re.split(r"\s+|,+|;+|\|", sEmail)
      for sCurrEmail in lsEmails:
        sNewEmail = sCurrEmail.strip()
        if re.match(sEmailRE, sNewEmail):
          DBInsertTransmission(sSample, "email", sNewEmail, "clinician", "void", sStatus)
    sEmail = CleanStrEmail(dRow['email_patient'])
    if sEmail and sEmail != "":
      lsEmails = re.split(r"\s+|,+|;+|\|", sEmail)
      for sCurrEmail in lsEmails:
        sNewEmail = sCurrEmail.strip()
        if re.match(sEmailRE, sNewEmail):
          DBInsertTransmission(sSample, "email", sNewEmail, "patient", "void", sStatus)
    sEmail = CleanStrEmail(dRow['email_manager'])
    if sEmail and sEmail != "":
      lsEmails = re.split(r"\s+|,+|;+|\|", sEmail)
      for sCurrEmail in lsEmails:
        sNewEmail = sCurrEmail.strip()
        if re.match(sEmailRE, sNewEmail):
          DBInsertTransmission(sSample, "email", sNewEmail, "manager", "void", sStatus)
  CommitMySql()
  UnlockTablesMySql()
  del lSourceTable
  del lEmailsTable

# ------------
# Main routine
# ------------

warnings.simplefilter("ignore")

LoadConfigXML()

print("")
Log(0, "#INSTITUTION# | #DEPARTMENT#")
Log(0, "Ferramenta de atualizacao de resultados do LIS para a plataforma COVID-DX-DB")
Log(0, "----------------------------------------------------------------------------")
Log(0, "")
Log(0, "Data e hora: " + datetime.datetime.now().strftime(sTimestampFormat))

try:
  MySqlCursorStart()
  lsFiles = glob.glob("{}\\*.xlsx".format(sSourceFolder))
  for sFile in lsFiles:
    StartTransactionMySql()
    LoadWorkSheetFromXLSX(sFile)
    CommitMySql()
  sFile = "{}\\emails.tsv".format(sSourceFolder)
  if os.path.isfile(sFile):
    LoadEmailsFromCSV(sFile)
  if os.path.isfile(sSourceRes):
    dSourceResults = listools.GetSelectTable(sSourceRes, uResSourceStartLine, uResSourceTrimLines, sSelResultsTable)
    LockTablesMySql(["tests", "audit_tests", "patients", "audit_patients", "groups", "groups AS B", "transmissions", "audit_transmissions"])
    StartTransactionMySql()
    for dRow in dSourceResults[1]:
      sTestCode = dRow['test_code'].lower()
      if sTestCode not in lsTests:
        continue
      if not re.match(sSampleMatch, dRow['sample']):
        Log(1, "")
        Log(1, "Erro: Ficheiros de resultados com dados invalidos na amostra {}.".format(dRow['sample']))
        CommitMySql()
        raise
      Log(0, "")
      Log(0, "A processar resultados da amostra {}...".format(dRow['sample']))
      ProcessTest(dRow)
      Log(0, "A processar dados de grupo da amostra {}...".format(dRow['sample']))
      ProcessGroupData(dRow)
    CommitMySql()
    UnlockTablesMySql()
    LoadEmailsFromCSV(sSourceRes)
  if os.path.isfile(sSourceWorklist):
    dSourceWorklist = listools.GetSelectTable(sSourceWorklist, uResSourceStartLine, uResSourceTrimLines, sSelWorklistTable)
    LockTablesMySql(["tests", "audit_tests", "patients", "audit_patients", "groups", "groups AS B", "transmissions", "audit_transmissions"])
    StartTransactionMySql()
    for dRow in dSourceWorklist[1]:
      sTestCode = dRow['test_code'].lower()
      sResult = GetResult(dRow[dTestRelations[sTestCode]['result_field']])
      if sTestCode in lsTests and not sResult:
        sMethod = dTestRelations[sTestCode]['method']
        Log(0, "")
        Log(0, "A processar dados da amostra {} com resultado pendente...".format(dRow['sample']))
        ProcessTest_2(
          dRow['sample'],
          dRow['accession'][0:6],
          GetName(dRow['name']),
         GetDate(dRow['birthday']),
          GetGender(dRow['gender']),
          GetRecord(dRow['record']),
          GetText(dRow['episode']),
          dRow['department'].strip(),
          GetDate(dRow['sample_date']),
          GetName(dRow['prescriber']),
          sMethod,
          GetText(dRow['address_1']),
          GetText(dRow['address_2']))
        Log(0, "A processar dados de grupo da amostra {}...".format(dRow['sample']))
        ProcessGroupData(dRow)
    CommitMySql()
    UnlockTablesMySql()
    LoadEmailsFromCSV(sSourceWorklist)
except Exception as dError:
  Log(0, "")
  Log(0, "Erro: Nao foi possivel continuar.")
  Log(0, "Mensagem de erro:\n" + str(dError))
  Log(0, "")
  CommitMySql()
  UnlockTablesMySql()
if dMySqlCursor:
  MySqlCursorClose()
LogSave()

print("")

if uErrorCount > 0:
  print("ATENCAO: Ocorreu pelo menos um erro. Verifique o registo de erros.")
  print("")
